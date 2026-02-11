"""
Views for API endpoints.
"""

from decimal import Decimal

from django.contrib.auth import get_user_model
from django.db import models
from django.db.models import Q
from drf_spectacular.types import OpenApiTypes
from drf_spectacular.utils import OpenApiParameter, OpenApiResponse, extend_schema, extend_schema_view
from rest_framework import status, viewsets
from rest_framework.decorators import action
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework.views import APIView
from rest_framework_simplejwt.tokens import RefreshToken

from .models import Booking, DateRateOverride, DeviceToken, FinancialCategory, FinancialEntry, FolioItem, GroupBooking, Guest, GuestMessage, HousekeepingTask, InspectionTemplate, LostAndFound, MaintenanceRequest, MessageTemplate, MinibarItem, MinibarSale, NightAudit, Notification, RatePlan, Room, RoomInspection, RoomType
from .permissions import IsManager, IsStaff, IsStaffOrManager
from .serializers import (
    BookingListSerializer,
    BookingSerializer,
    BookingStatusUpdateSerializer,
    CheckInSerializer,
    CheckOutSerializer,
    CurrencyConversionSerializer,
    DepositRecordSerializer,
    ExchangeRateSerializer,
    FinancialCategoryListSerializer,
    FinancialCategorySerializer,
    FinancialEntryListSerializer,
    FinancialEntrySerializer,
    FolioItemSerializer,
    GroupBookingCreateSerializer,
    GroupBookingListSerializer,
    GroupBookingSerializer,
    GroupBookingUpdateSerializer,
    GuestListSerializer,
    GuestSearchSerializer,
    GuestSerializer,
    HousekeepingTaskCreateSerializer,
    HousekeepingTaskListSerializer,
    HousekeepingTaskSerializer,
    HousekeepingTaskUpdateSerializer,
    LoginSerializer,
    LostAndFoundClaimSerializer,
    LostAndFoundCreateSerializer,
    LostAndFoundDisposeSerializer,
    LostAndFoundListSerializer,
    LostAndFoundSerializer,
    LostAndFoundUpdateSerializer,
    MaintenanceRequestCreateSerializer,
    MaintenanceRequestListSerializer,
    MaintenanceRequestSerializer,
    MaintenanceRequestUpdateSerializer,
    MinibarItemCreateSerializer,
    MinibarItemListSerializer,
    MinibarItemSerializer,
    MinibarItemUpdateSerializer,
    MinibarSaleBulkCreateSerializer,
    MinibarSaleCreateSerializer,
    MinibarSaleListSerializer,
    MinibarSaleSerializer,
    MinibarSaleUpdateSerializer,
    NightAuditCreateSerializer,
    NightAuditListSerializer,
    NightAuditSerializer,
    OutstandingDepositSerializer,
    PasswordChangeSerializer,
    PaymentListSerializer,
    PaymentSerializer,
    ReceiptDataSerializer,
    ReceiptGenerateSerializer,
    RoomAvailabilitySerializer,
    RoomListSerializer,
    RoomSerializer,
    RoomStatusUpdateSerializer,
    RoomTypeListSerializer,
    RoomTypeSerializer,
    UserProfileSerializer,
    # Phase 3: Room Inspection serializers
    InspectionTemplateSerializer,
    InspectionTemplateListSerializer,
    InspectionTemplateCreateSerializer,
    RoomInspectionSerializer,
    RoomInspectionListSerializer,
    RoomInspectionCreateSerializer,
    RoomInspectionUpdateSerializer,
    RoomInspectionCompleteSerializer,
    RoomInspectionStatisticsSerializer,
    # Phase 4: Report serializers
    OccupancyReportRequestSerializer,
    RevenueReportRequestSerializer,
    KPIReportRequestSerializer,
    ExpenseReportRequestSerializer,
    ChannelPerformanceRequestSerializer,
    GuestDemographicsRequestSerializer,
    ComparativeReportRequestSerializer,
    ExportReportRequestSerializer,
    # RatePlan and DateRateOverride serializers
    RatePlanSerializer,
    RatePlanListSerializer,
    RatePlanCreateSerializer,
    RatePlanUpdateSerializer,
    DateRateOverrideSerializer,
    DateRateOverrideListSerializer,
    DateRateOverrideCreateSerializer,
    DateRateOverrideUpdateSerializer,
    DateRateOverrideBulkCreateSerializer,
    # Phase 5: Notification serializers
    NotificationSerializer,
    NotificationListSerializer,
    DeviceTokenSerializer,
    NotificationPreferencesSerializer,
    # Phase 5.3: Guest Messaging serializers
    MessageTemplateSerializer,
    MessageTemplateListSerializer,
    GuestMessageSerializer,
    GuestMessageListSerializer,
    SendMessageSerializer,
    PreviewMessageSerializer,
)

User = get_user_model()


@extend_schema_view(
    post=extend_schema(
        summary="User login",
        description="Authenticate user and return JWT access and refresh tokens.",
        request=LoginSerializer,
        responses={
            200: OpenApiResponse(
                description="Login successful",
                response={
                    "type": "object",
                    "properties": {
                        "access": {"type": "string", "description": "JWT access token"},
                        "refresh": {"type": "string", "description": "JWT refresh token"},
                        "user": {"$ref": "#/components/schemas/UserProfile"},
                    },
                },
            ),
            400: OpenApiResponse(description="Invalid credentials"),
        },
        tags=["Authentication"],
    )
)
class LoginView(APIView):
    """User login endpoint."""

    permission_classes = []
    authentication_classes = []

    def post(self, request):
        """Login with username and password."""
        serializer = LoginSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        user = serializer.validated_data["user"]

        # Generate JWT tokens
        refresh = RefreshToken.for_user(user)

        # Get user profile data with related HotelUser
        user_with_profile = User.objects.select_related("hotel_profile").get(pk=user.pk)
        user_serializer = UserProfileSerializer(user_with_profile)

        return Response(
            {
                "access": str(refresh.access_token),
                "refresh": str(refresh),
                "user": user_serializer.data,
            },
            status=status.HTTP_200_OK,
        )


@extend_schema_view(
    post=extend_schema(
        summary="Logout",
        description="Blacklist the refresh token to logout the user.",
        request={
            "type": "object",
            "properties": {"refresh": {"type": "string", "description": "JWT refresh token"}},
            "required": ["refresh"],
        },
        responses={
            200: OpenApiResponse(description="Logout successful"),
            400: OpenApiResponse(description="Invalid token"),
        },
        tags=["Authentication"],
    )
)
class LogoutView(APIView):
    """User logout endpoint."""

    permission_classes = [IsAuthenticated]

    def post(self, request):
        """Blacklist the refresh token."""
        try:
            refresh_token = request.data.get("refresh")
            if not refresh_token:
                return Response(
                    {"detail": "Refresh token is required."},
                    status=status.HTTP_400_BAD_REQUEST,
                )

            token = RefreshToken(refresh_token)
            token.blacklist()

            return Response({"detail": "Đăng xuất thành công."}, status=status.HTTP_200_OK)
        except Exception as e:
            return Response({"detail": str(e)}, status=status.HTTP_400_BAD_REQUEST)


@extend_schema_view(
    get=extend_schema(
        summary="Get user profile",
        description="Retrieve the authenticated user's profile information.",
        responses={
            200: UserProfileSerializer,
            401: OpenApiResponse(description="Unauthorized"),
        },
        tags=["Authentication"],
    )
)
class UserProfileView(APIView):
    """User profile endpoint."""

    permission_classes = [IsAuthenticated]

    def get(self, request):
        """Get authenticated user's profile."""
        # Use select_related to avoid N+1 queries
        user = User.objects.select_related("hotel_profile").get(pk=request.user.pk)
        serializer = UserProfileSerializer(user)
        return Response(serializer.data, status=status.HTTP_200_OK)


@extend_schema_view(
    get=extend_schema(
        summary="List staff members",
        description="List all active staff members (users with hotel profiles).",
        responses={
            200: UserProfileSerializer(many=True),
            401: OpenApiResponse(description="Unauthorized"),
        },
        tags=["Authentication"],
    )
)
class StaffListView(APIView):
    """Staff list endpoint for task assignment."""

    permission_classes = [IsAuthenticated]

    def get(self, request):
        """Get list of all active staff members."""
        users = User.objects.filter(
            is_active=True,
            hotel_profile__isnull=False,
        ).select_related("hotel_profile")
        serializer = UserProfileSerializer(users, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)


@extend_schema_view(
    post=extend_schema(
        summary="Change password",
        description="Change the authenticated user's password.",
        request=PasswordChangeSerializer,
        responses={
            200: OpenApiResponse(description="Password changed successfully"),
            400: OpenApiResponse(description="Validation error"),
            401: OpenApiResponse(description="Unauthorized"),
        },
        tags=["Authentication"],
    )
)
class PasswordChangeView(APIView):
    """Password change endpoint."""

    permission_classes = [IsAuthenticated]

    def post(self, request):
        """Change user's password."""
        serializer = PasswordChangeSerializer(data=request.data, context={"request": request})
        serializer.is_valid(raise_exception=True)
        serializer.save()

        return Response(
            {"detail": "Mật khẩu đã được thay đổi thành công."}, status=status.HTTP_200_OK
        )


# ==================== Room Management Views ====================


@extend_schema_view(
    list=extend_schema(
        summary="List room types",
        description="Get list of all room types with room counts.",
        parameters=[
            OpenApiParameter(
                name="is_active",
                type=bool,
                description="Filter by active status",
                required=False,
            ),
        ],
        responses={200: RoomTypeListSerializer(many=True)},
        tags=["Room Management"],
    ),
    retrieve=extend_schema(
        summary="Get room type details",
        description="Get detailed information about a specific room type.",
        responses={200: RoomTypeSerializer},
        tags=["Room Management"],
    ),
    create=extend_schema(
        summary="Create room type",
        description="Create a new room type. Requires manager role.",
        request=RoomTypeSerializer,
        responses={201: RoomTypeSerializer},
        tags=["Room Management"],
    ),
    update=extend_schema(
        summary="Update room type",
        description="Update room type details. Requires manager role.",
        request=RoomTypeSerializer,
        responses={200: RoomTypeSerializer},
        tags=["Room Management"],
    ),
    partial_update=extend_schema(
        summary="Partially update room type",
        description="Partially update room type details. Requires manager role.",
        request=RoomTypeSerializer,
        responses={200: RoomTypeSerializer},
        tags=["Room Management"],
    ),
    destroy=extend_schema(
        summary="Delete room type",
        description="Delete a room type. Requires manager role. Cannot delete if rooms exist.",
        responses={204: None, 400: OpenApiResponse(description="Cannot delete - rooms exist")},
        tags=["Room Management"],
    ),
)
class RoomTypeViewSet(viewsets.ModelViewSet):
    """ViewSet for RoomType CRUD operations."""

    queryset = RoomType.objects.all()
    permission_classes = [IsAuthenticated, IsStaff]

    def get_serializer_class(self):
        """Return appropriate serializer class."""
        if self.action == "list":
            return RoomTypeListSerializer
        return RoomTypeSerializer

    def get_permissions(self):
        """Set permissions based on action."""
        if self.action in ["create", "update", "partial_update", "destroy"]:
            return [IsAuthenticated(), IsManager()]
        return [IsAuthenticated(), IsStaff()]

    def get_queryset(self):
        """Filter queryset based on query parameters."""
        queryset = super().get_queryset()

        # Filter by active status
        is_active = self.request.query_params.get("is_active")
        if is_active is not None:
            is_active_bool = is_active.lower() in ["true", "1", "yes"]
            queryset = queryset.filter(is_active=is_active_bool)

        return queryset.prefetch_related("rooms")

    def destroy(self, request, *args, **kwargs):
        """Delete room type, but prevent if rooms exist."""
        instance = self.get_object()

        # Check if any rooms exist for this type
        if instance.rooms.exists():
            return Response(
                {"detail": "Không thể xóa loại phòng đã có phòng. Hãy xóa hoặc chuyển phòng trước."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        return super().destroy(request, *args, **kwargs)


@extend_schema_view(
    list=extend_schema(
        summary="List rooms",
        description="Get list of all rooms with filtering options.",
        parameters=[
            OpenApiParameter(name="status", type=str, description="Filter by status", required=False),
            OpenApiParameter(
                name="room_type", type=int, description="Filter by room type ID", required=False
            ),
            OpenApiParameter(name="floor", type=int, description="Filter by floor", required=False),
            OpenApiParameter(name="is_active", type=bool, description="Filter by active status", required=False),
            OpenApiParameter(name="search", type=str, description="Search by room number or name", required=False),
        ],
        responses={200: RoomListSerializer(many=True)},
        tags=["Room Management"],
    ),
    retrieve=extend_schema(
        summary="Get room details",
        description="Get detailed information about a specific room.",
        responses={200: RoomSerializer},
        tags=["Room Management"],
    ),
    create=extend_schema(
        summary="Create room",
        description="Create a new room. Requires manager role.",
        request=RoomSerializer,
        responses={201: RoomSerializer},
        tags=["Room Management"],
    ),
    update=extend_schema(
        summary="Update room",
        description="Update room details. Requires manager role.",
        request=RoomSerializer,
        responses={200: RoomSerializer},
        tags=["Room Management"],
    ),
    partial_update=extend_schema(
        summary="Partially update room",
        description="Partially update room details. Requires manager role.",
        request=RoomSerializer,
        responses={200: RoomSerializer},
        tags=["Room Management"],
    ),
    destroy=extend_schema(
        summary="Delete room",
        description="Delete a room. Requires manager role.",
        responses={204: None},
        tags=["Room Management"],
    ),
)
class RoomViewSet(viewsets.ModelViewSet):
    """ViewSet for Room CRUD operations."""

    queryset = Room.objects.all()
    permission_classes = [IsAuthenticated, IsStaff]

    def get_serializer_class(self):
        """Return appropriate serializer class."""
        if self.action == "list":
            return RoomListSerializer
        elif self.action == "update_status":
            return RoomStatusUpdateSerializer
        return RoomSerializer

    def get_permissions(self):
        """Set permissions based on action."""
        if self.action in ["create", "update", "partial_update", "destroy"]:
            return [IsAuthenticated(), IsManager()]
        return [IsAuthenticated(), IsStaff()]

    def get_queryset(self):
        """Filter queryset based on query parameters."""
        queryset = super().get_queryset().select_related("room_type")

        # Filter by status
        status_param = self.request.query_params.get("status")
        if status_param:
            queryset = queryset.filter(status=status_param)

        # Filter by room type
        room_type_param = self.request.query_params.get("room_type")
        if room_type_param:
            queryset = queryset.filter(room_type_id=room_type_param)

        # Filter by floor
        floor_param = self.request.query_params.get("floor")
        if floor_param:
            queryset = queryset.filter(floor=floor_param)

        # Filter by active status
        is_active = self.request.query_params.get("is_active")
        if is_active is not None:
            is_active_bool = is_active.lower() in ["true", "1", "yes"]
            queryset = queryset.filter(is_active=is_active_bool)

        # Search by room number or name
        search_param = self.request.query_params.get("search")
        if search_param:
            queryset = queryset.filter(Q(number__icontains=search_param) | Q(name__icontains=search_param))

        return queryset

    @extend_schema(
        summary="Update room status",
        description="Update the status of a room (e.g., available, occupied, cleaning).",
        request=RoomStatusUpdateSerializer,
        responses={
            200: RoomSerializer,
            400: OpenApiResponse(description="Invalid status transition"),
        },
        tags=["Room Management"],
    )
    @action(detail=True, methods=["post"], url_path="update-status")
    def update_status(self, request, pk=None):
        """Update room status."""
        room = self.get_object()
        serializer = RoomStatusUpdateSerializer(data=request.data, context={"room": room})
        serializer.is_valid(raise_exception=True)

        # Update room status
        room.status = serializer.validated_data["status"]
        if "notes" in serializer.validated_data:
            room.notes = serializer.validated_data["notes"]
        room.save()

        # Return updated room
        return Response(RoomSerializer(room).data, status=status.HTTP_200_OK)

    @extend_schema(
        summary="Check room availability",
        description="Check which rooms are available for a given date range.",
        request=RoomAvailabilitySerializer,
        responses={
            200: OpenApiResponse(
                description="Available rooms",
                response={
                    "type": "object",
                    "properties": {
                        "available_rooms": {"type": "array", "items": {"$ref": "#/components/schemas/RoomList"}},
                        "total_available": {"type": "integer", "description": "Total number of available rooms"},
                        "check_in": {"type": "string", "format": "date"},
                        "check_out": {"type": "string", "format": "date"},
                        "room_type": {"type": "integer", "nullable": True},
                    },
                },
            )
        },
        tags=["Room Management"],
    )
    @action(detail=False, methods=["post"], url_path="check-availability")
    def check_availability(self, request):
        """Check room availability for date range."""
        serializer = RoomAvailabilitySerializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        check_in = serializer.validated_data["check_in"]
        check_out = serializer.validated_data["check_out"]
        room_type = serializer.validated_data.get("room_type")

        # For now, return all available rooms (no booking overlap check yet)
        # Will be enhanced when Booking model endpoints are implemented
        available_rooms = Room.objects.filter(
            is_active=True,
            status=Room.Status.AVAILABLE,
        ).select_related("room_type")

        if room_type:
            available_rooms = available_rooms.filter(room_type=room_type)

        return Response(
            {
                "available_rooms": RoomListSerializer(available_rooms, many=True).data,
                "total_available": available_rooms.count(),
                "check_in": check_in,
                "check_out": check_out,
                "room_type": room_type.id if room_type else None,
            },
            status=status.HTTP_200_OK,
        )


# ==================== Guest Management Views ====================


@extend_schema_view(
    list=extend_schema(
        summary="List guests",
        description="Get list of all guests with filtering options.",
        parameters=[
            OpenApiParameter(
                name="is_vip",
                type=bool,
                description="Filter by VIP status",
                required=False,
            ),
            OpenApiParameter(
                name="nationality",
                type=str,
                description="Filter by nationality",
                required=False,
            ),
            OpenApiParameter(
                name="search",
                type=str,
                description="Search by name, phone, or ID number",
                required=False,
            ),
        ],
        responses={200: GuestListSerializer(many=True)},
        tags=["Guest Management"],
    ),
    retrieve=extend_schema(
        summary="Get guest details",
        description="Get detailed information about a specific guest.",
        responses={200: GuestSerializer},
        tags=["Guest Management"],
    ),
    create=extend_schema(
        summary="Create new guest",
        description="Create a new guest record.",
        request=GuestSerializer,
        responses={201: GuestSerializer},
        tags=["Guest Management"],
    ),
    update=extend_schema(
        summary="Update guest",
        description="Update a guest's information.",
        request=GuestSerializer,
        responses={200: GuestSerializer},
        tags=["Guest Management"],
    ),
    partial_update=extend_schema(
        summary="Partial update guest",
        description="Partially update a guest's information.",
        request=GuestSerializer,
        responses={200: GuestSerializer},
        tags=["Guest Management"],
    ),
    destroy=extend_schema(
        summary="Delete guest",
        description="Delete a guest record. This is a soft delete that deactivates the guest.",
        responses={204: OpenApiResponse(description="Guest deleted successfully")},
        tags=["Guest Management"],
    ),
)
class GuestViewSet(viewsets.ModelViewSet):
    """ViewSet for managing guests."""

    permission_classes = [IsAuthenticated, IsStaff]
    serializer_class = GuestSerializer
    filterset_fields = ["is_vip", "nationality"]
    search_fields = ["full_name", "phone", "id_number"]
    ordering_fields = ["created_at", "full_name", "total_stays"]
    ordering = ["-created_at"]

    def get_queryset(self):
        """Get queryset with optional filtering and optimized annotations."""
        from django.db.models import Count
        
        queryset = Guest.objects.annotate(booking_count=Count("bookings")).all()

        # Filter by VIP status
        is_vip = self.request.query_params.get("is_vip")
        if is_vip is not None:
            queryset = queryset.filter(is_vip=is_vip.lower() == "true")

        # Filter by nationality
        nationality = self.request.query_params.get("nationality")
        if nationality:
            queryset = queryset.filter(nationality=nationality)

        # Search by name, phone, or ID number
        search = self.request.query_params.get("search")
        if search:
            queryset = queryset.filter(
                Q(full_name__icontains=search)
                | Q(phone__icontains=search)
                | Q(id_number__icontains=search)
            )

        return queryset.order_by(*self.ordering)

    def get_serializer_class(self):
        """Return appropriate serializer class."""
        if self.action == "list":
            return GuestListSerializer
        elif self.action == "search":
            return GuestSearchSerializer
        return GuestSerializer

    def get_permissions(self):
        """Set permissions based on action."""
        if self.action in ["create", "update", "partial_update", "destroy"]:
            return [IsAuthenticated(), IsManager()]
        return [IsAuthenticated(), IsStaff()]

    @extend_schema(
        summary="Search guests",
        description="Search for guests by name, phone, or ID number.",
        request=GuestSearchSerializer,
        responses={200: GuestListSerializer(many=True)},
        tags=["Guest Management"],
    )
    @action(detail=False, methods=["post"], url_path="search")
    def search(self, request):
        """Search for guests."""
        serializer = GuestSearchSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        query = serializer.validated_data["query"]
        guests = Guest.objects.filter(
            Q(full_name__icontains=query) | Q(phone__icontains=query) | Q(id_number__icontains=query)
        )

        return Response(
            GuestListSerializer(guests, many=True).data,
            status=status.HTTP_200_OK,
        )

    @extend_schema(
        summary="Get guest booking history",
        description="Get all bookings for a specific guest.",
        responses={
            200: OpenApiResponse(
                description="Guest booking history",
                response={
                    "type": "object",
                    "properties": {
                        "guest": {"$ref": "#/components/schemas/Guest"},
                        "bookings": {"type": "array", "items": {"$ref": "#/components/schemas/BookingList"}},
                        "total_bookings": {"type": "integer"},
                        "total_stays": {"type": "integer"},
                    },
                },
            )
        },
        tags=["Guest Management"],
    )
    @action(detail=True, methods=["get"], url_path="history")
    def history(self, request, pk=None):
        """Get booking history for a guest."""
        guest = self.get_object()
        bookings = Booking.objects.filter(guest=guest).select_related("room", "room__room_type").order_by(
            "-check_in_date"
        )

        return Response(
            {
                "guest": GuestSerializer(guest).data,
                "bookings": BookingListSerializer(bookings, many=True).data,
                "total_bookings": bookings.count(),
                "total_stays": guest.total_stays,
            },
            status=status.HTTP_200_OK,
        )

    @extend_schema(
        summary="Export temporary residence declaration",
        description=(
            "Export guest data for temporary residence declaration to police.\n"
            "Supports two official Vietnamese formats:\n"
            "- **dd10**: Mẫu ĐD10 - Sổ quản lý lưu trú (Vietnamese guests, Nghị định 144/2021)\n"
            "- **na17**: Mẫu NA17 - Phiếu khai báo tạm trú (Foreign guests, Thông tư 04/2015)\n"
            "- **all**: Both formats in one Excel workbook (separate sheets) or combined CSV\n"
            "\nReturns CSV or Excel file with guest information for a date range."
        ),
        parameters=[
            OpenApiParameter(
                name="date_from",
                type=str,
                description="Start date for export (YYYY-MM-DD). Defaults to today.",
                required=False,
            ),
            OpenApiParameter(
                name="date_to",
                type=str,
                description="End date for export (YYYY-MM-DD). Defaults to today.",
                required=False,
            ),
            OpenApiParameter(
                name="export_format",
                type=str,
                description="Export format: 'csv' or 'excel'. Defaults to 'excel'.",
                required=False,
            ),
            OpenApiParameter(
                name="form_type",
                type=str,
                description="Form type: 'dd10' (Vietnamese), 'na17' (Foreign), 'all' (both). Defaults to 'all'.",
                required=False,
            ),
        ],
        responses={
            200: OpenApiResponse(
                description="File download (CSV or Excel)",
            )
        },
        tags=["Guest Management"],
    )
    @action(detail=False, methods=["get"], url_path="declaration-export")
    def declaration_export(self, request):
        """Export temporary residence declaration for police reporting.

        Supports two official Vietnamese forms:
        - ĐD10 (Nghị định 144/2021): For Vietnamese guests → reported to Công an phường
        - NA17 (Thông tư 04/2015): For foreign guests → reported to Phòng Quản lý XNC
        """
        import csv
        import io
        from datetime import date

        from django.http import HttpResponse
        from django.utils import timezone as tz

        # ── Hotel establishment info (used in form headers) ──
        HOTEL_NAME = "Hoàng Lâm Heritage Suites"
        HOTEL_ADDRESS = "123 Đường ABC, Phường XYZ, TP.HCM"
        HOTEL_PHONE = "028 1234 5678"

        # ── Parse query parameters ──
        date_from = request.query_params.get("date_from")
        date_to = request.query_params.get("date_to")
        export_format = request.query_params.get("export_format", "excel").lower()
        form_type = request.query_params.get("form_type", "all").lower()

        today = date.today()
        try:
            date_from = date.fromisoformat(date_from) if date_from else today
            date_to = date.fromisoformat(date_to) if date_to else today
        except ValueError:
            return Response(
                {"detail": "Định dạng ngày không hợp lệ. Sử dụng YYYY-MM-DD."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        if date_from > date_to:
            return Response(
                {"detail": "Ngày bắt đầu không được lớn hơn ngày kết thúc."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        if form_type not in ("dd10", "na17", "all"):
            return Response(
                {"detail": "form_type không hợp lệ. Sử dụng: dd10, na17, all."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        # ── Fetch bookings ──
        bookings = Booking.objects.filter(
            check_in_date__gte=date_from,
            check_in_date__lte=date_to,
            status__in=[Booking.Status.CHECKED_IN, Booking.Status.CHECKED_OUT],
        ).select_related("guest", "room")

        # ── Separate Vietnamese and foreign guests ──
        vn_aliases = {"việt nam", "vietnam", "vn", "viet nam"}
        vietnamese_bookings = []
        foreign_bookings = []
        for booking in bookings:
            nationality = (booking.guest.nationality or "").strip().lower()
            if nationality in vn_aliases or not nationality:
                vietnamese_bookings.append(booking)
            else:
                foreign_bookings.append(booking)

        # ── Helper: format date DD/MM/YYYY ──
        def fmt_date(d):
            return d.strftime("%d/%m/%Y") if d else ""

        # ── Helper: gender display ──
        def fmt_gender(g):
            return "Nam" if g == "male" else ("Nữ" if g == "female" else "")

        # ── Helper: nationality display (Vietnamese name for official forms) ──
        nationality_map = {
            "vietnam": "Việt Nam", "united states": "Mỹ", "usa": "Mỹ",
            "china": "Trung Quốc", "south korea": "Hàn Quốc", "japan": "Nhật Bản",
            "france": "Pháp", "uk": "Anh", "australia": "Úc", "germany": "Đức",
            "russia": "Nga", "thailand": "Thái Lan",
        }

        def fmt_nationality(nat):
            return nationality_map.get((nat or "").strip().lower(), nat or "Việt Nam")

        # ═══════════════════════════════════════════════════════════
        # Build ĐD10 rows (Vietnamese guests - Nghị định 144/2021)
        # ═══════════════════════════════════════════════════════════
        DD10_HEADERS = [
            "STT", "Họ và tên", "Ngày sinh", "Giới tính", "Quốc tịch",
            "Số CMND/CCCD/Hộ chiếu", "Địa chỉ thường trú",
            "Ngày đến", "Ngày đi", "Số phòng", "Ghi chú",
        ]

        def build_dd10_rows(bkings):
            rows = []
            for i, bk in enumerate(bkings, 1):
                g = bk.guest
                rows.append([
                    i,
                    g.full_name,
                    fmt_date(g.date_of_birth),
                    fmt_gender(g.gender),
                    fmt_nationality(g.nationality),
                    g.id_number or "",
                    g.address or "",
                    fmt_date(bk.check_in_date),
                    fmt_date(bk.check_out_date),
                    bk.room.number,
                    bk.notes or "",
                ])
            return rows

        # ═══════════════════════════════════════════════════════════
        # Build NA17 rows (Foreign guests - Thông tư 04/2015/TT-BCA)
        # ═══════════════════════════════════════════════════════════
        NA17_HEADERS = [
            "STT", "Họ tên", "Giới tính", "Ngày tháng năm sinh",
            "Quốc tịch", "Số hộ chiếu", "Loại hộ chiếu",
            "Loại giấy tờ nhập cảnh", "Số giấy tờ", "Thời hạn",
            "Ngày cấp", "Cơ quan cấp",
            "Ngày nhập cảnh", "Cửa khẩu nhập cảnh",
            "Mục đích nhập cảnh",
            "Tạm trú từ ngày", "Tạm trú đến ngày",
            "Số phòng",
        ]

        def build_na17_rows(bkings):
            rows = []
            for i, bk in enumerate(bkings, 1):
                g = bk.guest
                rows.append([
                    i,
                    g.full_name,
                    fmt_gender(g.gender),
                    fmt_date(g.date_of_birth),
                    fmt_nationality(g.nationality),
                    g.id_number or "",
                    g.get_passport_type_display() if g.passport_type else "",
                    g.get_visa_type_display() if g.visa_type else "",
                    g.visa_number or "",
                    fmt_date(g.visa_expiry_date),
                    fmt_date(g.visa_issue_date),
                    g.visa_issuing_authority or "",
                    fmt_date(g.entry_date),
                    g.entry_port or "",
                    g.entry_purpose or "",
                    fmt_date(bk.check_in_date),
                    fmt_date(bk.check_out_date),
                    bk.room.number,
                ])
            return rows

        # ── Mark bookings as declared ──
        now = tz.now()
        booking_ids = [b.id for b in bookings]
        if booking_ids:
            Booking.objects.filter(id__in=booking_ids, declaration_submitted=False).update(
                declaration_submitted=True,
                declaration_submitted_at=now,
            )

        # ═══════════════════════════════════════════════════════════
        # EXCEL EXPORT (recommended - separate sheets per form)
        # ═══════════════════════════════════════════════════════════
        if export_format == "excel":
            try:
                import openpyxl
                from openpyxl.styles import Alignment, Font, PatternFill
            except ImportError:
                return Response(
                    {"detail": "Thư viện openpyxl chưa được cài đặt. Vui lòng sử dụng format=csv."},
                    status=status.HTTP_400_BAD_REQUEST,
                )

            wb = openpyxl.Workbook()
            wb.remove(wb.active)  # Remove default sheet

            header_font = Font(bold=True)
            header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
            title_font = Font(bold=True, size=14)

            def write_sheet(ws, title, headers, rows, establishment_header):
                """Write a properly formatted declaration sheet."""
                # Row 1: Establishment name
                ws.cell(row=1, column=1, value=establishment_header[0]).font = title_font
                ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
                # Row 2: Address
                ws.cell(row=2, column=1, value=f"Địa chỉ: {establishment_header[1]}")
                ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))
                # Row 3: Phone
                ws.cell(row=3, column=1, value=f"Điện thoại: {establishment_header[2]}")
                ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=len(headers))
                # Row 4: Empty
                # Row 5: Title
                ws.cell(row=5, column=1, value=title).font = Font(bold=True, size=12)
                ws.merge_cells(start_row=5, start_column=1, end_row=5, end_column=len(headers))
                ws.cell(row=5, column=1).alignment = Alignment(horizontal="center")
                # Row 6: Date range
                date_range_str = f"Từ ngày {fmt_date(date_from)} đến ngày {fmt_date(date_to)}"
                ws.cell(row=6, column=1, value=date_range_str)
                ws.merge_cells(start_row=6, start_column=1, end_row=6, end_column=len(headers))
                ws.cell(row=6, column=1).alignment = Alignment(horizontal="center")

                # Row 8: Column headers
                for col, h in enumerate(headers, 1):
                    cell = ws.cell(row=8, column=col, value=h)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal="center", wrap_text=True)

                # Data rows starting at row 9
                for row_idx, row_data in enumerate(rows, 9):
                    for col_idx, val in enumerate(row_data, 1):
                        ws.cell(row=row_idx, column=col_idx, value=val)

                # Auto-adjust column widths
                for col_idx in range(1, len(headers) + 1):
                    max_len = len(str(headers[col_idx - 1]))
                    for row_data in rows:
                        if col_idx - 1 < len(row_data):
                            max_len = max(max_len, len(str(row_data[col_idx - 1])))
                    ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = min(max_len + 4, 40)

            establishment_info = [HOTEL_NAME, HOTEL_ADDRESS, HOTEL_PHONE]

            if form_type in ("dd10", "all"):
                ws_dd10 = wb.create_sheet(title="ĐD10 - Khách Việt Nam")
                dd10_rows = build_dd10_rows(vietnamese_bookings)
                write_sheet(
                    ws_dd10,
                    "SỔ QUẢN LÝ LƯU TRÚ (Mẫu ĐD10)",
                    DD10_HEADERS,
                    dd10_rows,
                    establishment_info,
                )

            if form_type in ("na17", "all"):
                ws_na17 = wb.create_sheet(title="NA17 - Khách nước ngoài")
                na17_rows = build_na17_rows(foreign_bookings)
                na17_header_info = [
                    HOTEL_NAME,
                    HOTEL_ADDRESS,
                    HOTEL_PHONE,
                ]
                write_sheet(
                    ws_na17,
                    "PHIẾU KHAI BÁO TẠM TRÚ CHO NGƯỜI NƯỚC NGOÀI (Mẫu NA17)",
                    NA17_HEADERS,
                    na17_rows,
                    na17_header_info,
                )

            # If no sheets created (shouldn't happen but just in case)
            if not wb.sheetnames:
                wb.create_sheet(title="Trống")

            buffer = io.BytesIO()
            wb.save(buffer)
            buffer.seek(0)

            response = HttpResponse(
                buffer.getvalue(),
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
            filename = f"khai_bao_luu_tru_{date_from}_{date_to}.xlsx"
            response["Content-Disposition"] = f'attachment; filename="{filename}"'
            return response

        # ═══════════════════════════════════════════════════════════
        # CSV EXPORT (single file, sections separated by blank lines)
        # ═══════════════════════════════════════════════════════════
        else:
            buffer = io.StringIO()
            writer = csv.writer(buffer)

            def write_csv_section(writer, title, headers, rows, establishment_info):
                writer.writerow([establishment_info[0]])
                writer.writerow([f"Địa chỉ: {establishment_info[1]}"])
                writer.writerow([f"Điện thoại: {establishment_info[2]}"])
                writer.writerow([])
                writer.writerow([title])
                writer.writerow([f"Từ ngày {fmt_date(date_from)} đến ngày {fmt_date(date_to)}"])
                writer.writerow([])
                writer.writerow(headers)
                for row in rows:
                    writer.writerow(row)

            establishment_info = [HOTEL_NAME, HOTEL_ADDRESS, HOTEL_PHONE]

            if form_type in ("dd10", "all"):
                dd10_rows = build_dd10_rows(vietnamese_bookings)
                write_csv_section(
                    writer,
                    "SỔ QUẢN LÝ LƯU TRÚ (Mẫu ĐD10)",
                    DD10_HEADERS,
                    dd10_rows,
                    establishment_info,
                )

            if form_type == "all":
                writer.writerow([])
                writer.writerow(["=" * 80])
                writer.writerow([])

            if form_type in ("na17", "all"):
                na17_rows = build_na17_rows(foreign_bookings)
                write_csv_section(
                    writer,
                    "PHIẾU KHAI BÁO TẠM TRÚ CHO NGƯỜI NƯỚC NGOÀI (Mẫu NA17)",
                    NA17_HEADERS,
                    na17_rows,
                    establishment_info,
                )

            response = HttpResponse(
                buffer.getvalue(),
                content_type="text/csv; charset=utf-8-sig",
            )
            filename = f"khai_bao_luu_tru_{date_from}_{date_to}.csv"
            response["Content-Disposition"] = f'attachment; filename="{filename}"'
            return response


# ==================== Booking Management Views ====================


@extend_schema_view(
    list=extend_schema(
        summary="List bookings",
        description="Get list of all bookings with filtering options.",
        parameters=[
            OpenApiParameter(
                name="status",
                type=str,
                description="Filter by booking status",
                required=False,
            ),
            OpenApiParameter(
                name="source",
                type=str,
                description="Filter by booking source",
                required=False,
            ),
            OpenApiParameter(
                name="room",
                type=int,
                description="Filter by room ID",
                required=False,
            ),
            OpenApiParameter(
                name="guest",
                type=int,
                description="Filter by guest ID",
                required=False,
            ),
            OpenApiParameter(
                name="check_in_from",
                type=str,
                description="Filter bookings with check-in date from this date (YYYY-MM-DD)",
                required=False,
            ),
            OpenApiParameter(
                name="check_in_to",
                type=str,
                description="Filter bookings with check-in date until this date (YYYY-MM-DD)",
                required=False,
            ),
        ],
        responses={200: BookingListSerializer(many=True)},
        tags=["Booking Management"],
    ),
    retrieve=extend_schema(
        summary="Get booking details",
        description="Get detailed information about a specific booking.",
        responses={200: BookingSerializer},
        tags=["Booking Management"],
    ),
    create=extend_schema(
        summary="Create new booking",
        description="Create a new booking. Automatically checks for room availability and overlaps.",
        request=BookingSerializer,
        responses={201: BookingSerializer},
        tags=["Booking Management"],
    ),
    update=extend_schema(
        summary="Update booking",
        description="Update a booking's information. Checks for room availability if dates or room changed.",
        request=BookingSerializer,
        responses={200: BookingSerializer},
        tags=["Booking Management"],
    ),
    partial_update=extend_schema(
        summary="Partial update booking",
        description="Partially update a booking's information.",
        request=BookingSerializer,
        responses={200: BookingSerializer},
        tags=["Booking Management"],
    ),
    destroy=extend_schema(
        summary="Delete booking",
        description="Delete a booking record. This is a hard delete.",
        responses={204: OpenApiResponse(description="Booking deleted successfully")},
        tags=["Booking Management"],
    ),
)
class BookingViewSet(viewsets.ModelViewSet):
    """ViewSet for managing bookings."""

    permission_classes = [IsAuthenticated, IsStaff]
    serializer_class = BookingSerializer
    filterset_fields = ["status", "source", "room", "guest"]
    ordering_fields = ["created_at", "check_in_date", "check_out_date"]
    ordering = ["-created_at"]

    def get_queryset(self):
        """Get queryset with optional filtering."""
        queryset = Booking.objects.select_related("guest", "room", "room__room_type").all()

        # Filter by date range
        check_in_from = self.request.query_params.get("check_in_from")
        if check_in_from:
            queryset = queryset.filter(check_in_date__gte=check_in_from)

        check_in_to = self.request.query_params.get("check_in_to")
        if check_in_to:
            queryset = queryset.filter(check_in_date__lte=check_in_to)

        return queryset.order_by(*self.ordering)

    def get_serializer_class(self):
        """Return appropriate serializer class."""
        if self.action == "list":
            return BookingListSerializer
        elif self.action == "update_status":
            return BookingStatusUpdateSerializer
        elif self.action == "check_in":
            return CheckInSerializer
        elif self.action == "check_out":
            return CheckOutSerializer
        return BookingSerializer

    def perform_create(self, serializer):
        """Create booking and notify staff."""
        booking = serializer.save()

        from .services import PushNotificationService

        PushNotificationService.notify_staff(
            notification_type=Notification.NotificationType.BOOKING_CREATED,
            title=f"Đặt phòng mới: Phòng {booking.room.number}",
            body=f"{booking.guest.full_name} - {booking.check_in_date} → {booking.check_out_date}",
            data={
                "booking_id": str(booking.id),
                "room_number": booking.room.number,
                "action": "booking_created",
            },
            booking=booking,
            exclude_user=self.request.user,
        )

    @extend_schema(
        summary="Update booking status",
        description="Update the status of a booking (e.g., confirm, cancel, no-show).",
        request=BookingStatusUpdateSerializer,
        responses={200: BookingSerializer},
        tags=["Booking Management"],
    )
    @action(detail=True, methods=["post"], url_path="update-status")
    def update_status(self, request, pk=None):
        """Update booking status."""
        from django.db import transaction

        booking = self.get_object()
        old_status = booking.status
        serializer = BookingStatusUpdateSerializer(booking, data=request.data, partial=True)
        serializer.is_valid(raise_exception=True)

        with transaction.atomic():
            serializer.save()
            booking.refresh_from_db()

            # Sync room status based on booking status transition
            room = booking.room
            if booking.status == Booking.Status.CHECKED_IN:
                room.status = Room.Status.OCCUPIED
                room.save()
            elif booking.status == Booking.Status.CHECKED_OUT:
                room.status = Room.Status.CLEANING
                room.save()
            elif booking.status in [Booking.Status.CANCELLED, Booking.Status.NO_SHOW]:
                # Only revert room if it was occupied by this booking
                if old_status == Booking.Status.CHECKED_IN and room.status == Room.Status.OCCUPIED:
                    # Check if another active booking occupies this room
                    other_active = Booking.objects.filter(
                        room=room,
                        status=Booking.Status.CHECKED_IN,
                    ).exclude(pk=booking.pk).exists()
                    if not other_active:
                        room.status = Room.Status.AVAILABLE
                        room.save()

        # Send notification for status changes
        from .services import PushNotificationService

        if booking.status == Booking.Status.CONFIRMED:
            PushNotificationService.notify_staff(
                notification_type=Notification.NotificationType.BOOKING_CONFIRMED,
                title=f"Xác nhận: Phòng {booking.room.number}",
                body=f"{booking.guest.full_name} - {booking.check_in_date} → {booking.check_out_date}",
                data={
                    "booking_id": str(booking.id),
                    "room_number": booking.room.number,
                    "action": "booking_confirmed",
                },
                booking=booking,
                exclude_user=request.user,
            )
        elif booking.status == Booking.Status.CANCELLED:
            PushNotificationService.notify_staff(
                notification_type=Notification.NotificationType.BOOKING_CANCELLED,
                title=f"Hủy đặt phòng: Phòng {booking.room.number}",
                body=f"{booking.guest.full_name} - {booking.check_in_date}",
                data={
                    "booking_id": str(booking.id),
                    "room_number": booking.room.number,
                    "action": "booking_cancelled",
                },
                booking=booking,
                exclude_user=request.user,
            )

        return Response(
            BookingSerializer(booking).data,
            status=status.HTTP_200_OK,
        )

    @extend_schema(
        summary="Check in guest",
        description="Check in a guest for their booking. Updates status to CHECKED_IN.",
        request=CheckInSerializer,
        responses={200: BookingSerializer},
        tags=["Booking Management"],
    )
    @action(detail=True, methods=["post"], url_path="check-in")
    def check_in(self, request, pk=None):
        """Check in a guest."""
        from django.db import transaction
        from django.utils import timezone

        booking = self.get_object()

        if booking.status == Booking.Status.CHECKED_IN:
            return Response(
                {"detail": "Guest is already checked in."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        if booking.status not in [Booking.Status.PENDING, Booking.Status.CONFIRMED]:
            return Response(
                {"detail": f"Cannot check in booking with status {booking.get_status_display()}."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        serializer = CheckInSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        # Use atomic transaction to ensure consistency
        with transaction.atomic():
            booking.status = Booking.Status.CHECKED_IN
            booking.actual_check_in = serializer.validated_data.get("actual_check_in", timezone.now())
            booking.notes = serializer.validated_data.get("notes", booking.notes or "")
            booking.save()

            # Update room status to OCCUPIED
            room = booking.room
            room.status = Room.Status.OCCUPIED
            room.save()

        # Notify staff about check-in
        from .services import PushNotificationService

        PushNotificationService.notify_staff(
            notification_type=Notification.NotificationType.CHECKIN_COMPLETED,
            title=f"Check-in: Phòng {booking.room.number}",
            body=f"{booking.guest.full_name} đã nhận phòng {booking.room.number}",
            data={
                "booking_id": str(booking.id),
                "room_number": booking.room.number,
                "action": "check_in",
            },
            booking=booking,
            exclude_user=request.user,
        )

        return Response(
            BookingSerializer(booking).data,
            status=status.HTTP_200_OK,
        )

    @extend_schema(
        summary="Check out guest",
        description="Check out a guest from their booking. Updates status to CHECKED_OUT.",
        request=CheckOutSerializer,
        responses={200: BookingSerializer},
        tags=["Booking Management"],
    )
    @action(detail=True, methods=["post"], url_path="check-out")
    def check_out(self, request, pk=None):
        """Check out a guest."""
        from django.db import transaction
        from django.utils import timezone

        booking = self.get_object()

        if booking.status == Booking.Status.CHECKED_OUT:
            return Response(
                {"detail": "Guest is already checked out."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        if booking.status != Booking.Status.CHECKED_IN:
            return Response(
                {"detail": f"Cannot check out booking with status {booking.get_status_display()}."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        serializer = CheckOutSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        # Use atomic transaction to ensure consistency
        with transaction.atomic():
            booking.status = Booking.Status.CHECKED_OUT
            booking.actual_check_out = serializer.validated_data.get("actual_check_out", timezone.now())
            # Append to existing notes if provided
            additional_notes = serializer.validated_data.get("notes", "")
            if additional_notes:
                booking.notes = f"{booking.notes}\n{additional_notes}" if booking.notes else additional_notes
            # Store additional charges
            booking.additional_charges = serializer.validated_data.get("additional_charges", 0)
            booking.save()

            # Update room status to CLEANING
            room = booking.room
            room.status = Room.Status.CLEANING
            room.save()

            # Increment guest's total stays
            guest = booking.guest
            guest.total_stays += 1
            guest.save()

        # Notify staff about check-out
        from .services import PushNotificationService

        PushNotificationService.notify_staff(
            notification_type=Notification.NotificationType.CHECKOUT_COMPLETED,
            title=f"Check-out: Phòng {booking.room.number}",
            body=f"{booking.guest.full_name} đã trả phòng {booking.room.number}",
            data={
                "booking_id": str(booking.id),
                "room_number": booking.room.number,
                "action": "check_out",
            },
            booking=booking,
            exclude_user=request.user,
        )

        return Response(
            BookingSerializer(booking).data,
            status=status.HTTP_200_OK,
        )

    @action(detail=True, methods=["post"], url_path="record-early-checkin")
    def record_early_checkin(self, request, pk=None):
        """Record early check-in fee for a booking."""
        from django.db import transaction

        booking = self.get_object()

        if booking.status not in [Booking.Status.CHECKED_IN, Booking.Status.CONFIRMED]:
            return Response(
                {"detail": "Chỉ có thể ghi phí nhận sớm cho booking đã xác nhận hoặc đang ở."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        from .serializers import EarlyCheckInFeeSerializer

        serializer = EarlyCheckInFeeSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        hours = serializer.validated_data["hours"]
        fee = serializer.validated_data["fee"]
        notes = serializer.validated_data.get("notes", "")
        create_folio = serializer.validated_data.get("create_folio_item", True)

        with transaction.atomic():
            booking.early_check_in_hours = hours
            booking.early_check_in_fee = fee
            if notes:
                booking.notes = f"{booking.notes}\n[Nhận sớm] {notes}" if booking.notes else f"[Nhận sớm] {notes}"
            booking.save(update_fields=[
                "early_check_in_hours", "early_check_in_fee", "notes",
            ])

            # Optionally create a FolioItem for tracking
            # Mark as is_paid=True to prevent double-counting in additional_charges
            # (fees are already tracked via dedicated fields in balance_due)
            if create_folio and fee > 0:
                from datetime import date as date_today
                FolioItem.objects.get_or_create(
                    booking=booking,
                    item_type="early_checkin",
                    defaults={
                        "description": f"Phí nhận sớm ({hours}h)",
                        "unit_price": fee,
                        "total_price": fee,
                        "quantity": 1,
                        "date": date_today.today(),
                        "is_paid": True,
                        "created_by": request.user,
                    },
                )

        return Response(
            BookingSerializer(booking).data,
            status=status.HTTP_200_OK,
        )

    @action(detail=True, methods=["post"], url_path="record-late-checkout")
    def record_late_checkout(self, request, pk=None):
        """Record late check-out fee for a booking."""
        from django.db import transaction

        booking = self.get_object()

        if booking.status != Booking.Status.CHECKED_IN:
            return Response(
                {"detail": "Chỉ có thể ghi phí trả muộn cho booking đang ở."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        from .serializers import LateCheckOutFeeSerializer

        serializer = LateCheckOutFeeSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        hours = serializer.validated_data["hours"]
        fee = serializer.validated_data["fee"]
        notes = serializer.validated_data.get("notes", "")
        create_folio = serializer.validated_data.get("create_folio_item", True)

        with transaction.atomic():
            booking.late_check_out_hours = hours
            booking.late_check_out_fee = fee
            if notes:
                booking.notes = f"{booking.notes}\n[Trả muộn] {notes}" if booking.notes else f"[Trả muộn] {notes}"
            booking.save(update_fields=[
                "late_check_out_hours", "late_check_out_fee", "notes",
            ])

            # Optionally create a FolioItem for tracking
            # Mark as is_paid=True to prevent double-counting in additional_charges
            if create_folio and fee > 0:
                from datetime import date as date_today
                FolioItem.objects.get_or_create(
                    booking=booking,
                    item_type="late_checkout",
                    defaults={
                        "description": f"Phí trả muộn ({hours}h)",
                        "unit_price": fee,
                        "total_price": fee,
                        "quantity": 1,
                        "date": date_today.today(),
                        "is_paid": True,
                        "created_by": request.user,
                    },
                )

        return Response(
            BookingSerializer(booking).data,
            status=status.HTTP_200_OK,
        )

    @extend_schema(
        summary="Get today's bookings",
        description="Get all bookings for today (check-ins and check-outs).",
        responses={
            200: OpenApiResponse(
                description="Today's bookings",
                response={
                    "type": "object",
                    "properties": {
                        "check_ins": {"type": "array", "items": {"$ref": "#/components/schemas/BookingList"}},
                        "check_outs": {"type": "array", "items": {"$ref": "#/components/schemas/BookingList"}},
                        "total_check_ins": {"type": "integer"},
                        "total_check_outs": {"type": "integer"},
                    },
                },
            )
        },
        tags=["Booking Management"],
    )
    @action(detail=False, methods=["get"], url_path="today")
    def today(self, request):
        """Get today's bookings."""
        from datetime import date

        today = date.today()

        check_ins = (
            Booking.objects.filter(check_in_date=today)
            .select_related("guest", "room", "room__room_type")
            .order_by("check_in_date")
        )

        check_outs = (
            Booking.objects.filter(check_out_date=today)
            .select_related("guest", "room", "room__room_type")
            .order_by("check_out_date")
        )

        return Response(
            {
                "check_ins": BookingListSerializer(check_ins, many=True).data,
                "check_outs": BookingListSerializer(check_outs, many=True).data,
                "total_check_ins": check_ins.count(),
                "total_check_outs": check_outs.count(),
            },
            status=status.HTTP_200_OK,
        )

    @extend_schema(
        summary="Get calendar view",
        description="Get bookings for a date range, useful for calendar view.",
        parameters=[
            OpenApiParameter(
                name="start_date",
                type=str,
                description="Start date (YYYY-MM-DD)",
                required=True,
            ),
            OpenApiParameter(
                name="end_date",
                type=str,
                description="End date (YYYY-MM-DD)",
                required=True,
            ),
        ],
        responses={
            200: OpenApiResponse(
                description="Bookings in date range",
                response={
                    "type": "object",
                    "properties": {
                        "bookings": {"type": "array", "items": {"$ref": "#/components/schemas/BookingList"}},
                        "total": {"type": "integer"},
                        "start_date": {"type": "string", "format": "date"},
                        "end_date": {"type": "string", "format": "date"},
                    },
                },
            )
        },
        tags=["Booking Management"],
    )
    @action(detail=False, methods=["get"], url_path="calendar")
    def calendar(self, request):
        """Get bookings for calendar view."""
        from datetime import datetime

        start_date = request.query_params.get("start_date")
        end_date = request.query_params.get("end_date")

        if not start_date or not end_date:
            return Response(
                {"detail": "Cần cung cấp start_date và end_date."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        try:
            start_date = datetime.strptime(start_date, "%Y-%m-%d").date()
            end_date = datetime.strptime(end_date, "%Y-%m-%d").date()
        except ValueError:
            return Response(
                {"detail": "Định dạng ngày không hợp lệ. Sử dụng YYYY-MM-DD."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        bookings = (
            Booking.objects.filter(
                Q(check_in_date__range=[start_date, end_date]) | Q(check_out_date__range=[start_date, end_date])
            )
            .select_related("guest", "room", "room__room_type")
            .order_by("check_in_date")
        )

        return Response(
            {
                "bookings": BookingListSerializer(bookings, many=True).data,
                "total": bookings.count(),
                "start_date": start_date,
                "end_date": end_date,
            },
            status=status.HTTP_200_OK,
        )


# ==================== Dashboard Views ====================


@extend_schema_view(
    get=extend_schema(
        summary="Get dashboard summary",
        description="Get aggregated dashboard metrics including room status, today's stats, and revenue summary.",
        responses={
            200: OpenApiResponse(
                description="Dashboard summary",
                response={
                    "type": "object",
                    "properties": {
                        "room_status": {
                            "type": "object",
                            "properties": {
                                "total": {"type": "integer"},
                                "available": {"type": "integer"},
                                "occupied": {"type": "integer"},
                                "cleaning": {"type": "integer"},
                                "maintenance": {"type": "integer"},
                                "blocked": {"type": "integer"},
                            },
                        },
                        "today": {
                            "type": "object",
                            "properties": {
                                "date": {"type": "string", "format": "date"},
                                "check_ins": {"type": "integer"},
                                "check_outs": {"type": "integer"},
                                "pending_arrivals": {"type": "integer"},
                                "pending_departures": {"type": "integer"},
                            },
                        },
                        "occupancy": {
                            "type": "object",
                            "properties": {
                                "rate": {"type": "number", "format": "float"},
                                "occupied_rooms": {"type": "integer"},
                                "total_rooms": {"type": "integer"},
                            },
                        },
                        "bookings": {
                            "type": "object",
                            "properties": {
                                "pending": {"type": "integer"},
                                "confirmed": {"type": "integer"},
                                "checked_in": {"type": "integer"},
                            },
                        },
                    },
                },
            )
        },
        tags=["Dashboard"],
    )
)
class DashboardView(APIView):
    """Dashboard summary endpoint for hotel overview."""

    permission_classes = [IsAuthenticated, IsStaff]

    def get(self, request):
        """Get dashboard summary metrics."""
        from datetime import date

        from django.db.models import Count

        today = date.today()

        # Room status summary
        room_status = (
            Room.objects.filter(is_active=True)
            .values("status")
            .annotate(count=Count("id"))
        )
        room_status_dict = {item["status"]: item["count"] for item in room_status}
        total_rooms = Room.objects.filter(is_active=True).count()

        # Today's check-ins and check-outs
        today_check_ins = Booking.objects.filter(
            check_in_date=today,
            status__in=[Booking.Status.PENDING, Booking.Status.CONFIRMED],
        ).count()
        today_check_outs = Booking.objects.filter(
            check_out_date=today,
            status=Booking.Status.CHECKED_IN,
        ).count()

        # Already checked in/out today
        completed_check_ins = Booking.objects.filter(
            check_in_date=today,
            status=Booking.Status.CHECKED_IN,
        ).count()
        completed_check_outs = Booking.objects.filter(
            check_out_date=today,
            status=Booking.Status.CHECKED_OUT,
        ).count()

        # Booking status counts
        pending_bookings = Booking.objects.filter(status=Booking.Status.PENDING).count()
        confirmed_bookings = Booking.objects.filter(status=Booking.Status.CONFIRMED).count()
        checked_in_bookings = Booking.objects.filter(status=Booking.Status.CHECKED_IN).count()

        # Calculate occupancy
        occupied_rooms = room_status_dict.get(Room.Status.OCCUPIED, 0)
        occupancy_rate = (occupied_rooms / total_rooms * 100) if total_rooms > 0 else 0

        # Today's financial summary
        from django.db.models import Sum

        today_income = (
            FinancialEntry.objects.filter(
                date=today, entry_type=FinancialEntry.EntryType.INCOME
            ).aggregate(total=Sum("amount"))["total"]
            or 0
        )
        today_expense = (
            FinancialEntry.objects.filter(
                date=today, entry_type=FinancialEntry.EntryType.EXPENSE
            ).aggregate(total=Sum("amount"))["total"]
            or 0
        )

        return Response(
            {
                "room_status": {
                    "total": total_rooms,
                    "available": room_status_dict.get(Room.Status.AVAILABLE, 0),
                    "occupied": occupied_rooms,
                    "cleaning": room_status_dict.get(Room.Status.CLEANING, 0),
                    "maintenance": room_status_dict.get(Room.Status.MAINTENANCE, 0),
                    "blocked": room_status_dict.get(Room.Status.BLOCKED, 0),
                },
                "today": {
                    "date": today.isoformat(),
                    "check_ins": today_check_ins + completed_check_ins,
                    "check_outs": today_check_outs + completed_check_outs,
                    "pending_arrivals": today_check_ins,
                    "pending_departures": today_check_outs,
                    "revenue": float(today_income),
                    "expense": float(today_expense),
                },
                "occupancy": {
                    "rate": round(occupancy_rate, 1),
                    "occupied_rooms": occupied_rooms,
                    "total_rooms": total_rooms,
                },
                "bookings": {
                    "pending": pending_bookings,
                    "confirmed": confirmed_bookings,
                    "checked_in": checked_in_bookings,
                },
            },
            status=status.HTTP_200_OK,
        )


# ==================== Financial Management Views ====================


@extend_schema_view(
    list=extend_schema(
        summary="List financial categories",
        description="Get list of all financial categories.",
        parameters=[
            OpenApiParameter(
                name="category_type",
                type=str,
                description="Filter by category type (income/expense)",
                required=False,
            ),
            OpenApiParameter(
                name="is_active",
                type=bool,
                description="Filter by active status",
                required=False,
            ),
        ],
        responses={200: FinancialCategoryListSerializer(many=True)},
        tags=["Financial Management"],
    ),
    retrieve=extend_schema(
        summary="Get financial category details",
        description="Get detailed information about a specific financial category.",
        responses={200: FinancialCategorySerializer},
        tags=["Financial Management"],
    ),
    create=extend_schema(
        summary="Create financial category",
        description="Create a new financial category. Requires manager role.",
        request=FinancialCategorySerializer,
        responses={201: FinancialCategorySerializer},
        tags=["Financial Management"],
    ),
    update=extend_schema(
        summary="Update financial category",
        description="Update a financial category. Requires manager role.",
        request=FinancialCategorySerializer,
        responses={200: FinancialCategorySerializer},
        tags=["Financial Management"],
    ),
    partial_update=extend_schema(
        summary="Partially update financial category",
        description="Partially update a financial category. Requires manager role.",
        request=FinancialCategorySerializer,
        responses={200: FinancialCategorySerializer},
        tags=["Financial Management"],
    ),
    destroy=extend_schema(
        summary="Delete financial category",
        description="Delete a financial category. Requires manager role.",
        responses={204: None},
        tags=["Financial Management"],
    ),
)
class FinancialCategoryViewSet(viewsets.ModelViewSet):
    """ViewSet for FinancialCategory CRUD operations."""

    queryset = FinancialCategory.objects.all()
    permission_classes = [IsAuthenticated, IsStaff]

    def get_serializer_class(self):
        """Return appropriate serializer class."""
        if self.action == "list":
            return FinancialCategoryListSerializer
        return FinancialCategorySerializer

    def get_permissions(self):
        """Set permissions based on action."""
        if self.action in ["create", "update", "partial_update", "destroy"]:
            return [IsAuthenticated(), IsManager()]
        return [IsAuthenticated(), IsStaff()]

    def get_queryset(self):
        """Filter queryset based on query parameters."""
        from django.db.models import Count

        queryset = super().get_queryset().annotate(entry_count=Count("entries"))

        # Filter by category type
        category_type = self.request.query_params.get("category_type")
        if category_type:
            queryset = queryset.filter(category_type=category_type)

        # Filter by active status
        is_active = self.request.query_params.get("is_active")
        if is_active is not None:
            is_active_bool = is_active.lower() in ["true", "1", "yes"]
            queryset = queryset.filter(is_active=is_active_bool)

        return queryset


@extend_schema_view(
    list=extend_schema(
        summary="List financial entries",
        description="Get list of financial entries with filtering options.",
        parameters=[
            OpenApiParameter(
                name="entry_type",
                type=str,
                description="Filter by entry type (income/expense)",
                required=False,
            ),
            OpenApiParameter(
                name="category",
                type=int,
                description="Filter by category ID",
                required=False,
            ),
            OpenApiParameter(
                name="date_from",
                type=str,
                description="Filter entries from this date (YYYY-MM-DD)",
                required=False,
            ),
            OpenApiParameter(
                name="date_to",
                type=str,
                description="Filter entries until this date (YYYY-MM-DD)",
                required=False,
            ),
            OpenApiParameter(
                name="payment_method",
                type=str,
                description="Filter by payment method",
                required=False,
            ),
        ],
        responses={200: FinancialEntryListSerializer(many=True)},
        tags=["Financial Management"],
    ),
    retrieve=extend_schema(
        summary="Get financial entry details",
        description="Get detailed information about a specific financial entry.",
        responses={200: FinancialEntrySerializer},
        tags=["Financial Management"],
    ),
    create=extend_schema(
        summary="Create financial entry",
        description="Create a new financial entry (income or expense).",
        request=FinancialEntrySerializer,
        responses={201: FinancialEntrySerializer},
        tags=["Financial Management"],
    ),
    update=extend_schema(
        summary="Update financial entry",
        description="Update a financial entry. Requires manager role.",
        request=FinancialEntrySerializer,
        responses={200: FinancialEntrySerializer},
        tags=["Financial Management"],
    ),
    partial_update=extend_schema(
        summary="Partially update financial entry",
        description="Partially update a financial entry. Requires manager role.",
        request=FinancialEntrySerializer,
        responses={200: FinancialEntrySerializer},
        tags=["Financial Management"],
    ),
    destroy=extend_schema(
        summary="Delete financial entry",
        description="Delete a financial entry. Requires manager role.",
        responses={204: None},
        tags=["Financial Management"],
    ),
)
class FinancialEntryViewSet(viewsets.ModelViewSet):
    """ViewSet for FinancialEntry CRUD operations."""

    queryset = FinancialEntry.objects.all()
    permission_classes = [IsAuthenticated, IsStaff]

    def get_serializer_class(self):
        """Return appropriate serializer class."""
        if self.action == "list":
            return FinancialEntryListSerializer
        return FinancialEntrySerializer

    def get_permissions(self):
        """Set permissions based on action."""
        if self.action in ["update", "partial_update", "destroy"]:
            return [IsAuthenticated(), IsManager()]
        return [IsAuthenticated(), IsStaff()]

    def get_queryset(self):
        """Filter queryset based on query parameters."""
        queryset = super().get_queryset().select_related(
            "category", "booking", "booking__room", "booking__guest", "created_by"
        )

        # Filter by entry type
        entry_type = self.request.query_params.get("entry_type")
        if entry_type:
            queryset = queryset.filter(entry_type=entry_type)

        # Filter by category
        category = self.request.query_params.get("category")
        if category:
            queryset = queryset.filter(category_id=category)

        # Filter by date range
        date_from = self.request.query_params.get("date_from")
        if date_from:
            queryset = queryset.filter(date__gte=date_from)

        date_to = self.request.query_params.get("date_to")
        if date_to:
            queryset = queryset.filter(date__lte=date_to)

        # Filter by payment method
        payment_method = self.request.query_params.get("payment_method")
        if payment_method:
            queryset = queryset.filter(payment_method=payment_method)

        return queryset.order_by("-date", "-created_at")

    def perform_create(self, serializer):
        """Set created_by to current user."""
        serializer.save(created_by=self.request.user)

    @extend_schema(
        summary="Get daily financial summary",
        description="Get financial summary for a specific day.",
        parameters=[
            OpenApiParameter(
                name="date",
                type=str,
                description="Date for summary (YYYY-MM-DD). Defaults to today.",
                required=False,
            ),
        ],
        responses={
            200: OpenApiResponse(
                description="Daily financial summary",
                response={
                    "type": "object",
                    "properties": {
                        "date": {"type": "string", "format": "date"},
                        "total_income": {"type": "number"},
                        "total_expense": {"type": "number"},
                        "net_profit": {"type": "number"},
                        "income_entries": {"type": "integer"},
                        "expense_entries": {"type": "integer"},
                        "income_by_category": {"type": "array"},
                        "expense_by_category": {"type": "array"},
                    },
                },
            )
        },
        tags=["Financial Management"],
    )
    @action(detail=False, methods=["get"], url_path="daily-summary")
    def daily_summary(self, request):
        """Get financial summary for a day."""
        from datetime import date, datetime

        from django.db.models import Sum

        date_param = request.query_params.get("date")
        if date_param:
            try:
                summary_date = datetime.strptime(date_param, "%Y-%m-%d").date()
            except ValueError:
                return Response(
                    {"detail": "Định dạng ngày không hợp lệ. Sử dụng YYYY-MM-DD."},
                    status=status.HTTP_400_BAD_REQUEST,
                )
        else:
            summary_date = date.today()

        # Get totals
        entries = FinancialEntry.objects.filter(date=summary_date)

        total_income = (
            entries.filter(entry_type=FinancialEntry.EntryType.INCOME).aggregate(total=Sum("amount"))[
                "total"
            ]
            or 0
        )

        total_expense = (
            entries.filter(entry_type=FinancialEntry.EntryType.EXPENSE).aggregate(total=Sum("amount"))[
                "total"
            ]
            or 0
        )

        # Get breakdown by category
        income_by_category = list(
            entries.filter(entry_type=FinancialEntry.EntryType.INCOME)
            .values("category__name", "category__icon", "category__color")
            .annotate(total=Sum("amount"))
            .order_by("-total")
        )

        expense_by_category = list(
            entries.filter(entry_type=FinancialEntry.EntryType.EXPENSE)
            .values("category__name", "category__icon", "category__color")
            .annotate(total=Sum("amount"))
            .order_by("-total")
        )

        return Response(
            {
                "date": summary_date.isoformat(),
                "total_income": total_income,
                "total_expense": total_expense,
                "net_profit": total_income - total_expense,
                "income_entries": entries.filter(entry_type=FinancialEntry.EntryType.INCOME).count(),
                "expense_entries": entries.filter(entry_type=FinancialEntry.EntryType.EXPENSE).count(),
                "income_by_category": income_by_category,
                "expense_by_category": expense_by_category,
            },
            status=status.HTTP_200_OK,
        )

    @extend_schema(
        summary="Get monthly financial summary",
        description="Get financial summary for a specific month.",
        parameters=[
            OpenApiParameter(
                name="year",
                type=int,
                description="Year for summary. Defaults to current year.",
                required=False,
            ),
            OpenApiParameter(
                name="month",
                type=int,
                description="Month for summary (1-12). Defaults to current month.",
                required=False,
            ),
        ],
        responses={
            200: OpenApiResponse(
                description="Monthly financial summary",
                response={
                    "type": "object",
                    "properties": {
                        "year": {"type": "integer"},
                        "month": {"type": "integer"},
                        "total_income": {"type": "number"},
                        "total_expense": {"type": "number"},
                        "net_profit": {"type": "number"},
                        "profit_margin": {"type": "number"},
                        "income_by_category": {"type": "array"},
                        "expense_by_category": {"type": "array"},
                        "daily_totals": {"type": "array"},
                    },
                },
            )
        },
        tags=["Financial Management"],
    )
    @action(detail=False, methods=["get"], url_path="monthly-summary")
    def monthly_summary(self, request):
        """Get financial summary for a month."""
        from datetime import date

        from django.db.models import Sum
        from django.db.models.functions import TruncDate

        # Get year and month parameters
        year = request.query_params.get("year")
        month = request.query_params.get("month")

        today = date.today()
        year = int(year) if year else today.year
        month = int(month) if month else today.month

        if not (1 <= month <= 12):
            return Response(
                {"detail": "Tháng phải từ 1 đến 12."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        # Get entries for the month
        entries = FinancialEntry.objects.filter(date__year=year, date__month=month)

        total_income = (
            entries.filter(entry_type=FinancialEntry.EntryType.INCOME).aggregate(total=Sum("amount"))[
                "total"
            ]
            or 0
        )

        total_expense = (
            entries.filter(entry_type=FinancialEntry.EntryType.EXPENSE).aggregate(total=Sum("amount"))[
                "total"
            ]
            or 0
        )

        net_profit = total_income - total_expense
        profit_margin = (net_profit / total_income * 100) if total_income > 0 else 0

        # Get breakdown by category
        income_by_category = list(
            entries.filter(entry_type=FinancialEntry.EntryType.INCOME)
            .values("category__name", "category__icon", "category__color")
            .annotate(total=Sum("amount"))
            .order_by("-total")
        )

        expense_by_category = list(
            entries.filter(entry_type=FinancialEntry.EntryType.EXPENSE)
            .values("category__name", "category__icon", "category__color")
            .annotate(total=Sum("amount"))
            .order_by("-total")
        )

        # Get daily totals for chart - manual grouping for SQLite compatibility
        from collections import defaultdict

        daily_data = defaultdict(lambda: {"income": 0, "expense": 0})
        for entry in entries:
            day_str = entry.date.isoformat()
            if entry.entry_type == FinancialEntry.EntryType.INCOME:
                daily_data[day_str]["income"] += entry.amount
            else:
                daily_data[day_str]["expense"] += entry.amount

        daily_totals = [
            {"day": day, "income": data["income"], "expense": data["expense"]}
            for day, data in sorted(daily_data.items())
        ]

        return Response(
            {
                "year": year,
                "month": month,
                "total_income": total_income,
                "total_expense": total_expense,
                "net_profit": net_profit,
                "profit_margin": round(profit_margin, 1),
                "income_by_category": income_by_category,
                "expense_by_category": expense_by_category,
                "daily_totals": daily_totals,
            },
            status=status.HTTP_200_OK,
        )


# Import NightAudit model and serializers
from .models import NightAudit
from .serializers import (
    NightAuditCreateSerializer,
    NightAuditListSerializer,
    NightAuditSerializer,
)


@extend_schema_view(
    list=extend_schema(
        summary="List night audits",
        description="Get a paginated list of night audits with optional filtering.",
        parameters=[
            OpenApiParameter(
                name="status",
                type=str,
                description="Filter by status (draft, completed, closed)",
            ),
            OpenApiParameter(
                name="date_from",
                type=str,
                description="Filter from date (YYYY-MM-DD)",
            ),
            OpenApiParameter(
                name="date_to",
                type=str,
                description="Filter to date (YYYY-MM-DD)",
            ),
        ],
        tags=["Night Audit"],
    ),
    retrieve=extend_schema(
        summary="Get night audit details",
        description="Get detailed information about a specific night audit.",
        tags=["Night Audit"],
    ),
    create=extend_schema(
        summary="Create/generate night audit",
        description="Create a new night audit for a specific date. Statistics are calculated automatically.",
        tags=["Night Audit"],
    ),
    update=extend_schema(
        summary="Update night audit",
        description="Update night audit details (only for draft audits).",
        tags=["Night Audit"],
    ),
    partial_update=extend_schema(
        summary="Partially update night audit",
        description="Partially update night audit details (only for draft audits).",
        tags=["Night Audit"],
    ),
    destroy=extend_schema(
        summary="Delete night audit",
        description="Delete a night audit (only for draft audits).",
        tags=["Night Audit"],
    ),
)
class NightAuditViewSet(viewsets.ModelViewSet):
    """
    ViewSet for managing night audits.

    Endpoints:
    - GET /night-audits/ - List all night audits
    - POST /night-audits/ - Generate night audit for a date
    - GET /night-audits/{id}/ - Get audit details
    - PUT /night-audits/{id}/ - Update audit (draft only)
    - DELETE /night-audits/{id}/ - Delete audit (draft only)
    - POST /night-audits/{id}/close/ - Close the audit
    - POST /night-audits/{id}/recalculate/ - Recalculate statistics
    - GET /night-audits/today/ - Get or create today's audit
    - GET /night-audits/latest/ - Get the most recent audit
    """

    queryset = NightAudit.objects.all()
    permission_classes = [IsAuthenticated]

    def get_serializer_class(self):
        """Return appropriate serializer based on action."""
        if self.action == "list":
            return NightAuditListSerializer
        if self.action == "create":
            return NightAuditCreateSerializer
        return NightAuditSerializer

    def get_queryset(self):
        """Filter queryset based on query parameters."""
        queryset = NightAudit.objects.all()

        # Filter by status
        status_filter = self.request.query_params.get("status")
        if status_filter:
            queryset = queryset.filter(status=status_filter)

        # Filter by date range
        date_from = self.request.query_params.get("date_from")
        date_to = self.request.query_params.get("date_to")
        if date_from:
            queryset = queryset.filter(audit_date__gte=date_from)
        if date_to:
            queryset = queryset.filter(audit_date__lte=date_to)

        return queryset.order_by("-audit_date")

    def create(self, request, *args, **kwargs):
        """Create or generate a night audit for a specific date."""
        from django.utils import timezone

        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        audit_date = serializer.validated_data["audit_date"]
        notes = serializer.validated_data.get("notes", "")

        # Check if audit already exists for this date
        existing = NightAudit.objects.filter(audit_date=audit_date).first()
        if existing:
            return Response(
                {
                    "detail": f"Đã tồn tại kiểm toán cho ngày {audit_date}.",
                    "existing_id": existing.id,
                },
                status=status.HTTP_400_BAD_REQUEST,
            )

        # Create new audit
        audit = NightAudit.objects.create(
            audit_date=audit_date,
            notes=notes,
            performed_by=request.user,
            performed_at=timezone.now(),
            status=NightAudit.Status.DRAFT,
        )

        # Calculate statistics
        audit.calculate_statistics()
        audit.save()

        return Response(
            NightAuditSerializer(audit).data,
            status=status.HTTP_201_CREATED,
        )

    def update(self, request, *args, **kwargs):
        """Update night audit (only draft audits)."""
        instance = self.get_object()

        if instance.status == NightAudit.Status.CLOSED:
            return Response(
                {"detail": "Không thể chỉnh sửa kiểm toán đã đóng."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        return super().update(request, *args, **kwargs)

    def destroy(self, request, *args, **kwargs):
        """Delete night audit (only draft audits)."""
        instance = self.get_object()

        if instance.status == NightAudit.Status.CLOSED:
            return Response(
                {"detail": "Không thể xóa kiểm toán đã đóng."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        return super().destroy(request, *args, **kwargs)

    @extend_schema(
        summary="Close night audit",
        description="Close the audit - no more changes allowed after closing.",
        request=None,
        responses={200: NightAuditSerializer},
        tags=["Night Audit"],
    )
    @action(detail=True, methods=["post"], url_path="close")
    def close(self, request, pk=None):
        """Close the night audit."""
        audit = self.get_object()

        if audit.status == NightAudit.Status.CLOSED:
            return Response(
                {"detail": "Kiểm toán đã được đóng."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        audit.close_audit(request.user)
        return Response(NightAuditSerializer(audit).data)

    @extend_schema(
        summary="Recalculate statistics",
        description="Recalculate all statistics for the audit (only for non-closed audits).",
        request=None,
        responses={200: NightAuditSerializer},
        tags=["Night Audit"],
    )
    @action(detail=True, methods=["post"], url_path="recalculate")
    def recalculate(self, request, pk=None):
        """Recalculate audit statistics."""
        audit = self.get_object()

        if audit.status == NightAudit.Status.CLOSED:
            return Response(
                {"detail": "Không thể tính lại kiểm toán đã đóng."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        audit.calculate_statistics()
        audit.save()

        return Response(NightAuditSerializer(audit).data)

    @extend_schema(
        summary="Get today's audit",
        description="Get today's night audit. Creates a draft if not exists.",
        responses={200: NightAuditSerializer},
        tags=["Night Audit"],
    )
    @action(detail=False, methods=["get"], url_path="today")
    def today(self, request):
        """Get or create today's night audit."""
        from datetime import date

        from django.utils import timezone

        today = date.today()
        audit = NightAudit.objects.filter(audit_date=today).first()

        if not audit:
            # Create draft audit for today
            audit = NightAudit.objects.create(
                audit_date=today,
                performed_by=request.user,
                performed_at=timezone.now(),
                status=NightAudit.Status.DRAFT,
            )
            audit.calculate_statistics()
            audit.save()

        return Response(NightAuditSerializer(audit).data)

    @extend_schema(
        summary="Get latest audit",
        description="Get the most recent night audit.",
        responses={
            200: NightAuditSerializer,
            404: OpenApiResponse(description="No audits found"),
        },
        tags=["Night Audit"],
    )
    @action(detail=False, methods=["get"], url_path="latest")
    def latest(self, request):
        """Get the most recent night audit."""
        audit = NightAudit.objects.order_by("-audit_date").first()

        if not audit:
            return Response(
                {"detail": "Chưa có kiểm toán nào."},
                status=status.HTTP_404_NOT_FOUND,
            )

        return Response(NightAuditSerializer(audit).data)


# ============================================================
# Payment ViewSet (Phase 2.1.3)
# ============================================================


class PaymentViewSet(viewsets.ModelViewSet):
    """
    API endpoints for Payment management.

    Provides:
    - List payments (with filtering by booking, type, status, date range)
    - Create payment
    - Retrieve payment details
    - Update payment
    - Delete payment (only pending payments)
    - Record deposit action
    - Get deposits for a booking
    - Get outstanding deposits report
    """

    permission_classes = [IsAuthenticated, IsStaffOrManager]

    def get_queryset(self):
        from .models import Payment

        queryset = Payment.objects.select_related(
            "booking__room",
            "booking__guest",
            "created_by",
        ).all()

        # Filter by booking
        booking_id = self.request.query_params.get("booking")
        if booking_id:
            queryset = queryset.filter(booking_id=booking_id)

        # Filter by payment type
        payment_type = self.request.query_params.get("payment_type")
        if payment_type:
            queryset = queryset.filter(payment_type=payment_type)

        # Filter by status
        payment_status = self.request.query_params.get("status")
        if payment_status:
            queryset = queryset.filter(status=payment_status)

        # Filter by date range
        date_from = self.request.query_params.get("date_from")
        date_to = self.request.query_params.get("date_to")
        if date_from:
            queryset = queryset.filter(payment_date__date__gte=date_from)
        if date_to:
            queryset = queryset.filter(payment_date__date__lte=date_to)

        return queryset

    def get_serializer_class(self):
        from .serializers import (
            PaymentCreateSerializer,
            PaymentListSerializer,
            PaymentSerializer,
        )

        if self.action == "list":
            return PaymentListSerializer
        elif self.action == "create":
            return PaymentCreateSerializer
        return PaymentSerializer

    @extend_schema(
        summary="List payments",
        description="Get a list of payments with optional filtering.",
        parameters=[
            OpenApiParameter("booking", OpenApiTypes.INT, description="Filter by booking ID"),
            OpenApiParameter("payment_type", OpenApiTypes.STR, description="Filter by type (deposit, room_charge, etc.)"),
            OpenApiParameter("status", OpenApiTypes.STR, description="Filter by status"),
            OpenApiParameter("date_from", OpenApiTypes.DATE, description="Filter by date from"),
            OpenApiParameter("date_to", OpenApiTypes.DATE, description="Filter by date to"),
        ],
        tags=["Payments"],
    )
    def list(self, request, *args, **kwargs):
        return super().list(request, *args, **kwargs)

    @extend_schema(
        summary="Create payment",
        description="Create a new payment record.",
        tags=["Payments"],
    )
    def create(self, request, *args, **kwargs):
        return super().create(request, *args, **kwargs)

    @extend_schema(
        summary="Record deposit",
        description="Quick action to record a deposit payment for a booking.",
        request=DepositRecordSerializer,
        responses={201: PaymentSerializer},
        tags=["Payments"],
    )
    @action(detail=False, methods=["post"], url_path="record-deposit")
    def record_deposit(self, request):
        """Record a deposit payment for a booking."""
        from .models import Payment
        from .serializers import DepositRecordSerializer, PaymentSerializer

        serializer = DepositRecordSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        booking = Booking.objects.get(id=serializer.validated_data["booking_id"])

        payment = Payment.objects.create(
            booking=booking,
            payment_type=Payment.PaymentType.DEPOSIT,
            amount=serializer.validated_data["amount"],
            payment_method=serializer.validated_data["payment_method"],
            transaction_id=serializer.validated_data.get("transaction_id", ""),
            notes=serializer.validated_data.get("notes", ""),
            status=Payment.Status.COMPLETED,
            created_by=request.user,
        )

        # Update booking deposit amount
        from django.db.models import Sum

        total_deposits = Payment.objects.filter(
            booking=booking,
            payment_type=Payment.PaymentType.DEPOSIT,
            status=Payment.Status.COMPLETED,
        ).aggregate(total=Sum("amount"))["total"] or 0

        booking.deposit_amount = total_deposits
        booking.deposit_paid = total_deposits >= booking.total_amount * Decimal("0.3")
        booking.save(update_fields=["deposit_amount", "deposit_paid"])

        return Response(PaymentSerializer(payment).data, status=status.HTTP_201_CREATED)

    @extend_schema(
        summary="Get booking deposits",
        description="Get all deposit payments for a specific booking.",
        parameters=[
            OpenApiParameter("booking_id", OpenApiTypes.INT, location=OpenApiParameter.PATH),
        ],
        responses={200: PaymentListSerializer(many=True)},
        tags=["Payments"],
    )
    @action(detail=False, methods=["get"], url_path="booking/(?P<booking_id>[^/.]+)/deposits")
    def booking_deposits(self, request, booking_id=None):
        """Get all deposits for a booking."""
        from .models import Payment
        from .serializers import PaymentListSerializer

        deposits = Payment.objects.filter(
            booking_id=booking_id,
            payment_type=Payment.PaymentType.DEPOSIT,
        ).order_by("-payment_date")

        serializer = PaymentListSerializer(deposits, many=True)
        return Response(serializer.data)

    @extend_schema(
        summary="Outstanding deposits report",
        description="Get list of bookings with outstanding/pending deposits.",
        responses={200: OutstandingDepositSerializer(many=True)},
        tags=["Payments"],
    )
    @action(detail=False, methods=["get"], url_path="outstanding-deposits")
    def outstanding_deposits(self, request):
        """Get bookings with outstanding deposits."""
        from .serializers import OutstandingDepositSerializer

        bookings = Booking.objects.filter(
            deposit_paid=False,
            status__in=[
                Booking.Status.CONFIRMED,
                Booking.Status.CHECKED_IN,
            ],
        ).select_related("room__room_type", "guest").order_by("check_in_date")

        serializer = OutstandingDepositSerializer(bookings, many=True)
        return Response(serializer.data)


# ============================================================
# Folio Item ViewSet (Phase 2.1.4)
# ============================================================


class FolioItemViewSet(viewsets.ModelViewSet):
    """
    API endpoints for Folio Item management (room charges).

    Provides:
    - List folio items
    - Create folio item
    - Update folio item
    - Void folio item
    - Get booking folio
    """

    permission_classes = [IsAuthenticated, IsStaffOrManager]

    def get_queryset(self):
        from .models import FolioItem

        queryset = FolioItem.objects.select_related(
            "booking__room",
            "booking__guest",
            "created_by",
        ).all()

        # Filter by booking
        booking_id = self.request.query_params.get("booking")
        if booking_id:
            queryset = queryset.filter(booking_id=booking_id)

        # Filter by type
        item_type = self.request.query_params.get("item_type")
        if item_type:
            queryset = queryset.filter(item_type=item_type)

        # Filter by status
        include_voided = self.request.query_params.get("include_voided", "false").lower() == "true"
        if not include_voided:
            queryset = queryset.filter(is_voided=False)

        return queryset

    def get_serializer_class(self):
        from .serializers import FolioItemCreateSerializer, FolioItemSerializer

        if self.action == "create":
            return FolioItemCreateSerializer
        return FolioItemSerializer

    def create(self, request, *args, **kwargs):
        """Create folio item and return full serialized response."""
        from .serializers import FolioItemCreateSerializer, FolioItemSerializer

        serializer = FolioItemCreateSerializer(data=request.data, context={"request": request})
        serializer.is_valid(raise_exception=True)
        instance = serializer.save()
        return Response(
            FolioItemSerializer(instance).data,
            status=status.HTTP_201_CREATED,
        )

    @extend_schema(
        summary="Void folio item",
        description="Void a folio item (soft delete with reason).",
        request={"type": "object", "properties": {"reason": {"type": "string"}}},
        tags=["Folio Items"],
    )
    @action(detail=True, methods=["post"], url_path="void")
    def void(self, request, pk=None):
        """Void a folio item."""
        from .serializers import FolioItemSerializer

        item = self.get_object()

        if item.is_voided:
            return Response(
                {"detail": "Chi phí này đã bị hủy."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        if item.is_paid:
            return Response(
                {"detail": "Không thể hủy chi phí đã thanh toán."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        item.is_voided = True
        item.void_reason = request.data.get("reason", "")
        item.save()

        return Response(FolioItemSerializer(item).data)

    @extend_schema(
        summary="Get booking folio",
        description="Get all folio items for a specific booking.",
        parameters=[
            OpenApiParameter("booking_id", OpenApiTypes.INT, location=OpenApiParameter.PATH),
        ],
        responses={200: FolioItemSerializer(many=True)},
        tags=["Folio Items"],
    )
    @action(detail=False, methods=["get"], url_path="booking/(?P<booking_id>[^/.]+)")
    def booking_folio(self, request, booking_id=None):
        """Get all folio items for a booking."""
        from .models import FolioItem
        from .serializers import FolioItemSerializer

        items = FolioItem.objects.filter(
            booking_id=booking_id,
            is_voided=False,
        ).order_by("date", "created_at")

        serializer = FolioItemSerializer(items, many=True)

        # Calculate summary
        from django.db.models import Sum

        summary = items.aggregate(
            total=Sum("total_price"),
            paid=Sum("total_price", filter=models.Q(is_paid=True)),
            unpaid=Sum("total_price", filter=models.Q(is_paid=False)),
        )

        return Response({
            "items": serializer.data,
            "summary": {
                "total": summary["total"] or 0,
                "paid": summary["paid"] or 0,
                "unpaid": summary["unpaid"] or 0,
            },
        })


# ============================================================
# Exchange Rate ViewSet (Phase 2.6)
# ============================================================


class ExchangeRateViewSet(viewsets.ModelViewSet):
    """
    API endpoints for Exchange Rate management.

    Provides:
    - List exchange rates
    - Create exchange rate
    - Convert currency
    - Get latest rates
    """

    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        from .models import ExchangeRate

        queryset = ExchangeRate.objects.all()

        # Filter by currency pair
        from_currency = self.request.query_params.get("from_currency")
        to_currency = self.request.query_params.get("to_currency")
        if from_currency:
            queryset = queryset.filter(from_currency=from_currency.upper())
        if to_currency:
            queryset = queryset.filter(to_currency=to_currency.upper())

        # Filter by date
        date = self.request.query_params.get("date")
        if date:
            queryset = queryset.filter(date=date)

        return queryset

    def get_serializer_class(self):
        from .serializers import ExchangeRateCreateSerializer, ExchangeRateSerializer

        if self.action == "create":
            return ExchangeRateCreateSerializer
        return ExchangeRateSerializer

    @extend_schema(
        summary="Get latest rates",
        description="Get the latest exchange rates for all currency pairs.",
        responses={200: ExchangeRateSerializer(many=True)},
        tags=["Exchange Rates"],
    )
    @action(detail=False, methods=["get"], url_path="latest")
    def latest_rates(self, request):
        """Get latest exchange rates."""
        from .models import ExchangeRate
        from .serializers import ExchangeRateSerializer

        # Get latest rate for each currency pair
        from django.db.models import Max

        latest_dates = ExchangeRate.objects.values(
            "from_currency", "to_currency"
        ).annotate(latest_date=Max("date"))

        rates = []
        for item in latest_dates:
            rate = ExchangeRate.objects.filter(
                from_currency=item["from_currency"],
                to_currency=item["to_currency"],
                date=item["latest_date"],
            ).first()
            if rate:
                rates.append(rate)

        serializer = ExchangeRateSerializer(rates, many=True)
        return Response(serializer.data)

    @extend_schema(
        summary="Convert currency",
        description="Convert an amount from one currency to another.",
        request=CurrencyConversionSerializer,
        responses={
            200: {"type": "object", "properties": {
                "original_amount": {"type": "number"},
                "from_currency": {"type": "string"},
                "converted_amount": {"type": "number"},
                "to_currency": {"type": "string"},
                "rate": {"type": "number"},
                "date": {"type": "string", "format": "date"},
            }},
        },
        tags=["Exchange Rates"],
    )
    @action(detail=False, methods=["post"], url_path="convert")
    def convert(self, request):
        """Convert currency."""
        from .models import ExchangeRate
        from .serializers import CurrencyConversionSerializer

        serializer = CurrencyConversionSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        amount = serializer.validated_data["amount"]
        from_currency = serializer.validated_data["from_currency"].upper()
        to_currency = serializer.validated_data["to_currency"].upper()
        date = serializer.validated_data.get("date")

        # If same currency, no conversion needed
        if from_currency == to_currency:
            return Response({
                "original_amount": amount,
                "from_currency": from_currency,
                "converted_amount": amount,
                "to_currency": to_currency,
                "rate": 1,
                "date": date or "N/A",
            })

        # Find exchange rate
        if date:
            rate = ExchangeRate.objects.filter(
                from_currency=from_currency,
                to_currency=to_currency,
                date=date,
            ).first()
        else:
            # Get latest rate
            rate = ExchangeRate.objects.filter(
                from_currency=from_currency,
                to_currency=to_currency,
            ).order_by("-date").first()

        if not rate:
            return Response(
                {"detail": f"Không tìm thấy tỷ giá cho {from_currency} sang {to_currency}."},
                status=status.HTTP_404_NOT_FOUND,
            )

        converted_amount = amount * rate.rate

        return Response({
            "original_amount": amount,
            "from_currency": from_currency,
            "converted_amount": converted_amount,
            "to_currency": to_currency,
            "rate": rate.rate,
            "date": rate.date,
        })


# ============================================================
# Receipt Generation ViewSet (Phase 2.8)
# ============================================================


class ReceiptViewSet(viewsets.ViewSet):
    """
    API endpoints for Receipt generation.

    Provides:
    - Generate receipt for booking
    - Generate receipt for payment
    - Get receipt data (JSON)
    - Download receipt (PDF)
    """

    permission_classes = [IsAuthenticated, IsStaffOrManager]

    def _get_receipt_data(self, booking, payment=None, include_folio=True):
        """Generate receipt data for a booking."""
        from django.utils import timezone
        from .models import FolioItem

        # Generate receipt number
        date_str = timezone.now().strftime("%Y%m%d")
        receipt_count = booking.payments.filter(
            receipt_generated=True,
        ).count() + 1
        receipt_number = f"INV-{booking.room.number}-{date_str}-{receipt_count:03d}"

        # Get folio items
        folio_items = []
        if include_folio:
            items = FolioItem.objects.filter(
                booking=booking,
                is_voided=False,
            ).order_by("date")
            folio_items = [
                {
                    "date": item.date.isoformat(),
                    "description": item.description,
                    "quantity": item.quantity,
                    "unit_price": float(item.unit_price),
                    "total": float(item.total_price),
                }
                for item in items
            ]

        return {
            "receipt_number": receipt_number,
            "receipt_date": timezone.now().isoformat(),
            "hotel_name": "Hoàng Lâm Heritage Suites",
            "hotel_address": "123 Đường ABC, Phường XYZ, TP.HCM",
            "hotel_phone": "028 1234 5678",
            "guest_name": booking.guest.full_name,
            "guest_phone": booking.guest.phone,
            "guest_id_number": booking.guest.id_number or "",
            "room_number": booking.room.number,
            "room_type": booking.room.room_type.name,
            "check_in_date": booking.check_in_date.isoformat(),
            "check_out_date": booking.check_out_date.isoformat(),
            "nights": booking.nights,
            "room_total": float(booking.total_amount),
            "additional_charges": float(booking.additional_charges),
            "total_amount": float(booking.total_amount + booking.additional_charges),
            "deposit_paid": float(booking.deposit_amount),
            "balance_due": float(booking.balance_due),
            "folio_items": folio_items,
            "payment_method": booking.get_payment_method_display(),
            "created_by": self.request.user.get_full_name() or self.request.user.username,
        }

    @extend_schema(
        summary="Generate receipt data",
        description="Generate receipt data for a booking (JSON format).",
        request=ReceiptGenerateSerializer,
        responses={200: ReceiptDataSerializer},
        tags=["Receipts"],
    )
    @action(detail=False, methods=["post"], url_path="generate")
    def generate(self, request):
        """Generate receipt data."""
        from .serializers import ReceiptGenerateSerializer

        serializer = ReceiptGenerateSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        booking_id = serializer.validated_data.get("booking_id")
        payment_id = serializer.validated_data.get("payment_id")
        include_folio = serializer.validated_data.get("include_folio", True)

        if payment_id:
            from .models import Payment

            payment = Payment.objects.select_related(
                "booking__room__room_type",
                "booking__guest",
            ).get(id=payment_id)
            booking = payment.booking
        else:
            booking = Booking.objects.select_related(
                "room__room_type",
                "guest",
            ).get(id=booking_id)
            payment = None

        receipt_data = self._get_receipt_data(booking, payment, include_folio)
        return Response(receipt_data)

    @extend_schema(
        summary="Download receipt PDF",
        description="Download receipt as PDF file.",
        parameters=[
            OpenApiParameter("booking_id", OpenApiTypes.INT, description="Booking ID"),
        ],
        tags=["Receipts"],
    )
    @action(detail=False, methods=["get"], url_path="download/(?P<booking_id>[^/.]+)")
    def download(self, request, booking_id=None):
        """Download receipt as PDF."""
        from django.http import HttpResponse

        try:
            booking = Booking.objects.select_related(
                "room__room_type",
                "guest",
            ).get(id=booking_id)
        except Booking.DoesNotExist:
            return Response(
                {"detail": "Không tìm thấy đặt phòng."},
                status=status.HTTP_404_NOT_FOUND,
            )

        receipt_data = self._get_receipt_data(booking)

        # Try to generate PDF
        try:
            from reportlab.lib.pagesizes import A4
            from reportlab.pdfgen import canvas
            from reportlab.lib.units import cm
            from io import BytesIO

            buffer = BytesIO()
            p = canvas.Canvas(buffer, pagesize=A4)
            width, height = A4

            # Header
            p.setFont("Helvetica-Bold", 16)
            p.drawString(2 * cm, height - 2 * cm, receipt_data["hotel_name"])

            p.setFont("Helvetica", 10)
            p.drawString(2 * cm, height - 2.6 * cm, receipt_data["hotel_address"])
            p.drawString(2 * cm, height - 3 * cm, f"Tel: {receipt_data['hotel_phone']}")

            # Receipt info
            p.setFont("Helvetica-Bold", 14)
            p.drawString(2 * cm, height - 4 * cm, f"HÓA ĐƠN #{receipt_data['receipt_number']}")

            p.setFont("Helvetica", 10)
            y = height - 5 * cm

            # Guest info
            p.drawString(2 * cm, y, f"Khách hàng: {receipt_data['guest_name']}")
            y -= 0.5 * cm
            p.drawString(2 * cm, y, f"SĐT: {receipt_data['guest_phone']}")
            y -= 0.5 * cm
            if receipt_data["guest_id_number"]:
                p.drawString(2 * cm, y, f"CCCD/Passport: {receipt_data['guest_id_number']}")
                y -= 0.5 * cm

            # Booking info
            y -= 0.5 * cm
            p.drawString(2 * cm, y, f"Phòng: {receipt_data['room_number']} - {receipt_data['room_type']}")
            y -= 0.5 * cm
            p.drawString(2 * cm, y, f"Ngày: {receipt_data['check_in_date']} - {receipt_data['check_out_date']} ({receipt_data['nights']} đêm)")
            y -= cm

            # Financial
            p.setFont("Helvetica-Bold", 10)
            p.drawString(2 * cm, y, "Chi tiết thanh toán:")
            y -= 0.5 * cm

            p.setFont("Helvetica", 10)
            p.drawString(2 * cm, y, f"Tiền phòng: {receipt_data['room_total']:,.0f} VND")
            y -= 0.5 * cm
            if receipt_data["additional_charges"] > 0:
                p.drawString(2 * cm, y, f"Chi phí phát sinh: {receipt_data['additional_charges']:,.0f} VND")
                y -= 0.5 * cm
            p.drawString(2 * cm, y, f"Tổng cộng: {receipt_data['total_amount']:,.0f} VND")
            y -= 0.5 * cm
            p.drawString(2 * cm, y, f"Đã đặt cọc: {receipt_data['deposit_paid']:,.0f} VND")
            y -= 0.5 * cm
            p.setFont("Helvetica-Bold", 10)
            p.drawString(2 * cm, y, f"Còn lại: {receipt_data['balance_due']:,.0f} VND")

            # Footer
            p.setFont("Helvetica", 8)
            p.drawString(2 * cm, 2 * cm, f"Người lập: {receipt_data['created_by']}")
            p.drawString(2 * cm, 1.5 * cm, f"Ngày: {receipt_data['receipt_date']}")

            p.showPage()
            p.save()

            buffer.seek(0)
            response = HttpResponse(buffer.getvalue(), content_type="application/pdf")
            response["Content-Disposition"] = f'attachment; filename="receipt_{booking_id}.pdf"'
            return response

        except ImportError:
            # reportlab not installed, return JSON data instead
            return Response({
                "detail": "PDF generation not available. Returning JSON data.",
                "data": receipt_data,
            })


@extend_schema_view(
    list=extend_schema(
        summary="List housekeeping tasks",
        description="List all housekeeping tasks with filtering options.",
        parameters=[
            OpenApiParameter(
                name="room",
                type=OpenApiTypes.INT,
                description="Filter by room ID",
            ),
            OpenApiParameter(
                name="status",
                type=OpenApiTypes.STR,
                description="Filter by status (pending, in_progress, completed, verified)",
            ),
            OpenApiParameter(
                name="task_type",
                type=OpenApiTypes.STR,
                description="Filter by task type (cleaning, turndown, inspection, deep_clean, laundry)",
            ),
            OpenApiParameter(
                name="assigned_to",
                type=OpenApiTypes.INT,
                description="Filter by assigned user ID",
            ),
            OpenApiParameter(
                name="scheduled_date",
                type=OpenApiTypes.DATE,
                description="Filter by scheduled date (YYYY-MM-DD)",
            ),
            OpenApiParameter(
                name="priority",
                type=OpenApiTypes.STR,
                description="Filter by priority (low, medium, high, urgent)",
            ),
        ],
        tags=["Housekeeping"],
    ),
    retrieve=extend_schema(
        summary="Get housekeeping task details",
        description="Get detailed information about a specific housekeeping task.",
        tags=["Housekeeping"],
    ),
    create=extend_schema(
        summary="Create housekeeping task",
        description="Create a new housekeeping task.",
        tags=["Housekeeping"],
    ),
    update=extend_schema(
        summary="Update housekeeping task",
        description="Update a housekeeping task.",
        tags=["Housekeeping"],
    ),
    partial_update=extend_schema(
        summary="Partial update housekeeping task",
        description="Partially update a housekeeping task.",
        tags=["Housekeeping"],
    ),
    destroy=extend_schema(
        summary="Delete housekeeping task",
        description="Delete a housekeeping task.",
        tags=["Housekeeping"],
    ),
)
class HousekeepingTaskViewSet(viewsets.ModelViewSet):
    """ViewSet for managing housekeeping tasks."""

    permission_classes = [IsAuthenticated, IsStaffOrManager]
    queryset = HousekeepingTask.objects.all()

    def get_serializer_class(self):
        if self.action == "list":
            return HousekeepingTaskListSerializer
        elif self.action == "create":
            return HousekeepingTaskCreateSerializer
        elif self.action in ["update", "partial_update"]:
            return HousekeepingTaskUpdateSerializer
        return HousekeepingTaskSerializer

    def get_queryset(self):
        queryset = HousekeepingTask.objects.select_related(
            "room", "assigned_to", "created_by"
        ).order_by("-created_at")

        # Filter by room
        room = self.request.query_params.get("room")
        if room:
            queryset = queryset.filter(room_id=room)

        # Filter by status
        status_filter = self.request.query_params.get("status")
        if status_filter:
            queryset = queryset.filter(status=status_filter)

        # Filter by task type
        task_type = self.request.query_params.get("task_type")
        if task_type:
            queryset = queryset.filter(task_type=task_type)

        # Filter by assigned user
        assigned_to = self.request.query_params.get("assigned_to")
        if assigned_to:
            queryset = queryset.filter(assigned_to_id=assigned_to)

        # Filter by scheduled date
        scheduled_date = self.request.query_params.get("scheduled_date")
        if scheduled_date:
            queryset = queryset.filter(scheduled_date=scheduled_date)

        # Filter by priority
        priority = self.request.query_params.get("priority")
        if priority:
            queryset = queryset.filter(priority=priority)

        return queryset

    def perform_create(self, serializer):
        serializer.save(created_by=self.request.user)

    @extend_schema(
        summary="Assign housekeeping task",
        description="Assign a housekeeping task to a staff member.",
        request={
            "type": "object",
            "properties": {
                "assigned_to": {
                    "type": "integer",
                    "description": "User ID of the staff member to assign",
                },
            },
            "required": ["assigned_to"],
        },
        responses={
            200: HousekeepingTaskSerializer,
            400: OpenApiResponse(description="Invalid request"),
            404: OpenApiResponse(description="Task or user not found"),
        },
        tags=["Housekeeping"],
    )
    @action(detail=True, methods=["post"])
    def assign(self, request, pk=None):
        """Assign the task to a staff member."""
        task = self.get_object()
        assigned_to_id = request.data.get("assigned_to")

        if not assigned_to_id:
            return Response(
                {"detail": "assigned_to is required."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        try:
            user = User.objects.get(pk=assigned_to_id)
        except User.DoesNotExist:
            return Response(
                {"detail": "User not found."},
                status=status.HTTP_404_NOT_FOUND,
            )

        task.assigned_to = user
        if task.status == "pending":
            task.status = "in_progress"
        task.save()

        serializer = HousekeepingTaskSerializer(task)
        return Response(serializer.data)

    @extend_schema(
        summary="Complete housekeeping task",
        description="Mark a housekeeping task as completed.",
        request={
            "type": "object",
            "properties": {
                "notes": {
                    "type": "string",
                    "description": "Optional completion notes",
                },
            },
        },
        responses={
            200: HousekeepingTaskSerializer,
            400: OpenApiResponse(description="Task already completed or verified"),
        },
        tags=["Housekeeping"],
    )
    @action(detail=True, methods=["post"])
    def complete(self, request, pk=None):
        """Mark the task as completed."""
        from django.utils import timezone

        task = self.get_object()

        if task.status in ["completed", "verified"]:
            return Response(
                {"detail": f"Cannot complete task with status '{task.status}'."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        task.status = "completed"
        task.completed_at = timezone.now()
        if request.data.get("notes"):
            task.notes = request.data.get("notes")
        task.save()

        # Update room status to available if it was a cleaning task
        if task.task_type in ["checkout_clean", "stay_clean", "deep_clean"] and task.room:
            task.room.status = Room.Status.AVAILABLE
            task.room.save()

        serializer = HousekeepingTaskSerializer(task)
        return Response(serializer.data)

    @extend_schema(
        summary="Verify housekeeping task",
        description="Verify a completed housekeeping task (manager only).",
        request={
            "type": "object",
            "properties": {
                "notes": {
                    "type": "string",
                    "description": "Optional verification notes",
                },
            },
        },
        responses={
            200: HousekeepingTaskSerializer,
            400: OpenApiResponse(description="Task not completed"),
        },
        tags=["Housekeeping"],
    )
    @action(detail=True, methods=["post"])
    def verify(self, request, pk=None):
        """Verify the completed task."""
        task = self.get_object()

        if task.status != "completed":
            return Response(
                {"detail": "Only completed tasks can be verified."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        task.status = "verified"
        if request.data.get("notes"):
            task.notes = f"{task.notes}\nVerified: {request.data.get('notes')}" if task.notes else f"Verified: {request.data.get('notes')}"
        task.save()

        serializer = HousekeepingTaskSerializer(task)
        return Response(serializer.data)

    @extend_schema(
        summary="Get tasks for today",
        description="Get all housekeeping tasks scheduled for today.",
        responses={
            200: HousekeepingTaskListSerializer(many=True),
        },
        tags=["Housekeeping"],
    )
    @action(detail=False, methods=["get"])
    def today(self, request):
        """Get tasks scheduled for today."""
        from django.utils import timezone

        today = timezone.now().date()
        queryset = self.get_queryset().filter(scheduled_date=today)
        serializer = HousekeepingTaskListSerializer(queryset, many=True)
        return Response(serializer.data)

    @extend_schema(
        summary="Get my tasks",
        description="Get housekeeping tasks assigned to the current user.",
        responses={
            200: HousekeepingTaskListSerializer(many=True),
        },
        tags=["Housekeeping"],
    )
    @action(detail=False, methods=["get"])
    def my_tasks(self, request):
        """Get tasks assigned to the current user."""
        queryset = self.get_queryset().filter(
            assigned_to=request.user,
            status__in=["pending", "in_progress"],
        )
        serializer = HousekeepingTaskListSerializer(queryset, many=True)
        return Response(serializer.data)


@extend_schema_view(
    list=extend_schema(
        summary="List maintenance requests",
        description="List all maintenance requests with filtering options.",
        parameters=[
            OpenApiParameter(
                name="room",
                type=OpenApiTypes.INT,
                description="Filter by room ID",
            ),
            OpenApiParameter(
                name="status",
                type=OpenApiTypes.STR,
                description="Filter by status (pending, assigned, in_progress, on_hold, completed, cancelled)",
            ),
            OpenApiParameter(
                name="priority",
                type=OpenApiTypes.STR,
                description="Filter by priority (low, medium, high, urgent)",
            ),
            OpenApiParameter(
                name="category",
                type=OpenApiTypes.STR,
                description="Filter by category (electrical, plumbing, ac_heating, furniture, appliance, structural, safety, other)",
            ),
            OpenApiParameter(
                name="assigned_to",
                type=OpenApiTypes.INT,
                description="Filter by assigned user ID",
            ),
        ],
        tags=["Maintenance"],
    ),
    retrieve=extend_schema(
        summary="Get maintenance request details",
        description="Get detailed information about a specific maintenance request.",
        tags=["Maintenance"],
    ),
    create=extend_schema(
        summary="Create maintenance request",
        description="Create a new maintenance request.",
        tags=["Maintenance"],
    ),
    update=extend_schema(
        summary="Update maintenance request",
        description="Update a maintenance request.",
        tags=["Maintenance"],
    ),
    partial_update=extend_schema(
        summary="Partial update maintenance request",
        description="Partially update a maintenance request.",
        tags=["Maintenance"],
    ),
    destroy=extend_schema(
        summary="Delete maintenance request",
        description="Delete a maintenance request.",
        tags=["Maintenance"],
    ),
)
class MaintenanceRequestViewSet(viewsets.ModelViewSet):
    """ViewSet for managing maintenance requests."""

    permission_classes = [IsAuthenticated, IsStaffOrManager]
    queryset = MaintenanceRequest.objects.all()

    def get_serializer_class(self):
        if self.action == "list":
            return MaintenanceRequestListSerializer
        elif self.action == "create":
            return MaintenanceRequestCreateSerializer
        elif self.action in ["update", "partial_update"]:
            return MaintenanceRequestUpdateSerializer
        return MaintenanceRequestSerializer

    def get_queryset(self):
        queryset = MaintenanceRequest.objects.select_related(
            "room", "reported_by", "assigned_to", "completed_by"
        ).order_by("-created_at")

        # Filter by room
        room = self.request.query_params.get("room")
        if room:
            queryset = queryset.filter(room_id=room)

        # Filter by status
        status_filter = self.request.query_params.get("status")
        if status_filter:
            queryset = queryset.filter(status=status_filter)

        # Filter by priority
        priority = self.request.query_params.get("priority")
        if priority:
            queryset = queryset.filter(priority=priority)

        # Filter by category
        category = self.request.query_params.get("category")
        if category:
            queryset = queryset.filter(category=category)

        # Filter by assigned user
        assigned_to = self.request.query_params.get("assigned_to")
        if assigned_to:
            queryset = queryset.filter(assigned_to_id=assigned_to)

        return queryset

    def perform_create(self, serializer):
        serializer.save(reported_by=self.request.user)

    @extend_schema(
        summary="Assign maintenance request",
        description="Assign a maintenance request to a staff member.",
        request={
            "type": "object",
            "properties": {
                "assigned_to": {
                    "type": "integer",
                    "description": "User ID of the staff member to assign",
                },
                "estimated_cost": {
                    "type": "number",
                    "description": "Estimated cost for the repair",
                },
            },
            "required": ["assigned_to"],
        },
        responses={
            200: MaintenanceRequestSerializer,
            400: OpenApiResponse(description="Invalid request"),
            404: OpenApiResponse(description="Request or user not found"),
        },
        tags=["Maintenance"],
    )
    @action(detail=True, methods=["post"])
    def assign(self, request, pk=None):
        """Assign the request to a staff member."""
        maintenance_request = self.get_object()
        assigned_to_id = request.data.get("assigned_to")

        if not assigned_to_id:
            return Response(
                {"detail": "assigned_to is required."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        try:
            user = User.objects.get(pk=assigned_to_id)
        except User.DoesNotExist:
            return Response(
                {"detail": "User not found."},
                status=status.HTTP_404_NOT_FOUND,
            )

        maintenance_request.assign(user)

        # Update estimated cost if provided
        if request.data.get("estimated_cost"):
            maintenance_request.estimated_cost = request.data.get("estimated_cost")
            maintenance_request.save()

        serializer = MaintenanceRequestSerializer(maintenance_request)
        return Response(serializer.data)

    @extend_schema(
        summary="Complete maintenance request",
        description="Mark a maintenance request as completed.",
        request={
            "type": "object",
            "properties": {
                "actual_cost": {
                    "type": "number",
                    "description": "Actual cost of the repair",
                },
                "resolution_notes": {
                    "type": "string",
                    "description": "Notes about the resolution",
                },
            },
        },
        responses={
            200: MaintenanceRequestSerializer,
            400: OpenApiResponse(description="Request not in progress"),
        },
        tags=["Maintenance"],
    )
    @action(detail=True, methods=["post"])
    def complete(self, request, pk=None):
        """Mark the request as completed."""
        maintenance_request = self.get_object()

        if maintenance_request.status not in ["assigned", "in_progress"]:
            return Response(
                {"detail": f"Cannot complete request with status '{maintenance_request.status}'."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        actual_cost = request.data.get("actual_cost")
        resolution_notes = request.data.get("resolution_notes", "")

        # Update actual cost if provided
        if actual_cost is not None:
            maintenance_request.actual_cost = actual_cost
            maintenance_request.save()

        # Use the model's complete method
        maintenance_request.complete(request.user, resolution_notes)

        serializer = MaintenanceRequestSerializer(maintenance_request)
        return Response(serializer.data)

    @extend_schema(
        summary="Put maintenance request on hold",
        description="Put a maintenance request on hold.",
        request={
            "type": "object",
            "properties": {
                "reason": {
                    "type": "string",
                    "description": "Reason for putting on hold",
                },
            },
        },
        responses={
            200: MaintenanceRequestSerializer,
            400: OpenApiResponse(description="Invalid status transition"),
        },
        tags=["Maintenance"],
    )
    @action(detail=True, methods=["post"])
    def hold(self, request, pk=None):
        """Put the request on hold."""
        maintenance_request = self.get_object()

        if maintenance_request.status in ["completed", "cancelled"]:
            return Response(
                {"detail": f"Cannot hold request with status '{maintenance_request.status}'."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        maintenance_request.status = "on_hold"
        if request.data.get("reason"):
            maintenance_request.resolution_notes = f"{maintenance_request.resolution_notes}\nOn hold: {request.data.get('reason')}" if maintenance_request.resolution_notes else f"On hold: {request.data.get('reason')}"
        maintenance_request.save()

        serializer = MaintenanceRequestSerializer(maintenance_request)
        return Response(serializer.data)

    @extend_schema(
        summary="Resume maintenance request",
        description="Resume a maintenance request from on hold.",
        responses={
            200: MaintenanceRequestSerializer,
            400: OpenApiResponse(description="Request not on hold"),
        },
        tags=["Maintenance"],
    )
    @action(detail=True, methods=["post"])
    def resume(self, request, pk=None):
        """Resume the request from on hold."""
        maintenance_request = self.get_object()

        if maintenance_request.status != "on_hold":
            return Response(
                {"detail": "Request is not on hold."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        maintenance_request.status = "in_progress" if maintenance_request.assigned_to else "assigned"
        maintenance_request.save()

        serializer = MaintenanceRequestSerializer(maintenance_request)
        return Response(serializer.data)

    @extend_schema(
        summary="Cancel maintenance request",
        description="Cancel a maintenance request.",
        request={
            "type": "object",
            "properties": {
                "reason": {
                    "type": "string",
                    "description": "Reason for cancellation",
                },
            },
        },
        responses={
            200: MaintenanceRequestSerializer,
            400: OpenApiResponse(description="Request already completed"),
        },
        tags=["Maintenance"],
    )
    @action(detail=True, methods=["post"])
    def cancel(self, request, pk=None):
        """Cancel the request."""
        maintenance_request = self.get_object()

        if maintenance_request.status == "completed":
            return Response(
                {"detail": "Cannot cancel completed request."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        maintenance_request.status = "cancelled"
        if request.data.get("reason"):
            maintenance_request.resolution_notes = f"{maintenance_request.resolution_notes}\nCancelled: {request.data.get('reason')}" if maintenance_request.resolution_notes else f"Cancelled: {request.data.get('reason')}"
        maintenance_request.save()

        serializer = MaintenanceRequestSerializer(maintenance_request)
        return Response(serializer.data)

    @extend_schema(
        summary="Get urgent requests",
        description="Get all urgent and high priority maintenance requests that are not completed.",
        responses={
            200: MaintenanceRequestListSerializer(many=True),
        },
        tags=["Maintenance"],
    )
    @action(detail=False, methods=["get"])
    def urgent(self, request):
        """Get urgent and high priority requests."""
        queryset = self.get_queryset().filter(
            priority__in=["urgent", "high"],
            status__in=["pending", "assigned", "in_progress", "on_hold"],
        )
        serializer = MaintenanceRequestListSerializer(queryset, many=True)
        return Response(serializer.data)

    @extend_schema(
        summary="Get my requests",
        description="Get maintenance requests assigned to the current user.",
        responses={
            200: MaintenanceRequestListSerializer(many=True),
        },
        tags=["Maintenance"],
    )
    @action(detail=False, methods=["get"])
    def my_requests(self, request):
        """Get requests assigned to the current user."""
        queryset = self.get_queryset().filter(
            assigned_to=request.user,
            status__in=["assigned", "in_progress", "on_hold"],
        )
        serializer = MaintenanceRequestListSerializer(queryset, many=True)
        return Response(serializer.data)


# ============================================================
# Minibar ViewSets (Phase 3.4)
# ============================================================


class MinibarItemViewSet(viewsets.ModelViewSet):
    """
    ViewSet for managing minibar items.

    Provides CRUD operations for minibar inventory items.
    """

    queryset = MinibarItem.objects.all()
    permission_classes = [IsAuthenticated]

    def get_serializer_class(self):
        if self.action == "create":
            return MinibarItemCreateSerializer
        elif self.action in ["update", "partial_update"]:
            return MinibarItemUpdateSerializer
        elif self.action == "list":
            return MinibarItemListSerializer
        return MinibarItemSerializer

    def get_queryset(self):
        queryset = MinibarItem.objects.all()

        # Filter by active status
        is_active = self.request.query_params.get("is_active")
        if is_active is not None:
            queryset = queryset.filter(is_active=is_active.lower() == "true")

        # Filter by category
        category = self.request.query_params.get("category")
        if category:
            queryset = queryset.filter(category__iexact=category)

        # Search by name
        search = self.request.query_params.get("search")
        if search:
            queryset = queryset.filter(name__icontains=search)

        return queryset.order_by("category", "name")

    @extend_schema(
        summary="List minibar items",
        description="Get a list of all minibar items with optional filtering.",
        parameters=[
            {
                "name": "is_active",
                "in": "query",
                "description": "Filter by active status",
                "required": False,
                "schema": {"type": "boolean"},
            },
            {
                "name": "category",
                "in": "query",
                "description": "Filter by category",
                "required": False,
                "schema": {"type": "string"},
            },
            {
                "name": "search",
                "in": "query",
                "description": "Search by name",
                "required": False,
                "schema": {"type": "string"},
            },
        ],
        responses={200: MinibarItemListSerializer(many=True)},
        tags=["Minibar"],
    )
    def list(self, request, *args, **kwargs):
        return super().list(request, *args, **kwargs)

    @extend_schema(
        summary="Create minibar item",
        description="Create a new minibar item.",
        request=MinibarItemCreateSerializer,
        responses={201: MinibarItemSerializer},
        tags=["Minibar"],
    )
    def create(self, request, *args, **kwargs):
        return super().create(request, *args, **kwargs)

    @extend_schema(
        summary="Get minibar item",
        description="Get details of a specific minibar item.",
        responses={200: MinibarItemSerializer},
        tags=["Minibar"],
    )
    def retrieve(self, request, *args, **kwargs):
        return super().retrieve(request, *args, **kwargs)

    @extend_schema(
        summary="Update minibar item",
        description="Update a minibar item.",
        request=MinibarItemUpdateSerializer,
        responses={200: MinibarItemSerializer},
        tags=["Minibar"],
    )
    def update(self, request, *args, **kwargs):
        return super().update(request, *args, **kwargs)

    @extend_schema(
        summary="Partial update minibar item",
        description="Partially update a minibar item.",
        request=MinibarItemUpdateSerializer,
        responses={200: MinibarItemSerializer},
        tags=["Minibar"],
    )
    def partial_update(self, request, *args, **kwargs):
        return super().partial_update(request, *args, **kwargs)

    @extend_schema(
        summary="Delete minibar item",
        description="Delete a minibar item.",
        responses={204: None},
        tags=["Minibar"],
    )
    def destroy(self, request, *args, **kwargs):
        return super().destroy(request, *args, **kwargs)

    @extend_schema(
        summary="Toggle item active status",
        description="Toggle the active status of a minibar item.",
        responses={200: MinibarItemSerializer},
        tags=["Minibar"],
    )
    @action(detail=True, methods=["post"])
    def toggle_active(self, request, pk=None):
        """Toggle the active status of a minibar item."""
        item = self.get_object()
        item.is_active = not item.is_active
        item.save()
        serializer = MinibarItemSerializer(item)
        return Response(serializer.data)

    @extend_schema(
        summary="Get active items",
        description="Get all active minibar items for POS display.",
        responses={200: MinibarItemListSerializer(many=True)},
        tags=["Minibar"],
    )
    @action(detail=False, methods=["get"])
    def active(self, request):
        """Get all active minibar items."""
        queryset = MinibarItem.objects.filter(is_active=True).order_by("category", "name")
        serializer = MinibarItemListSerializer(queryset, many=True)
        return Response(serializer.data)

    @extend_schema(
        summary="Get categories",
        description="Get list of distinct minibar item categories.",
        responses={200: {"type": "array", "items": {"type": "string"}}},
        tags=["Minibar"],
    )
    @action(detail=False, methods=["get"])
    def categories(self, request):
        """Get distinct minibar item categories."""
        categories = (
            MinibarItem.objects.filter(is_active=True)
            .values_list("category", flat=True)
            .distinct()
            .order_by("category")
        )
        return Response(list(filter(None, categories)))


class MinibarSaleViewSet(viewsets.ModelViewSet):
    """
    ViewSet for managing minibar sales.

    Provides CRUD operations for minibar sales/charges.
    """

    queryset = MinibarSale.objects.all()
    permission_classes = [IsAuthenticated]

    def get_serializer_class(self):
        if self.action == "create":
            return MinibarSaleCreateSerializer
        elif self.action in ["update", "partial_update"]:
            return MinibarSaleUpdateSerializer
        elif self.action == "list":
            return MinibarSaleListSerializer
        elif self.action == "bulk_create":
            return MinibarSaleBulkCreateSerializer
        return MinibarSaleSerializer

    def get_queryset(self):
        queryset = MinibarSale.objects.select_related(
            "booking", "booking__room", "item", "created_by"
        )

        # Filter by booking
        booking_id = self.request.query_params.get("booking")
        if booking_id:
            queryset = queryset.filter(booking_id=booking_id)

        # Filter by room
        room_id = self.request.query_params.get("room")
        if room_id:
            queryset = queryset.filter(booking__room_id=room_id)

        # Filter by date range
        date_from = self.request.query_params.get("date_from")
        date_to = self.request.query_params.get("date_to")
        if date_from:
            queryset = queryset.filter(date__gte=date_from)
        if date_to:
            queryset = queryset.filter(date__lte=date_to)

        # Filter by charged status
        is_charged = self.request.query_params.get("is_charged")
        if is_charged is not None:
            queryset = queryset.filter(is_charged=is_charged.lower() == "true")

        return queryset.order_by("-date", "-created_at")

    @extend_schema(
        summary="List minibar sales",
        description="Get a list of minibar sales with optional filtering.",
        parameters=[
            {
                "name": "booking",
                "in": "query",
                "description": "Filter by booking ID",
                "required": False,
                "schema": {"type": "integer"},
            },
            {
                "name": "room",
                "in": "query",
                "description": "Filter by room ID",
                "required": False,
                "schema": {"type": "integer"},
            },
            {
                "name": "date_from",
                "in": "query",
                "description": "Filter sales from this date",
                "required": False,
                "schema": {"type": "string", "format": "date"},
            },
            {
                "name": "date_to",
                "in": "query",
                "description": "Filter sales up to this date",
                "required": False,
                "schema": {"type": "string", "format": "date"},
            },
            {
                "name": "is_charged",
                "in": "query",
                "description": "Filter by charged status",
                "required": False,
                "schema": {"type": "boolean"},
            },
        ],
        responses={200: MinibarSaleListSerializer(many=True)},
        tags=["Minibar"],
    )
    def list(self, request, *args, **kwargs):
        return super().list(request, *args, **kwargs)

    @extend_schema(
        summary="Create minibar sale",
        description="Create a new minibar sale/charge.",
        request=MinibarSaleCreateSerializer,
        responses={201: MinibarSaleSerializer},
        tags=["Minibar"],
    )
    def create(self, request, *args, **kwargs):
        serializer = MinibarSaleCreateSerializer(data=request.data, context={"request": request})
        serializer.is_valid(raise_exception=True)
        instance = serializer.save()
        response_serializer = MinibarSaleSerializer(instance)
        return Response(response_serializer.data, status=status.HTTP_201_CREATED)

    @extend_schema(
        summary="Get minibar sale",
        description="Get details of a specific minibar sale.",
        responses={200: MinibarSaleSerializer},
        tags=["Minibar"],
    )
    def retrieve(self, request, *args, **kwargs):
        return super().retrieve(request, *args, **kwargs)

    @extend_schema(
        summary="Update minibar sale",
        description="Update a minibar sale.",
        request=MinibarSaleUpdateSerializer,
        responses={200: MinibarSaleSerializer},
        tags=["Minibar"],
    )
    def update(self, request, *args, **kwargs):
        instance = self.get_object()
        serializer = MinibarSaleUpdateSerializer(instance, data=request.data, context={"request": request})
        serializer.is_valid(raise_exception=True)
        instance = serializer.save()
        response_serializer = MinibarSaleSerializer(instance)
        return Response(response_serializer.data)

    @extend_schema(
        summary="Partial update minibar sale",
        description="Partially update a minibar sale.",
        request=MinibarSaleUpdateSerializer,
        responses={200: MinibarSaleSerializer},
        tags=["Minibar"],
    )
    def partial_update(self, request, *args, **kwargs):
        instance = self.get_object()
        serializer = MinibarSaleUpdateSerializer(instance, data=request.data, partial=True, context={"request": request})
        serializer.is_valid(raise_exception=True)
        instance = serializer.save()
        response_serializer = MinibarSaleSerializer(instance)
        return Response(response_serializer.data)

    @extend_schema(
        summary="Delete minibar sale",
        description="Delete a minibar sale.",
        responses={204: None},
        tags=["Minibar"],
    )
    def destroy(self, request, *args, **kwargs):
        return super().destroy(request, *args, **kwargs)

    @extend_schema(
        summary="Bulk create minibar sales",
        description="Create multiple minibar sales at once for a booking.",
        request=MinibarSaleBulkCreateSerializer,
        responses={201: MinibarSaleSerializer(many=True)},
        tags=["Minibar"],
    )
    @action(detail=False, methods=["post"])
    def bulk_create(self, request):
        """Create multiple minibar sales at once."""
        serializer = MinibarSaleBulkCreateSerializer(
            data=request.data, context={"request": request}
        )
        serializer.is_valid(raise_exception=True)
        sales = serializer.save()
        response_serializer = MinibarSaleSerializer(sales, many=True)
        return Response(response_serializer.data, status=status.HTTP_201_CREATED)

    @extend_schema(
        summary="Mark as charged",
        description="Mark a minibar sale as charged to the room folio.",
        responses={200: MinibarSaleSerializer},
        tags=["Minibar"],
    )
    @action(detail=True, methods=["post"])
    def mark_charged(self, request, pk=None):
        """Mark a minibar sale as charged."""
        sale = self.get_object()
        if sale.is_charged:
            return Response(
                {"detail": "Sale đã được tính phí."},
                status=status.HTTP_400_BAD_REQUEST,
            )
        sale.is_charged = True
        sale.save()
        serializer = MinibarSaleSerializer(sale)
        return Response(serializer.data)

    @extend_schema(
        summary="Unmark as charged",
        description="Unmark a minibar sale as charged (reverse charge).",
        responses={200: MinibarSaleSerializer},
        tags=["Minibar"],
    )
    @action(detail=True, methods=["post"])
    def unmark_charged(self, request, pk=None):
        """Unmark a minibar sale as charged."""
        sale = self.get_object()
        if not sale.is_charged:
            return Response(
                {"detail": "Sale chưa được tính phí."},
                status=status.HTTP_400_BAD_REQUEST,
            )
        sale.is_charged = False
        sale.save()
        serializer = MinibarSaleSerializer(sale)
        return Response(serializer.data)

    @extend_schema(
        summary="Get uncharged sales for booking",
        description="Get all uncharged minibar sales for a specific booking.",
        parameters=[
            {
                "name": "booking",
                "in": "query",
                "description": "Booking ID",
                "required": True,
                "schema": {"type": "integer"},
            },
        ],
        responses={200: MinibarSaleListSerializer(many=True)},
        tags=["Minibar"],
    )
    @action(detail=False, methods=["get"])
    def uncharged(self, request):
        """Get uncharged sales for a booking."""
        booking_id = request.query_params.get("booking")
        if not booking_id:
            return Response(
                {"detail": "booking parameter is required."},
                status=status.HTTP_400_BAD_REQUEST,
            )
        queryset = MinibarSale.objects.filter(
            booking_id=booking_id, is_charged=False
        ).order_by("-date")
        serializer = MinibarSaleListSerializer(queryset, many=True)
        return Response(serializer.data)

    @extend_schema(
        summary="Charge all to room",
        description="Charge all uncharged minibar sales for a booking to the room folio.",
        request={"type": "object", "properties": {"booking": {"type": "integer"}}},
        responses={
            200: {
                "type": "object",
                "properties": {
                    "charged_count": {"type": "integer"},
                    "total_amount": {"type": "number"},
                },
            }
        },
        tags=["Minibar"],
    )
    @action(detail=False, methods=["post"])
    def charge_all(self, request):
        """Charge all uncharged sales for a booking."""
        booking_id = request.data.get("booking")
        if not booking_id:
            return Response(
                {"detail": "booking parameter is required."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        sales = MinibarSale.objects.filter(booking_id=booking_id, is_charged=False)
        total = sum(sale.total for sale in sales)
        count = sales.count()

        sales.update(is_charged=True)

        return Response(
            {
                "charged_count": count,
                "total_amount": total,
            }
        )

    @extend_schema(
        summary="Get sales summary for booking",
        description="Get a summary of all minibar sales for a booking.",
        parameters=[
            {
                "name": "booking",
                "in": "query",
                "description": "Booking ID",
                "required": True,
                "schema": {"type": "integer"},
            },
        ],
        responses={
            200: {
                "type": "object",
                "properties": {
                    "total_sales": {"type": "integer"},
                    "total_amount": {"type": "number"},
                    "charged_amount": {"type": "number"},
                    "uncharged_amount": {"type": "number"},
                    "items": {"type": "array"},
                },
            }
        },
        tags=["Minibar"],
    )
    @action(detail=False, methods=["get"])
    def summary(self, request):
        """Get minibar sales summary for a booking."""
        booking_id = request.query_params.get("booking")
        if not booking_id:
            return Response(
                {"detail": "booking parameter is required."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        sales = MinibarSale.objects.filter(booking_id=booking_id)
        total_amount = sum(sale.total for sale in sales)
        charged_amount = sum(sale.total for sale in sales if sale.is_charged)
        uncharged_amount = sum(sale.total for sale in sales if not sale.is_charged)

        # Group by item
        from django.db.models import Sum

        items_summary = (
            sales.values("item__name")
            .annotate(total_quantity=Sum("quantity"), total_amount=Sum("total"))
            .order_by("item__name")
        )

        return Response(
            {
                "total_sales": sales.count(),
                "total_amount": total_amount,
                "charged_amount": charged_amount,
                "uncharged_amount": uncharged_amount,
                "items": list(items_summary),
            }
        )


# ============================================================================
# PHASE 4: REPORT VIEWS
# ============================================================================


class OccupancyReportView(APIView):
    """
    Occupancy report endpoint.
    Returns daily/weekly/monthly occupancy data with revenue.
    """
    
    permission_classes = [IsAuthenticated]
    
    @extend_schema(
        summary="Get occupancy report",
        description="Get occupancy statistics for a date range with optional grouping.",
        parameters=[
            OpenApiParameter("start_date", OpenApiTypes.DATE, required=True),
            OpenApiParameter("end_date", OpenApiTypes.DATE, required=True),
            OpenApiParameter("group_by", OpenApiTypes.STR, enum=["day", "week", "month"]),
            OpenApiParameter("room_type", OpenApiTypes.INT),
        ],
        responses={
            200: OpenApiResponse(
                description="Occupancy report data",
                response={
                    "type": "object",
                    "properties": {
                        "summary": {"type": "object"},
                        "data": {"type": "array"},
                    },
                },
            ),
        },
        tags=["Reports"],
    )
    def get(self, request):
        from datetime import timedelta
        from django.db.models import Count, Sum, Q
        from django.db.models.functions import TruncDate, TruncWeek, TruncMonth
        
        serializer = OccupancyReportRequestSerializer(data=request.query_params)
        serializer.is_valid(raise_exception=True)
        
        start_date = serializer.validated_data["start_date"]
        end_date = serializer.validated_data["end_date"]
        group_by = serializer.validated_data.get("group_by", "day")
        room_type_id = serializer.validated_data.get("room_type")
        
        # Get total active rooms (optionally filtered by room type)
        rooms_query = Room.objects.filter(is_active=True)
        if room_type_id:
            rooms_query = rooms_query.filter(room_type_id=room_type_id)
        total_rooms = rooms_query.count()
        
        if total_rooms == 0:
            return Response({
                "summary": {
                    "total_rooms": 0,
                    "total_room_nights": 0,
                    "occupied_nights": 0,
                    "average_occupancy": 0,
                    "total_revenue": 0,
                },
                "data": [],
            })
        
        # Calculate occupancy for each day in range
        data = []
        current_date = start_date
        
        # Get all bookings that overlap with date range
        bookings_query = Booking.objects.filter(
            status__in=["confirmed", "checked_in", "checked_out"],
            check_in_date__lte=end_date,
            check_out_date__gt=start_date,
        )
        if room_type_id:
            bookings_query = bookings_query.filter(room__room_type_id=room_type_id)
        
        daily_data = {}
        while current_date <= end_date:
            # Count rooms occupied on this date
            occupied = bookings_query.filter(
                check_in_date__lte=current_date,
                check_out_date__gt=current_date,
            ).count()
            
            # Get revenue for this date (from bookings that include this night)
            day_bookings = bookings_query.filter(
                check_in_date__lte=current_date,
                check_out_date__gt=current_date,
            )
            day_revenue = sum(b.nightly_rate for b in day_bookings)
            
            daily_data[current_date] = {
                "date": current_date,
                "total_rooms": total_rooms,
                "occupied_rooms": occupied,
                "available_rooms": total_rooms - occupied,
                "occupancy_rate": round((occupied / total_rooms) * 100, 2) if total_rooms > 0 else 0,
                "revenue": day_revenue,
            }
            current_date += timedelta(days=1)
        
        # Group data if needed
        if group_by == "day":
            data = list(daily_data.values())
        elif group_by == "week":
            from collections import defaultdict
            weekly = defaultdict(lambda: {"occupied_sum": 0, "revenue_sum": 0, "days": 0})
            for date, day_data in daily_data.items():
                week_start = date - timedelta(days=date.weekday())
                weekly[week_start]["occupied_sum"] += day_data["occupied_rooms"]
                weekly[week_start]["revenue_sum"] += day_data["revenue"]
                weekly[week_start]["days"] += 1
            
            for week_start, week_data in sorted(weekly.items()):
                avg_occupied = week_data["occupied_sum"] / week_data["days"]
                data.append({
                    "period": f"Week of {week_start.isoformat()}",
                    "date": week_start,
                    "total_rooms": total_rooms,
                    "occupied_rooms": round(avg_occupied, 1),
                    "available_rooms": round(total_rooms - avg_occupied, 1),
                    "occupancy_rate": round((avg_occupied / total_rooms) * 100, 2) if total_rooms > 0 else 0,
                    "revenue": week_data["revenue_sum"],
                })
        elif group_by == "month":
            from collections import defaultdict
            monthly = defaultdict(lambda: {"occupied_sum": 0, "revenue_sum": 0, "days": 0})
            for date, day_data in daily_data.items():
                month_key = date.replace(day=1)
                monthly[month_key]["occupied_sum"] += day_data["occupied_rooms"]
                monthly[month_key]["revenue_sum"] += day_data["revenue"]
                monthly[month_key]["days"] += 1
            
            for month_start, month_data in sorted(monthly.items()):
                avg_occupied = month_data["occupied_sum"] / month_data["days"]
                data.append({
                    "period": month_start.strftime("%Y-%m"),
                    "date": month_start,
                    "total_rooms": total_rooms,
                    "occupied_rooms": round(avg_occupied, 1),
                    "available_rooms": round(total_rooms - avg_occupied, 1),
                    "occupancy_rate": round((avg_occupied / total_rooms) * 100, 2) if total_rooms > 0 else 0,
                    "revenue": month_data["revenue_sum"],
                })
        
        # Calculate summary
        total_days = (end_date - start_date).days + 1
        total_room_nights = total_rooms * total_days
        occupied_nights = sum(d["occupied_rooms"] for d in daily_data.values())
        total_revenue = sum(d["revenue"] for d in daily_data.values())
        
        summary = {
            "total_rooms": total_rooms,
            "total_room_nights": total_room_nights,
            "occupied_nights": round(occupied_nights),
            "average_occupancy": round((occupied_nights / total_room_nights) * 100, 2) if total_room_nights > 0 else 0,
            "total_revenue": total_revenue,
        }
        
        return Response({
            "summary": summary,
            "data": data,
        })


class RevenueReportView(APIView):
    """
    Revenue report endpoint.
    Returns revenue breakdown by source with expenses and profit.
    """
    
    permission_classes = [IsAuthenticated]
    
    @extend_schema(
        summary="Get revenue report",
        description="Get revenue and expense data for a date range.",
        parameters=[
            OpenApiParameter("start_date", OpenApiTypes.DATE, required=True),
            OpenApiParameter("end_date", OpenApiTypes.DATE, required=True),
            OpenApiParameter("group_by", OpenApiTypes.STR, enum=["day", "week", "month"]),
            OpenApiParameter("category", OpenApiTypes.INT),
        ],
        responses={200: OpenApiResponse(description="Revenue report data")},
        tags=["Reports"],
    )
    def get(self, request):
        from datetime import timedelta
        from collections import defaultdict
        from django.db.models import Sum
        
        serializer = RevenueReportRequestSerializer(data=request.query_params)
        serializer.is_valid(raise_exception=True)
        
        start_date = serializer.validated_data["start_date"]
        end_date = serializer.validated_data["end_date"]
        group_by = serializer.validated_data.get("group_by", "day")
        category_id = serializer.validated_data.get("category")
        
        # Get financial entries
        income_query = FinancialEntry.objects.filter(
            entry_type="income",
            date__gte=start_date,
            date__lte=end_date,
        )
        expense_query = FinancialEntry.objects.filter(
            entry_type="expense",
            date__gte=start_date,
            date__lte=end_date,
        )
        
        if category_id:
            income_query = income_query.filter(category_id=category_id)
            expense_query = expense_query.filter(category_id=category_id)
        
        # Get minibar sales
        minibar_query = MinibarSale.objects.filter(
            date__gte=start_date,
            date__lte=end_date,
        )
        
        # Build daily data
        daily_data = {}
        current_date = start_date
        
        while current_date <= end_date:
            # Room revenue (from financial entries linked to bookings)
            day_room_revenue = income_query.filter(
                date=current_date,
                booking__isnull=False,
            ).aggregate(total=Sum("amount"))["total"] or 0
            
            # Additional revenue (income not linked to bookings)
            day_additional = income_query.filter(
                date=current_date,
                booking__isnull=True,
            ).aggregate(total=Sum("amount"))["total"] or 0
            
            # Minibar revenue
            day_minibar = minibar_query.filter(date=current_date).aggregate(
                total=Sum("total")
            )["total"] or 0
            
            # Expenses
            day_expenses = expense_query.filter(date=current_date).aggregate(
                total=Sum("amount")
            )["total"] or 0
            
            total_revenue = day_room_revenue + day_additional + day_minibar
            net_profit = total_revenue - day_expenses
            
            daily_data[current_date] = {
                "date": current_date,
                "room_revenue": day_room_revenue,
                "additional_revenue": day_additional,
                "minibar_revenue": day_minibar,
                "total_revenue": total_revenue,
                "total_expenses": day_expenses,
                "net_profit": net_profit,
                "profit_margin": round((net_profit / total_revenue) * 100, 2) if total_revenue > 0 else 0,
            }
            current_date += timedelta(days=1)
        
        # Group data
        if group_by == "day":
            data = list(daily_data.values())
        elif group_by == "week":
            weekly = defaultdict(lambda: {
                "room_revenue": 0, "additional_revenue": 0, "minibar_revenue": 0,
                "total_expenses": 0,
            })
            for date, day_data in daily_data.items():
                week_start = date - timedelta(days=date.weekday())
                weekly[week_start]["room_revenue"] += day_data["room_revenue"]
                weekly[week_start]["additional_revenue"] += day_data["additional_revenue"]
                weekly[week_start]["minibar_revenue"] += day_data["minibar_revenue"]
                weekly[week_start]["total_expenses"] += day_data["total_expenses"]
            
            data = []
            for week_start, week_data in sorted(weekly.items()):
                total_rev = week_data["room_revenue"] + week_data["additional_revenue"] + week_data["minibar_revenue"]
                net = total_rev - week_data["total_expenses"]
                data.append({
                    "period": f"Week of {week_start.isoformat()}",
                    "date": week_start,
                    "room_revenue": week_data["room_revenue"],
                    "additional_revenue": week_data["additional_revenue"],
                    "minibar_revenue": week_data["minibar_revenue"],
                    "total_revenue": total_rev,
                    "total_expenses": week_data["total_expenses"],
                    "net_profit": net,
                    "profit_margin": round((net / total_rev) * 100, 2) if total_rev > 0 else 0,
                })
        elif group_by == "month":
            monthly = defaultdict(lambda: {
                "room_revenue": 0, "additional_revenue": 0, "minibar_revenue": 0,
                "total_expenses": 0,
            })
            for date, day_data in daily_data.items():
                month_key = date.replace(day=1)
                monthly[month_key]["room_revenue"] += day_data["room_revenue"]
                monthly[month_key]["additional_revenue"] += day_data["additional_revenue"]
                monthly[month_key]["minibar_revenue"] += day_data["minibar_revenue"]
                monthly[month_key]["total_expenses"] += day_data["total_expenses"]
            
            data = []
            for month_start, month_data in sorted(monthly.items()):
                total_rev = month_data["room_revenue"] + month_data["additional_revenue"] + month_data["minibar_revenue"]
                net = total_rev - month_data["total_expenses"]
                data.append({
                    "period": month_start.strftime("%Y-%m"),
                    "date": month_start,
                    "room_revenue": month_data["room_revenue"],
                    "additional_revenue": month_data["additional_revenue"],
                    "minibar_revenue": month_data["minibar_revenue"],
                    "total_revenue": total_rev,
                    "total_expenses": month_data["total_expenses"],
                    "net_profit": net,
                    "profit_margin": round((net / total_rev) * 100, 2) if total_rev > 0 else 0,
                })
        
        # Calculate summary
        summary = {
            "room_revenue": sum(d["room_revenue"] for d in daily_data.values()),
            "additional_revenue": sum(d["additional_revenue"] for d in daily_data.values()),
            "minibar_revenue": sum(d["minibar_revenue"] for d in daily_data.values()),
            "total_revenue": sum(d["total_revenue"] for d in daily_data.values()),
            "total_expenses": sum(d["total_expenses"] for d in daily_data.values()),
            "net_profit": sum(d["net_profit"] for d in daily_data.values()),
        }
        summary["profit_margin"] = round(
            (summary["net_profit"] / summary["total_revenue"]) * 100, 2
        ) if summary["total_revenue"] > 0 else 0
        
        return Response({
            "summary": summary,
            "data": data,
        })


class KPIReportView(APIView):
    """
    KPI report endpoint (RevPAR, ADR, etc.).
    Returns key performance indicators for the hotel.
    """
    
    permission_classes = [IsAuthenticated]
    
    @extend_schema(
        summary="Get KPI report",
        description="Get key performance indicators including RevPAR and ADR.",
        parameters=[
            OpenApiParameter("start_date", OpenApiTypes.DATE, required=True),
            OpenApiParameter("end_date", OpenApiTypes.DATE, required=True),
            OpenApiParameter("compare_previous", OpenApiTypes.BOOL),
        ],
        responses={200: OpenApiResponse(description="KPI report data")},
        tags=["Reports"],
    )
    def get(self, request):
        from datetime import timedelta
        from django.db.models import Sum, Count
        
        serializer = KPIReportRequestSerializer(data=request.query_params)
        serializer.is_valid(raise_exception=True)
        
        start_date = serializer.validated_data["start_date"]
        end_date = serializer.validated_data["end_date"]
        compare_previous = serializer.validated_data.get("compare_previous", True)
        
        def calculate_kpis(start, end):
            total_days = (end - start).days + 1
            total_rooms = Room.objects.filter(is_active=True).count()
            total_room_nights = total_rooms * total_days
            
            if total_room_nights == 0:
                return None
            
            # Get completed bookings in range
            bookings = Booking.objects.filter(
                status__in=["checked_out", "checked_in", "confirmed"],
                check_in_date__lte=end,
                check_out_date__gt=start,
            )
            
            # Calculate room nights sold
            room_nights_sold = 0
            room_revenue = 0
            for booking in bookings:
                # Calculate nights within the date range
                booking_start = max(booking.check_in_date, start)
                booking_end = min(booking.check_out_date, end + timedelta(days=1))
                nights = (booking_end - booking_start).days
                if nights > 0:
                    room_nights_sold += nights
                    room_revenue += booking.nightly_rate * nights
            
            # Get total revenue and expenses from financial entries
            total_revenue = FinancialEntry.objects.filter(
                entry_type="income",
                date__gte=start,
                date__lte=end,
            ).aggregate(total=Sum("amount"))["total"] or 0
            
            total_expenses = FinancialEntry.objects.filter(
                entry_type="expense",
                date__gte=start,
                date__lte=end,
            ).aggregate(total=Sum("amount"))["total"] or 0
            
            # Add minibar revenue
            minibar_revenue = MinibarSale.objects.filter(
                date__gte=start,
                date__lte=end,
            ).aggregate(total=Sum("total"))["total"] or 0
            
            total_revenue += minibar_revenue
            
            # Calculate KPIs
            occupancy_rate = (room_nights_sold / total_room_nights) * 100 if total_room_nights > 0 else 0
            adr = room_revenue / room_nights_sold if room_nights_sold > 0 else 0
            revpar = room_revenue / total_room_nights if total_room_nights > 0 else 0
            
            return {
                "period_start": start,
                "period_end": end,
                "revpar": round(revpar),
                "adr": round(adr),
                "occupancy_rate": round(occupancy_rate, 2),
                "total_room_nights_available": total_room_nights,
                "total_room_nights_sold": room_nights_sold,
                "total_room_revenue": room_revenue,
                "total_revenue": total_revenue,
                "total_expenses": total_expenses,
                "net_profit": total_revenue - total_expenses,
            }
        
        current_kpis = calculate_kpis(start_date, end_date)
        
        if current_kpis is None:
            return Response({
                "current": None,
                "previous": None,
                "changes": None,
            })
        
        # Calculate previous period for comparison
        previous_kpis = None
        changes = None
        
        if compare_previous:
            period_length = (end_date - start_date).days + 1
            previous_end = start_date - timedelta(days=1)
            previous_start = previous_end - timedelta(days=period_length - 1)
            previous_kpis = calculate_kpis(previous_start, previous_end)
            
            if previous_kpis:
                def calc_change(current, previous):
                    if previous == 0:
                        return None
                    return round(((current - previous) / previous) * 100, 2)
                
                changes = {
                    "revpar_change": calc_change(current_kpis["revpar"], previous_kpis["revpar"]),
                    "adr_change": calc_change(current_kpis["adr"], previous_kpis["adr"]),
                    "occupancy_change": calc_change(current_kpis["occupancy_rate"], previous_kpis["occupancy_rate"]),
                    "revenue_change": calc_change(current_kpis["total_revenue"], previous_kpis["total_revenue"]),
                }
        
        return Response({
            "current": current_kpis,
            "previous": previous_kpis,
            "changes": changes,
        })


class ExpenseReportView(APIView):
    """
    Expense breakdown report by category.
    """
    
    permission_classes = [IsAuthenticated]
    
    @extend_schema(
        summary="Get expense breakdown report",
        description="Get expenses grouped by category for a date range.",
        parameters=[
            OpenApiParameter("start_date", OpenApiTypes.DATE, required=True),
            OpenApiParameter("end_date", OpenApiTypes.DATE, required=True),
        ],
        responses={200: OpenApiResponse(description="Expense report data")},
        tags=["Reports"],
    )
    def get(self, request):
        from django.db.models import Sum, Count
        
        serializer = ExpenseReportRequestSerializer(data=request.query_params)
        serializer.is_valid(raise_exception=True)
        
        start_date = serializer.validated_data["start_date"]
        end_date = serializer.validated_data["end_date"]
        
        # Get expenses grouped by category
        expenses = FinancialEntry.objects.filter(
            entry_type="expense",
            date__gte=start_date,
            date__lte=end_date,
        ).values(
            "category__id",
            "category__name",
            "category__icon",
            "category__color",
        ).annotate(
            total_amount=Sum("amount"),
            transaction_count=Count("id"),
        ).order_by("-total_amount")
        
        total_expenses = sum(e["total_amount"] for e in expenses)
        
        data = []
        for expense in expenses:
            data.append({
                "category_id": expense["category__id"],
                "category_name": expense["category__name"],
                "category_icon": expense["category__icon"],
                "category_color": expense["category__color"],
                "total_amount": expense["total_amount"],
                "transaction_count": expense["transaction_count"],
                "percentage": round((expense["total_amount"] / total_expenses) * 100, 2) if total_expenses > 0 else 0,
            })
        
        return Response({
            "summary": {
                "total_expenses": total_expenses,
                "category_count": len(data),
                "transaction_count": sum(e["transaction_count"] for e in data),
            },
            "data": data,
        })


class ChannelPerformanceView(APIView):
    """
    Channel (booking source) performance report.
    """
    
    permission_classes = [IsAuthenticated]
    
    @extend_schema(
        summary="Get channel performance report",
        description="Get booking performance by source/channel.",
        parameters=[
            OpenApiParameter("start_date", OpenApiTypes.DATE, required=True),
            OpenApiParameter("end_date", OpenApiTypes.DATE, required=True),
        ],
        responses={200: OpenApiResponse(description="Channel performance data")},
        tags=["Reports"],
    )
    def get(self, request):
        from django.db.models import Sum, Count, Avg, F
        
        serializer = ChannelPerformanceRequestSerializer(data=request.query_params)
        serializer.is_valid(raise_exception=True)
        
        start_date = serializer.validated_data["start_date"]
        end_date = serializer.validated_data["end_date"]
        
        # Get bookings in range
        all_bookings = Booking.objects.filter(
            check_in_date__lte=end_date,
            check_out_date__gt=start_date,
        )
        
        # Group by source
        source_data = all_bookings.values("source").annotate(
            booking_count=Count("id"),
            total_revenue=Sum("total_amount"),
            cancelled_count=Count("id", filter=models.Q(status="cancelled")),
        ).order_by("-total_revenue")
        
        total_revenue = sum(s["total_revenue"] or 0 for s in source_data)
        
        data = []
        for source in source_data:
            source_code = source["source"]
            source_display = dict(Booking.Source.choices).get(source_code, source_code)
            
            # Calculate total nights for this source
            source_bookings = all_bookings.filter(source=source_code)
            total_nights = sum(b.nights for b in source_bookings)
            
            booking_count = source["booking_count"]
            cancelled_count = source["cancelled_count"]
            revenue = source["total_revenue"] or 0
            
            data.append({
                "source": source_code,
                "source_display": source_display,
                "booking_count": booking_count,
                "total_nights": total_nights,
                "total_revenue": revenue,
                "average_rate": round(revenue / total_nights) if total_nights > 0 else 0,
                "cancellation_count": cancelled_count,
                "cancellation_rate": round((cancelled_count / booking_count) * 100, 2) if booking_count > 0 else 0,
                "percentage_of_revenue": round((revenue / total_revenue) * 100, 2) if total_revenue > 0 else 0,
            })
        
        return Response({
            "summary": {
                "total_bookings": sum(s["booking_count"] for s in data),
                "total_revenue": total_revenue,
                "total_nights": sum(s["total_nights"] for s in data),
                "total_cancellations": sum(s["cancellation_count"] for s in data),
            },
            "data": data,
        })


class GuestDemographicsView(APIView):
    """
    Guest demographics report.
    """
    
    permission_classes = [IsAuthenticated]
    
    @extend_schema(
        summary="Get guest demographics report",
        description="Get guest statistics by nationality or other criteria.",
        parameters=[
            OpenApiParameter("start_date", OpenApiTypes.DATE, required=True),
            OpenApiParameter("end_date", OpenApiTypes.DATE, required=True),
            OpenApiParameter("group_by", OpenApiTypes.STR, enum=["nationality", "source", "room_type"]),
        ],
        responses={200: OpenApiResponse(description="Guest demographics data")},
        tags=["Reports"],
    )
    def get(self, request):
        from django.db.models import Sum, Count, Avg
        
        serializer = GuestDemographicsRequestSerializer(data=request.query_params)
        serializer.is_valid(raise_exception=True)
        
        start_date = serializer.validated_data["start_date"]
        end_date = serializer.validated_data["end_date"]
        group_by = serializer.validated_data.get("group_by", "nationality")
        
        # Get bookings in range
        bookings = Booking.objects.filter(
            status__in=["confirmed", "checked_in", "checked_out"],
            check_in_date__lte=end_date,
            check_out_date__gt=start_date,
        ).select_related("guest", "room__room_type")
        
        total_revenue = sum(b.total_amount for b in bookings)
        
        if group_by == "nationality":
            # Group by guest nationality
            data_dict = {}
            for booking in bookings:
                nationality = booking.guest.nationality or "Unknown"
                if nationality not in data_dict:
                    data_dict[nationality] = {
                        "guest_ids": set(),
                        "booking_count": 0,
                        "total_nights": 0,
                        "total_revenue": 0,
                    }
                data_dict[nationality]["guest_ids"].add(booking.guest_id)
                data_dict[nationality]["booking_count"] += 1
                data_dict[nationality]["total_nights"] += booking.nights
                data_dict[nationality]["total_revenue"] += booking.total_amount
            
            data = []
            for nationality, stats in sorted(data_dict.items(), key=lambda x: -x[1]["total_revenue"]):
                guest_count = len(stats["guest_ids"])
                data.append({
                    "nationality": nationality,
                    "guest_count": guest_count,
                    "booking_count": stats["booking_count"],
                    "total_nights": stats["total_nights"],
                    "total_revenue": stats["total_revenue"],
                    "percentage": round((stats["total_revenue"] / total_revenue) * 100, 2) if total_revenue > 0 else 0,
                    "average_stay": round(stats["total_nights"] / stats["booking_count"], 2) if stats["booking_count"] > 0 else 0,
                })
        else:
            # Default to nationality grouping
            data = []
        
        return Response({
            "summary": {
                "total_guests": len(set(b.guest_id for b in bookings)),
                "total_bookings": bookings.count(),
                "total_revenue": total_revenue,
            },
            "data": data,
        })


class ComparativeReportView(APIView):
    """
    Period-over-period comparative report.
    """
    
    permission_classes = [IsAuthenticated]
    
    @extend_schema(
        summary="Get comparative report",
        description="Compare metrics between two periods.",
        parameters=[
            OpenApiParameter("current_start", OpenApiTypes.DATE, required=True),
            OpenApiParameter("current_end", OpenApiTypes.DATE, required=True),
            OpenApiParameter("comparison_type", OpenApiTypes.STR, enum=["previous_period", "previous_year", "custom"]),
            OpenApiParameter("previous_start", OpenApiTypes.DATE),
            OpenApiParameter("previous_end", OpenApiTypes.DATE),
        ],
        responses={200: OpenApiResponse(description="Comparative report data")},
        tags=["Reports"],
    )
    def get(self, request):
        from datetime import timedelta
        from dateutil.relativedelta import relativedelta
        from django.db.models import Sum, Count
        
        serializer = ComparativeReportRequestSerializer(data=request.query_params)
        serializer.is_valid(raise_exception=True)
        
        current_start = serializer.validated_data["current_start"]
        current_end = serializer.validated_data["current_end"]
        comparison_type = serializer.validated_data.get("comparison_type", "previous_period")
        
        # Calculate previous period dates
        if comparison_type == "previous_period":
            period_length = (current_end - current_start).days + 1
            previous_end = current_start - timedelta(days=1)
            previous_start = previous_end - timedelta(days=period_length - 1)
        elif comparison_type == "previous_year":
            previous_start = current_start - relativedelta(years=1)
            previous_end = current_end - relativedelta(years=1)
        else:  # custom
            previous_start = serializer.validated_data.get("previous_start")
            previous_end = serializer.validated_data.get("previous_end")
        
        def get_metrics(start, end):
            total_rooms = Room.objects.filter(is_active=True).count()
            total_days = (end - start).days + 1
            total_room_nights = total_rooms * total_days
            
            # Revenue
            revenue = FinancialEntry.objects.filter(
                entry_type="income",
                date__gte=start,
                date__lte=end,
            ).aggregate(total=Sum("amount"))["total"] or 0
            
            # Expenses
            expenses = FinancialEntry.objects.filter(
                entry_type="expense",
                date__gte=start,
                date__lte=end,
            ).aggregate(total=Sum("amount"))["total"] or 0
            
            # Bookings
            bookings = Booking.objects.filter(
                status__in=["confirmed", "checked_in", "checked_out"],
                check_in_date__lte=end,
                check_out_date__gt=start,
            )
            booking_count = bookings.count()
            room_nights_sold = sum(b.nights for b in bookings)
            room_revenue = sum(b.total_amount for b in bookings)
            
            # Occupancy
            occupancy = (room_nights_sold / total_room_nights) * 100 if total_room_nights > 0 else 0
            
            # ADR
            adr = room_revenue / room_nights_sold if room_nights_sold > 0 else 0
            
            # RevPAR
            revpar = room_revenue / total_room_nights if total_room_nights > 0 else 0
            
            return {
                "revenue": revenue,
                "expenses": expenses,
                "net_profit": revenue - expenses,
                "booking_count": booking_count,
                "occupancy_rate": round(occupancy, 2),
                "adr": round(adr),
                "revpar": round(revpar),
            }
        
        current_metrics = get_metrics(current_start, current_end)
        previous_metrics = get_metrics(previous_start, previous_end) if previous_start and previous_end else None
        
        def calc_change(current, previous):
            if previous is None or previous == 0:
                return None
            return round(((current - previous) / previous) * 100, 2)
        
        comparisons = []
        metrics_list = [
            ("revenue", "Doanh thu"),
            ("expenses", "Chi phí"),
            ("net_profit", "Lợi nhuận"),
            ("booking_count", "Số đặt phòng"),
            ("occupancy_rate", "Tỷ lệ lấp đầy"),
            ("adr", "Giá trung bình/đêm"),
            ("revpar", "RevPAR"),
        ]
        
        for metric_key, metric_name in metrics_list:
            current_val = current_metrics[metric_key]
            previous_val = previous_metrics[metric_key] if previous_metrics else None
            
            comparisons.append({
                "metric": metric_name,
                "metric_key": metric_key,
                "current_period_value": current_val,
                "previous_period_value": previous_val,
                "change_amount": current_val - previous_val if previous_val is not None else None,
                "change_percentage": calc_change(current_val, previous_val),
            })
        
        return Response({
            "current_period": {
                "start": current_start,
                "end": current_end,
                "metrics": current_metrics,
            },
            "previous_period": {
                "start": previous_start,
                "end": previous_end,
                "metrics": previous_metrics,
            } if previous_metrics else None,
            "comparisons": comparisons,
        })


class ExportReportView(APIView):
    """
    Export reports to Excel/CSV.
    """
    
    permission_classes = [IsAuthenticated]
    
    @extend_schema(
        summary="Export report to file",
        description="Export report data to Excel or CSV format.",
        parameters=[
            OpenApiParameter("report_type", OpenApiTypes.STR, required=True, 
                           enum=["occupancy", "revenue", "expenses", "kpi", "channels", "demographics"]),
            OpenApiParameter("start_date", OpenApiTypes.DATE, required=True),
            OpenApiParameter("end_date", OpenApiTypes.DATE, required=True),
            OpenApiParameter("format", OpenApiTypes.STR, enum=["xlsx", "csv"]),
        ],
        responses={200: OpenApiResponse(description="File download")},
        tags=["Reports"],
    )
    def get(self, request):
        import csv
        import io
        from django.http import HttpResponse
        
        serializer = ExportReportRequestSerializer(data=request.query_params)
        serializer.is_valid(raise_exception=True)
        
        report_type = serializer.validated_data["report_type"]
        start_date = serializer.validated_data["start_date"]
        end_date = serializer.validated_data["end_date"]
        export_format = serializer.validated_data.get("format", "xlsx")
        
        # Get report data based on type
        if report_type == "occupancy":
            view = OccupancyReportView()
            view.request = request
            response_data = view.get(request).data
            data = response_data.get("data", [])
            headers = ["Date", "Total Rooms", "Occupied", "Available", "Occupancy %", "Revenue"]
            rows = [[
                d.get("date") or d.get("period"),
                d["total_rooms"],
                d["occupied_rooms"],
                d["available_rooms"],
                d["occupancy_rate"],
                d["revenue"],
            ] for d in data]
        elif report_type == "revenue":
            view = RevenueReportView()
            view.request = request
            response_data = view.get(request).data
            data = response_data.get("data", [])
            headers = ["Date", "Room Revenue", "Additional", "Minibar", "Total Revenue", "Expenses", "Net Profit", "Margin %"]
            rows = [[
                d.get("date") or d.get("period"),
                d["room_revenue"],
                d["additional_revenue"],
                d["minibar_revenue"],
                d["total_revenue"],
                d["total_expenses"],
                d["net_profit"],
                d["profit_margin"],
            ] for d in data]
        elif report_type == "expenses":
            view = ExpenseReportView()
            view.request = request
            response_data = view.get(request).data
            data = response_data.get("data", [])
            headers = ["Category", "Amount", "Transactions", "Percentage"]
            rows = [[
                d["category_name"],
                d["total_amount"],
                d["transaction_count"],
                d["percentage"],
            ] for d in data]
        elif report_type == "channels":
            view = ChannelPerformanceView()
            view.request = request
            response_data = view.get(request).data
            data = response_data.get("data", [])
            headers = ["Channel", "Bookings", "Nights", "Revenue", "Avg Rate", "Cancellations", "Cancel %", "Revenue %"]
            rows = [[
                d["source_display"],
                d["booking_count"],
                d["total_nights"],
                d["total_revenue"],
                d["average_rate"],
                d["cancellation_count"],
                d["cancellation_rate"],
                d["percentage_of_revenue"],
            ] for d in data]
        elif report_type == "demographics":
            view = GuestDemographicsView()
            view.request = request
            response_data = view.get(request).data
            data = response_data.get("data", [])
            headers = ["Nationality", "Guests", "Bookings", "Nights", "Revenue", "Percentage", "Avg Stay"]
            rows = [[
                d["nationality"],
                d["guest_count"],
                d["booking_count"],
                d["total_nights"],
                d["total_revenue"],
                d["percentage"],
                d["average_stay"],
            ] for d in data]
        else:
            return Response({"detail": "Invalid report type"}, status=400)
        
        # Generate file
        if export_format == "csv":
            output = io.StringIO()
            writer = csv.writer(output)
            writer.writerow(headers)
            writer.writerows(rows)
            
            response = HttpResponse(output.getvalue(), content_type="text/csv")
            response["Content-Disposition"] = f'attachment; filename="{report_type}_report_{start_date}_{end_date}.csv"'
            return response
        else:  # xlsx
            try:
                import openpyxl
                from openpyxl.utils import get_column_letter
                
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = report_type.title()
                
                # Headers
                for col, header in enumerate(headers, 1):
                    ws.cell(row=1, column=col, value=header)
                
                # Data
                for row_idx, row in enumerate(rows, 2):
                    for col_idx, value in enumerate(row, 1):
                        ws.cell(row=row_idx, column=col_idx, value=value)
                
                # Auto-width columns
                for col_idx in range(1, len(headers) + 1):
                    ws.column_dimensions[get_column_letter(col_idx)].width = 15
                
                output = io.BytesIO()
                wb.save(output)
                output.seek(0)
                
                response = HttpResponse(
                    output.getvalue(),
                    content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
                response["Content-Disposition"] = f'attachment; filename="{report_type}_report_{start_date}_{end_date}.xlsx"'
                return response
            except ImportError:
                # Fallback to CSV if openpyxl not installed
                output = io.StringIO()
                writer = csv.writer(output)
                writer.writerow(headers)
                writer.writerows(rows)
                
                response = HttpResponse(output.getvalue(), content_type="text/csv")
                response["Content-Disposition"] = f'attachment; filename="{report_type}_report_{start_date}_{end_date}.csv"'
                return response


# ============================================================
# Lost & Found ViewSet (Phase 3)
# ============================================================


@extend_schema_view(
    list=extend_schema(
        summary="List lost and found items",
        description="Get list of all lost and found items with optional filtering.",
        parameters=[
            OpenApiParameter("status", OpenApiTypes.STR, description="Filter by status (found/stored/claimed/donated/disposed)"),
            OpenApiParameter("category", OpenApiTypes.STR, description="Filter by category"),
            OpenApiParameter("room", OpenApiTypes.INT, description="Filter by room ID"),
            OpenApiParameter("guest", OpenApiTypes.INT, description="Filter by guest ID"),
            OpenApiParameter("found_date_from", OpenApiTypes.DATE, description="Filter by found date (from)"),
            OpenApiParameter("found_date_to", OpenApiTypes.DATE, description="Filter by found date (to)"),
            OpenApiParameter("search", OpenApiTypes.STR, description="Search in item name and description"),
        ],
        tags=["Lost & Found"],
    ),
    retrieve=extend_schema(
        summary="Get lost and found item details",
        tags=["Lost & Found"],
    ),
    create=extend_schema(
        summary="Create lost and found item",
        tags=["Lost & Found"],
    ),
    update=extend_schema(
        summary="Update lost and found item",
        tags=["Lost & Found"],
    ),
    partial_update=extend_schema(
        summary="Partially update lost and found item",
        tags=["Lost & Found"],
    ),
    destroy=extend_schema(
        summary="Delete lost and found item",
        tags=["Lost & Found"],
    ),
)
class LostAndFoundViewSet(viewsets.ModelViewSet):
    """
    ViewSet for managing lost and found items.

    Endpoints:
    - GET /lost-found/ - List all items
    - POST /lost-found/ - Create new item
    - GET /lost-found/{id}/ - Get item details
    - PUT /lost-found/{id}/ - Update item
    - DELETE /lost-found/{id}/ - Delete item
    - POST /lost-found/{id}/claim/ - Mark as claimed/returned
    - POST /lost-found/{id}/dispose/ - Mark as disposed/donated
    - POST /lost-found/{id}/store/ - Mark as stored
    - GET /lost-found/statistics/ - Get summary statistics
    """

    queryset = LostAndFound.objects.all()
    permission_classes = [IsAuthenticated, IsStaffOrManager]

    def get_serializer_class(self):
        """Return appropriate serializer based on action."""
        if self.action == "list":
            return LostAndFoundListSerializer
        if self.action == "create":
            return LostAndFoundCreateSerializer
        if self.action in ["update", "partial_update"]:
            return LostAndFoundUpdateSerializer
        if self.action == "claim":
            return LostAndFoundClaimSerializer
        if self.action == "dispose":
            return LostAndFoundDisposeSerializer
        return LostAndFoundSerializer

    def get_queryset(self):
        """Filter queryset based on query parameters."""
        queryset = LostAndFound.objects.select_related(
            "room", "guest", "booking", "found_by", "claimed_by_staff"
        )

        # Filter by status
        status_filter = self.request.query_params.get("status")
        if status_filter:
            queryset = queryset.filter(status=status_filter)

        # Filter by category
        category = self.request.query_params.get("category")
        if category:
            queryset = queryset.filter(category=category)

        # Filter by room
        room_id = self.request.query_params.get("room")
        if room_id:
            queryset = queryset.filter(room_id=room_id)

        # Filter by guest
        guest_id = self.request.query_params.get("guest")
        if guest_id:
            queryset = queryset.filter(guest_id=guest_id)

        # Filter by found date range
        found_from = self.request.query_params.get("found_date_from")
        found_to = self.request.query_params.get("found_date_to")
        if found_from:
            queryset = queryset.filter(found_date__gte=found_from)
        if found_to:
            queryset = queryset.filter(found_date__lte=found_to)

        # Search
        search = self.request.query_params.get("search")
        if search:
            queryset = queryset.filter(
                Q(item_name__icontains=search) | Q(description__icontains=search)
            )

        return queryset.order_by("-found_date", "-created_at")

    def perform_create(self, serializer):
        """Set found_by to current user if not specified."""
        if not serializer.validated_data.get("found_by"):
            serializer.save(found_by=self.request.user)
        else:
            serializer.save()

    @extend_schema(
        summary="Claim item",
        description="Mark item as claimed/returned to guest.",
        request=LostAndFoundClaimSerializer,
        responses={200: LostAndFoundSerializer},
        tags=["Lost & Found"],
    )
    @action(detail=True, methods=["post"], url_path="claim")
    def claim(self, request, pk=None):
        """Mark item as claimed/returned to guest."""
        item = self.get_object()
        serializer = self.get_serializer(data=request.data, context={"item": item})
        serializer.is_valid(raise_exception=True)

        notes = serializer.validated_data.get("notes", "")
        item.claim(request.user, notes)

        return Response(LostAndFoundSerializer(item).data)

    @extend_schema(
        summary="Dispose item",
        description="Mark item as disposed or donated.",
        request=LostAndFoundDisposeSerializer,
        responses={200: LostAndFoundSerializer},
        tags=["Lost & Found"],
    )
    @action(detail=True, methods=["post"], url_path="dispose")
    def dispose(self, request, pk=None):
        """Mark item as disposed or donated."""
        item = self.get_object()
        serializer = self.get_serializer(data=request.data, context={"item": item})
        serializer.is_valid(raise_exception=True)

        method = serializer.validated_data.get("method", "disposed")
        item.dispose(method)

        return Response(LostAndFoundSerializer(item).data)

    @extend_schema(
        summary="Store item",
        description="Mark item as stored (in lost & found storage).",
        request=None,
        responses={200: LostAndFoundSerializer},
        tags=["Lost & Found"],
    )
    @action(detail=True, methods=["post"], url_path="store")
    def store(self, request, pk=None):
        """Mark item as stored."""
        item = self.get_object()

        if item.status in [LostAndFound.Status.CLAIMED, LostAndFound.Status.DISPOSED, LostAndFound.Status.DONATED]:
            return Response(
                {"detail": "Không thể lưu giữ vật phẩm đã được xử lý."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        item.status = LostAndFound.Status.STORED
        item.save()

        return Response(LostAndFoundSerializer(item).data)

    @extend_schema(
        summary="Get statistics",
        description="Get summary statistics for lost and found items.",
        responses={
            200: OpenApiResponse(
                description="Statistics data",
                response={
                    "type": "object",
                    "properties": {
                        "total_items": {"type": "integer"},
                        "by_status": {"type": "object"},
                        "by_category": {"type": "object"},
                        "unclaimed_value": {"type": "number"},
                        "recent_items": {"type": "array"},
                    },
                },
            )
        },
        tags=["Lost & Found"],
    )
    @action(detail=False, methods=["get"], url_path="statistics")
    def statistics(self, request):
        """Get summary statistics for lost and found items."""
        from django.db.models import Count, Sum

        # Total items
        total = LostAndFound.objects.count()

        # By status
        by_status = dict(
            LostAndFound.objects.values_list("status")
            .annotate(count=Count("id"))
            .values_list("status", "count")
        )

        # By category
        by_category = dict(
            LostAndFound.objects.values_list("category")
            .annotate(count=Count("id"))
            .values_list("category", "count")
        )

        # Unclaimed value
        unclaimed = LostAndFound.objects.filter(
            status__in=[LostAndFound.Status.FOUND, LostAndFound.Status.STORED]
        ).aggregate(total=Sum("estimated_value"))["total"] or 0

        # Recent items
        recent = LostAndFoundListSerializer(
            LostAndFound.objects.order_by("-created_at")[:5], many=True
        ).data

        return Response({
            "total_items": total,
            "by_status": by_status,
            "by_category": by_category,
            "unclaimed_value": unclaimed,
            "recent_items": recent,
        })


# ============================================================
# Group Booking ViewSet (Phase 3)
# ============================================================


@extend_schema_view(
    list=extend_schema(
        summary="List group bookings",
        description="Get list of all group bookings with optional filtering.",
        parameters=[
            OpenApiParameter("status", OpenApiTypes.STR, description="Filter by status"),
            OpenApiParameter("date_from", OpenApiTypes.DATE, description="Filter by check-in date (from)"),
            OpenApiParameter("date_to", OpenApiTypes.DATE, description="Filter by check-in date (to)"),
            OpenApiParameter("search", OpenApiTypes.STR, description="Search in name, contact, company"),
        ],
        tags=["Group Booking"],
    ),
    retrieve=extend_schema(
        summary="Get group booking details",
        tags=["Group Booking"],
    ),
    create=extend_schema(
        summary="Create group booking",
        tags=["Group Booking"],
    ),
    update=extend_schema(
        summary="Update group booking",
        tags=["Group Booking"],
    ),
    partial_update=extend_schema(
        summary="Partially update group booking",
        tags=["Group Booking"],
    ),
    destroy=extend_schema(
        summary="Delete group booking",
        tags=["Group Booking"],
    ),
)
class GroupBookingViewSet(viewsets.ModelViewSet):
    """
    ViewSet for managing group bookings.

    Endpoints:
    - GET /group-bookings/ - List all group bookings
    - POST /group-bookings/ - Create new group booking
    - GET /group-bookings/{id}/ - Get group booking details
    - PUT /group-bookings/{id}/ - Update group booking
    - DELETE /group-bookings/{id}/ - Delete group booking
    - POST /group-bookings/{id}/confirm/ - Confirm group booking
    - POST /group-bookings/{id}/check-in/ - Check in group
    - POST /group-bookings/{id}/check-out/ - Check out group
    - POST /group-bookings/{id}/cancel/ - Cancel group booking
    - POST /group-bookings/{id}/assign-rooms/ - Assign rooms to group
    """

    queryset = GroupBooking.objects.all()
    permission_classes = [IsAuthenticated, IsStaffOrManager]

    def get_serializer_class(self):
        """Return appropriate serializer based on action."""
        if self.action == "list":
            return GroupBookingListSerializer
        if self.action == "create":
            return GroupBookingCreateSerializer
        if self.action in ["update", "partial_update"]:
            return GroupBookingUpdateSerializer
        return GroupBookingSerializer

    def get_queryset(self):
        """Filter queryset based on query parameters."""
        queryset = GroupBooking.objects.prefetch_related("rooms").select_related("created_by")

        # Filter by status
        status_filter = self.request.query_params.get("status")
        if status_filter:
            queryset = queryset.filter(status=status_filter)

        # Filter by date range
        date_from = self.request.query_params.get("date_from")
        date_to = self.request.query_params.get("date_to")
        if date_from:
            queryset = queryset.filter(check_in_date__gte=date_from)
        if date_to:
            queryset = queryset.filter(check_in_date__lte=date_to)

        # Search
        search = self.request.query_params.get("search")
        if search:
            queryset = queryset.filter(
                Q(name__icontains=search)
                | Q(contact_name__icontains=search)
                | Q(company__icontains=search)
                | Q(contact_phone__icontains=search)
            )

        return queryset.order_by("-check_in_date", "-created_at")

    @extend_schema(
        summary="Confirm group booking",
        description="Change status from tentative to confirmed.",
        request=None,
        responses={200: GroupBookingSerializer},
        tags=["Group Booking"],
    )
    @action(detail=True, methods=["post"], url_path="confirm")
    def confirm(self, request, pk=None):
        """Confirm the group booking."""
        group = self.get_object()

        if group.status != GroupBooking.Status.TENTATIVE:
            return Response(
                {"detail": f"Không thể xác nhận đặt phòng ở trạng thái {group.get_status_display()}."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        group.status = GroupBooking.Status.CONFIRMED
        group.save()

        return Response(GroupBookingSerializer(group).data)

    @extend_schema(
        summary="Check in group",
        description="Check in the entire group.",
        request=None,
        responses={200: GroupBookingSerializer},
        tags=["Group Booking"],
    )
    @action(detail=True, methods=["post"], url_path="check-in")
    def check_in(self, request, pk=None):
        """Check in the group."""
        from django.utils import timezone

        group = self.get_object()

        if group.status not in [GroupBooking.Status.TENTATIVE, GroupBooking.Status.CONFIRMED]:
            return Response(
                {"detail": f"Không thể nhận phòng cho đặt phòng ở trạng thái {group.get_status_display()}."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        group.status = GroupBooking.Status.CHECKED_IN
        group.actual_check_in = timezone.now()
        group.save()

        return Response(GroupBookingSerializer(group).data)

    @extend_schema(
        summary="Check out group",
        description="Check out the entire group.",
        request=None,
        responses={200: GroupBookingSerializer},
        tags=["Group Booking"],
    )
    @action(detail=True, methods=["post"], url_path="check-out")
    def check_out(self, request, pk=None):
        """Check out the group."""
        from django.utils import timezone

        group = self.get_object()

        if group.status != GroupBooking.Status.CHECKED_IN:
            return Response(
                {"detail": f"Không thể trả phòng cho đặt phòng ở trạng thái {group.get_status_display()}."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        group.status = GroupBooking.Status.CHECKED_OUT
        group.actual_check_out = timezone.now()
        group.save()

        # Free up all assigned rooms
        group.rooms.update(status=Room.Status.CLEANING)

        return Response(GroupBookingSerializer(group).data)

    @extend_schema(
        summary="Cancel group booking",
        description="Cancel the group booking.",
        request=None,
        responses={200: GroupBookingSerializer},
        tags=["Group Booking"],
    )
    @action(detail=True, methods=["post"], url_path="cancel")
    def cancel(self, request, pk=None):
        """Cancel the group booking."""
        group = self.get_object()

        if group.status in [GroupBooking.Status.CHECKED_OUT, GroupBooking.Status.CANCELLED]:
            return Response(
                {"detail": f"Không thể hủy đặt phòng ở trạng thái {group.get_status_display()}."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        group.status = GroupBooking.Status.CANCELLED
        group.save()

        return Response(GroupBookingSerializer(group).data)

    @extend_schema(
        summary="Assign rooms to group",
        description="Assign specific rooms to the group booking.",
        request={
            "type": "object",
            "properties": {
                "room_ids": {"type": "array", "items": {"type": "integer"}},
            },
            "required": ["room_ids"],
        },
        responses={200: GroupBookingSerializer},
        tags=["Group Booking"],
    )
    @action(detail=True, methods=["post"], url_path="assign-rooms")
    def assign_rooms(self, request, pk=None):
        """Assign rooms to the group booking."""
        group = self.get_object()

        room_ids = request.data.get("room_ids", [])
        if not room_ids:
            return Response(
                {"detail": "Vui lòng cung cấp danh sách phòng."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        rooms = Room.objects.filter(id__in=room_ids, is_active=True)
        if rooms.count() != len(room_ids):
            return Response(
                {"detail": "Một số phòng không tồn tại hoặc không hoạt động."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        group.rooms.set(rooms)
        group.save()

        return Response(GroupBookingSerializer(group).data)


# ==============================================================================
# Room Inspection ViewSets
# ==============================================================================


@extend_schema_view(
    list=extend_schema(
        summary="List inspection templates",
        description="Get a list of all inspection templates.",
        tags=["Room Inspection"],
    ),
    retrieve=extend_schema(
        summary="Get template details",
        description="Get detailed information about an inspection template.",
        tags=["Room Inspection"],
    ),
    create=extend_schema(
        summary="Create template",
        description="Create a new inspection template.",
        tags=["Room Inspection"],
    ),
    update=extend_schema(
        summary="Update template",
        description="Update an existing inspection template.",
        tags=["Room Inspection"],
    ),
    partial_update=extend_schema(
        summary="Partially update template",
        description="Partially update an existing inspection template.",
        tags=["Room Inspection"],
    ),
    destroy=extend_schema(
        summary="Delete template",
        description="Delete an inspection template.",
        tags=["Room Inspection"],
    ),
)
class InspectionTemplateViewSet(viewsets.ModelViewSet):
    """ViewSet for managing inspection templates."""

    permission_classes = [IsAuthenticated, IsStaffOrManager]
    filterset_fields = ["inspection_type", "room_type", "is_default", "is_active"]
    search_fields = ["name"]
    ordering_fields = ["name", "created_at"]
    ordering = ["name"]

    def get_queryset(self):
        return InspectionTemplate.objects.select_related("room_type", "created_by").all()

    def get_serializer_class(self):
        if self.action == "list":
            return InspectionTemplateListSerializer
        if self.action == "create":
            return InspectionTemplateCreateSerializer
        return InspectionTemplateSerializer

    @extend_schema(
        summary="Get default templates",
        description="Get default inspection templates by type.",
        parameters=[
            OpenApiParameter(
                name="inspection_type",
                type=str,
                required=False,
                description="Filter by inspection type",
            ),
        ],
        responses={200: InspectionTemplateListSerializer(many=True)},
        tags=["Room Inspection"],
    )
    @action(detail=False, methods=["get"], url_path="defaults")
    def defaults(self, request):
        """Get default inspection templates."""
        queryset = self.get_queryset().filter(is_default=True, is_active=True)

        inspection_type = request.query_params.get("inspection_type")
        if inspection_type:
            queryset = queryset.filter(inspection_type=inspection_type)

        serializer = InspectionTemplateListSerializer(queryset, many=True)
        return Response(serializer.data)


@extend_schema_view(
    list=extend_schema(
        summary="List room inspections",
        description="Get a list of all room inspections with optional filters.",
        parameters=[
            OpenApiParameter(name="room", type=int, description="Filter by room ID"),
            OpenApiParameter(name="status", type=str, description="Filter by status"),
            OpenApiParameter(name="inspection_type", type=str, description="Filter by type"),
            OpenApiParameter(name="from_date", type=str, description="Filter from date (YYYY-MM-DD)"),
            OpenApiParameter(name="to_date", type=str, description="Filter to date (YYYY-MM-DD)"),
        ],
        tags=["Room Inspection"],
    ),
    retrieve=extend_schema(
        summary="Get inspection details",
        description="Get detailed information about a room inspection.",
        tags=["Room Inspection"],
    ),
    create=extend_schema(
        summary="Create inspection",
        description="Create a new room inspection.",
        tags=["Room Inspection"],
    ),
    update=extend_schema(
        summary="Update inspection",
        description="Update an existing room inspection.",
        tags=["Room Inspection"],
    ),
    partial_update=extend_schema(
        summary="Partially update inspection",
        description="Partially update an existing room inspection.",
        tags=["Room Inspection"],
    ),
    destroy=extend_schema(
        summary="Delete inspection",
        description="Delete a room inspection.",
        tags=["Room Inspection"],
    ),
)
class RoomInspectionViewSet(viewsets.ModelViewSet):
    """ViewSet for managing room inspections."""

    permission_classes = [IsAuthenticated, IsStaffOrManager]
    filterset_fields = ["room", "status", "inspection_type", "inspector"]
    search_fields = ["room__number", "notes"]
    ordering_fields = ["scheduled_date", "completed_at", "score", "created_at"]
    ordering = ["-scheduled_date"]

    def get_queryset(self):
        queryset = RoomInspection.objects.select_related(
            "room", "room__room_type", "booking", "booking__guest", "inspector"
        ).all()

        # Date filters
        from_date = self.request.query_params.get("from_date")
        to_date = self.request.query_params.get("to_date")

        if from_date:
            queryset = queryset.filter(scheduled_date__gte=from_date)
        if to_date:
            queryset = queryset.filter(scheduled_date__lte=to_date)

        return queryset

    def get_serializer_class(self):
        if self.action == "list":
            return RoomInspectionListSerializer
        if self.action == "create":
            return RoomInspectionCreateSerializer
        if self.action in ["update", "partial_update"]:
            return RoomInspectionUpdateSerializer
        if self.action == "complete":
            return RoomInspectionCompleteSerializer
        if self.action == "statistics":
            return RoomInspectionStatisticsSerializer
        return RoomInspectionSerializer

    @extend_schema(
        summary="Complete inspection",
        description="Mark a room inspection as completed with results.",
        request=RoomInspectionCompleteSerializer,
        responses={200: RoomInspectionSerializer},
        tags=["Room Inspection"],
    )
    @action(detail=True, methods=["post"], url_path="complete")
    def complete(self, request, pk=None):
        """Complete the room inspection with checklist results."""
        inspection = self.get_object()

        if inspection.status == RoomInspection.Status.COMPLETED:
            return Response(
                {"detail": "Kiểm tra này đã hoàn thành."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        serializer = RoomInspectionCompleteSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        inspection.checklist_items = serializer.validated_data["checklist_items"]
        inspection.images = serializer.validated_data.get("images", [])
        inspection.notes = serializer.validated_data.get("notes", "")
        inspection.action_required = serializer.validated_data.get("action_required", "")
        inspection.complete(request.user)

        return Response(RoomInspectionSerializer(inspection).data)

    @extend_schema(
        summary="Start inspection",
        description="Mark an inspection as in progress.",
        request=None,
        responses={200: RoomInspectionSerializer},
        tags=["Room Inspection"],
    )
    @action(detail=True, methods=["post"], url_path="start")
    def start(self, request, pk=None):
        """Start the room inspection."""
        inspection = self.get_object()

        if inspection.status != RoomInspection.Status.PENDING:
            return Response(
                {"detail": "Chỉ có thể bắt đầu kiểm tra đang chờ."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        inspection.status = RoomInspection.Status.IN_PROGRESS
        inspection.inspector = request.user
        inspection.save()

        return Response(RoomInspectionSerializer(inspection).data)

    @extend_schema(
        summary="Get inspection statistics",
        description="Get statistics about room inspections.",
        parameters=[
            OpenApiParameter(name="from_date", type=str, description="From date (YYYY-MM-DD)"),
            OpenApiParameter(name="to_date", type=str, description="To date (YYYY-MM-DD)"),
        ],
        responses={200: RoomInspectionStatisticsSerializer},
        tags=["Room Inspection"],
    )
    @action(detail=False, methods=["get"], url_path="statistics")
    def statistics(self, request):
        """Get room inspection statistics."""
        from django.db.models import Avg, Count, Sum

        queryset = self.get_queryset()

        # Calculate statistics
        total_inspections = queryset.count()
        completed_inspections = queryset.filter(status=RoomInspection.Status.COMPLETED).count()
        pending_inspections = queryset.filter(status=RoomInspection.Status.PENDING).count()
        requires_action = queryset.filter(status=RoomInspection.Status.REQUIRES_ACTION).count()

        # Average score from completed inspections
        avg_score = queryset.filter(
            status__in=[RoomInspection.Status.COMPLETED, RoomInspection.Status.REQUIRES_ACTION]
        ).aggregate(avg=Avg("score"))["avg"] or 0

        # Total issues
        issue_stats = queryset.aggregate(
            total_issues=Sum("issues_found"),
            critical_issues=Sum("critical_issues"),
        )

        # Inspections by type
        by_type = queryset.values("inspection_type").annotate(count=Count("id"))
        inspections_by_type = {item["inspection_type"]: item["count"] for item in by_type}

        # Inspections by room (top 10 with most inspections)
        by_room = queryset.values("room__number").annotate(
            count=Count("id"),
            avg_score=Avg("score"),
        ).order_by("-count")[:10]
        inspections_by_room = list(by_room)

        data = {
            "total_inspections": total_inspections,
            "completed_inspections": completed_inspections,
            "pending_inspections": pending_inspections,
            "requires_action": requires_action,
            "average_score": round(avg_score, 2),
            "total_issues": issue_stats["total_issues"] or 0,
            "critical_issues": issue_stats["critical_issues"] or 0,
            "inspections_by_type": inspections_by_type,
            "inspections_by_room": inspections_by_room,
        }

        return Response(data)

    @extend_schema(
        summary="Create from checkout",
        description="Auto-create checkout inspection for a booking.",
        request={
            "type": "object",
            "properties": {
                "booking_id": {"type": "integer"},
                "template_id": {"type": "integer"},
            },
            "required": ["booking_id"],
        },
        responses={201: RoomInspectionSerializer},
        tags=["Room Inspection"],
    )
    @action(detail=False, methods=["post"], url_path="from-checkout")
    def from_checkout(self, request):
        """Create checkout inspection from a booking."""
        from django.utils import timezone

        booking_id = request.data.get("booking_id")
        template_id = request.data.get("template_id")

        if not booking_id:
            return Response(
                {"detail": "Vui lòng cung cấp booking_id."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        try:
            booking = Booking.objects.get(id=booking_id)
        except Booking.DoesNotExist:
            return Response(
                {"detail": "Đặt phòng không tồn tại."},
                status=status.HTTP_404_NOT_FOUND,
            )

        # Get template or use default
        checklist_items = []
        if template_id:
            try:
                template = InspectionTemplate.objects.get(id=template_id, is_active=True)
                checklist_items = [
                    {
                        "category": item.get("category", ""),
                        "item": item.get("item", ""),
                        "critical": item.get("critical", False),
                        "passed": None,
                        "notes": "",
                    }
                    for item in template.items
                ]
            except InspectionTemplate.DoesNotExist:
                pass
        else:
            # Try to find default template for room type
            template = InspectionTemplate.objects.filter(
                inspection_type=RoomInspection.InspectionType.CHECKOUT,
                is_default=True,
                is_active=True,
            ).filter(
                Q(room_type=booking.room.room_type) | Q(room_type__isnull=True)
            ).first()

            if template:
                checklist_items = [
                    {
                        "category": item.get("category", ""),
                        "item": item.get("item", ""),
                        "critical": item.get("critical", False),
                        "passed": None,
                        "notes": "",
                    }
                    for item in template.items
                ]

        # Create inspection
        inspection = RoomInspection.objects.create(
            room=booking.room,
            booking=booking,
            inspection_type=RoomInspection.InspectionType.CHECKOUT,
            scheduled_date=timezone.now().date(),
            checklist_items=checklist_items,
        )

        return Response(
            RoomInspectionSerializer(inspection).data,
            status=status.HTTP_201_CREATED,
        )

    @extend_schema(
        summary="Get pending inspections",
        description="Get all pending inspections for today.",
        responses={200: RoomInspectionListSerializer(many=True)},
        tags=["Room Inspection"],
    )
    @action(detail=False, methods=["get"], url_path="pending-today")
    def pending_today(self, request):
        """Get pending inspections for today."""
        from django.utils import timezone

        today = timezone.now().date()
        queryset = self.get_queryset().filter(
            scheduled_date=today,
            status=RoomInspection.Status.PENDING,
        )

        serializer = RoomInspectionListSerializer(queryset, many=True)
        return Response(serializer.data)


# ============================================================
# RatePlan ViewSet
# ============================================================


@extend_schema_view(
    list=extend_schema(
        summary="List rate plans",
        description="Get all rate plans, optionally filtered by room type.",
        parameters=[
            OpenApiParameter("room_type", OpenApiTypes.INT, description="Filter by room type ID"),
            OpenApiParameter("is_active", OpenApiTypes.BOOL, description="Filter by active status"),
        ],
        responses={200: RatePlanListSerializer(many=True)},
        tags=["Rate Plans"],
    ),
    retrieve=extend_schema(
        summary="Get rate plan details",
        description="Get detailed information about a specific rate plan.",
        responses={200: RatePlanSerializer},
        tags=["Rate Plans"],
    ),
    create=extend_schema(
        summary="Create rate plan",
        description="Create a new rate plan.",
        request=RatePlanCreateSerializer,
        responses={201: RatePlanSerializer},
        tags=["Rate Plans"],
    ),
    update=extend_schema(
        summary="Update rate plan",
        description="Update an existing rate plan.",
        request=RatePlanUpdateSerializer,
        responses={200: RatePlanSerializer},
        tags=["Rate Plans"],
    ),
    partial_update=extend_schema(
        summary="Partial update rate plan",
        description="Partially update a rate plan.",
        request=RatePlanUpdateSerializer,
        responses={200: RatePlanSerializer},
        tags=["Rate Plans"],
    ),
    destroy=extend_schema(
        summary="Delete rate plan",
        description="Delete a rate plan.",
        responses={204: None},
        tags=["Rate Plans"],
    ),
)
class RatePlanViewSet(viewsets.ModelViewSet):
    """ViewSet for managing RatePlans."""

    permission_classes = [IsAuthenticated, IsStaffOrManager]
    queryset = RatePlan.objects.select_related("room_type").all()

    def get_serializer_class(self):
        if self.action == "list":
            return RatePlanListSerializer
        elif self.action == "create":
            return RatePlanCreateSerializer
        elif self.action in ["update", "partial_update"]:
            return RatePlanUpdateSerializer
        return RatePlanSerializer

    def get_queryset(self):
        queryset = super().get_queryset()

        # Filter by room type
        room_type = self.request.query_params.get("room_type")
        if room_type:
            queryset = queryset.filter(room_type_id=room_type)

        # Filter by active status
        is_active = self.request.query_params.get("is_active")
        if is_active is not None:
            is_active_bool = is_active.lower() in ("true", "1", "yes")
            queryset = queryset.filter(is_active=is_active_bool)

        return queryset

    @extend_schema(
        summary="Get active rate plans for room type",
        description="Get all active rate plans for a specific room type, including currently valid ones.",
        parameters=[
            OpenApiParameter("room_type_id", OpenApiTypes.INT, description="Room type ID", required=True, location=OpenApiParameter.PATH),
        ],
        responses={200: RatePlanListSerializer(many=True)},
        tags=["Rate Plans"],
    )
    @action(detail=False, methods=["get"], url_path="by-room-type/(?P<room_type_id>[^/.]+)")
    def by_room_type(self, request, room_type_id=None):
        """Get active rate plans for a room type."""
        from django.utils import timezone

        today = timezone.now().date()
        queryset = self.get_queryset().filter(
            room_type_id=room_type_id,
            is_active=True,
        ).filter(
            Q(valid_from__isnull=True) | Q(valid_from__lte=today),
            Q(valid_to__isnull=True) | Q(valid_to__gte=today),
        )

        serializer = RatePlanListSerializer(queryset, many=True)
        return Response(serializer.data)


# ============================================================
# DateRateOverride ViewSet
# ============================================================


@extend_schema_view(
    list=extend_schema(
        summary="List date rate overrides",
        description="Get all date rate overrides, optionally filtered by room type and date range.",
        parameters=[
            OpenApiParameter("room_type", OpenApiTypes.INT, description="Filter by room type ID"),
            OpenApiParameter("start_date", OpenApiTypes.DATE, description="Filter from date"),
            OpenApiParameter("end_date", OpenApiTypes.DATE, description="Filter to date"),
        ],
        responses={200: DateRateOverrideListSerializer(many=True)},
        tags=["Date Rate Overrides"],
    ),
    retrieve=extend_schema(
        summary="Get date rate override details",
        description="Get detailed information about a specific date rate override.",
        responses={200: DateRateOverrideSerializer},
        tags=["Date Rate Overrides"],
    ),
    create=extend_schema(
        summary="Create date rate override",
        description="Create a new date rate override.",
        request=DateRateOverrideCreateSerializer,
        responses={201: DateRateOverrideSerializer},
        tags=["Date Rate Overrides"],
    ),
    update=extend_schema(
        summary="Update date rate override",
        description="Update an existing date rate override.",
        request=DateRateOverrideUpdateSerializer,
        responses={200: DateRateOverrideSerializer},
        tags=["Date Rate Overrides"],
    ),
    partial_update=extend_schema(
        summary="Partial update date rate override",
        description="Partially update a date rate override.",
        request=DateRateOverrideUpdateSerializer,
        responses={200: DateRateOverrideSerializer},
        tags=["Date Rate Overrides"],
    ),
    destroy=extend_schema(
        summary="Delete date rate override",
        description="Delete a date rate override.",
        responses={204: None},
        tags=["Date Rate Overrides"],
    ),
)
class DateRateOverrideViewSet(viewsets.ModelViewSet):
    """ViewSet for managing DateRateOverrides."""

    permission_classes = [IsAuthenticated, IsStaffOrManager]
    queryset = DateRateOverride.objects.select_related("room_type").all()

    def get_serializer_class(self):
        if self.action == "list":
            return DateRateOverrideListSerializer
        elif self.action == "create":
            return DateRateOverrideCreateSerializer
        elif self.action in ["update", "partial_update"]:
            return DateRateOverrideUpdateSerializer
        elif self.action == "bulk_create":
            return DateRateOverrideBulkCreateSerializer
        return DateRateOverrideSerializer

    def get_queryset(self):
        queryset = super().get_queryset()

        # Filter by room type
        room_type = self.request.query_params.get("room_type")
        if room_type:
            queryset = queryset.filter(room_type_id=room_type)

        # Filter by date range
        start_date = self.request.query_params.get("start_date")
        if start_date:
            queryset = queryset.filter(date__gte=start_date)

        end_date = self.request.query_params.get("end_date")
        if end_date:
            queryset = queryset.filter(date__lte=end_date)

        return queryset

    @extend_schema(
        summary="Bulk create date rate overrides",
        description="Create date rate overrides for a date range. Existing overrides for the same room type and dates will be updated.",
        request=DateRateOverrideBulkCreateSerializer,
        responses={201: DateRateOverrideSerializer(many=True)},
        tags=["Date Rate Overrides"],
    )
    @action(detail=False, methods=["post"], url_path="bulk-create")
    def bulk_create(self, request):
        """Bulk create date rate overrides for a date range."""
        from datetime import timedelta

        serializer = DateRateOverrideBulkCreateSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        data = serializer.validated_data
        room_type = data["room_type"]
        start_date = data["start_date"]
        end_date = data["end_date"]
        rate = data["rate"]
        reason = data.get("reason", "")
        closed_to_arrival = data.get("closed_to_arrival", False)
        closed_to_departure = data.get("closed_to_departure", False)
        min_stay = data.get("min_stay")

        created_overrides = []
        current_date = start_date
        while current_date <= end_date:
            override, created = DateRateOverride.objects.update_or_create(
                room_type=room_type,
                date=current_date,
                defaults={
                    "rate": rate,
                    "reason": reason,
                    "closed_to_arrival": closed_to_arrival,
                    "closed_to_departure": closed_to_departure,
                    "min_stay": min_stay,
                },
            )
            created_overrides.append(override)
            current_date += timedelta(days=1)

        return Response(
            DateRateOverrideSerializer(created_overrides, many=True).data,
            status=status.HTTP_201_CREATED,
        )

    @extend_schema(
        summary="Get overrides for room type",
        description="Get all date rate overrides for a specific room type in a date range.",
        parameters=[
            OpenApiParameter("room_type_id", OpenApiTypes.INT, description="Room type ID", required=True, location=OpenApiParameter.PATH),
            OpenApiParameter("start_date", OpenApiTypes.DATE, description="Start date", required=True),
            OpenApiParameter("end_date", OpenApiTypes.DATE, description="End date", required=True),
        ],
        responses={200: DateRateOverrideListSerializer(many=True)},
        tags=["Date Rate Overrides"],
    )
    @action(detail=False, methods=["get"], url_path="by-room-type/(?P<room_type_id>[^/.]+)")
    def by_room_type(self, request, room_type_id=None):
        """Get date rate overrides for a room type in a date range."""
        start_date = request.query_params.get("start_date")
        end_date = request.query_params.get("end_date")

        if not start_date or not end_date:
            return Response(
                {"detail": "start_date and end_date are required."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        queryset = self.get_queryset().filter(
            room_type_id=room_type_id,
            date__gte=start_date,
            date__lte=end_date,
        )

        serializer = DateRateOverrideListSerializer(queryset, many=True)
        return Response(serializer.data)


# ===== Phase 5: Notification Views =====


class NotificationViewSet(viewsets.ReadOnlyModelViewSet):
    """ViewSet for viewing and managing notifications."""

    permission_classes = [IsAuthenticated]
    serializer_class = NotificationSerializer

    def get_queryset(self):
        """Return notifications for the authenticated user."""
        return Notification.objects.filter(
            recipient=self.request.user
        ).order_by("-created_at")

    def get_serializer_class(self):
        if self.action == "list":
            return NotificationListSerializer
        return NotificationSerializer

    @extend_schema(
        summary="Mark notification as read",
        responses={200: NotificationSerializer},
        tags=["Notifications"],
    )
    @action(detail=True, methods=["post"], url_path="read")
    def mark_read(self, request, pk=None):
        """Mark a single notification as read."""
        notification = self.get_object()
        notification.mark_read()
        return Response(NotificationSerializer(notification).data)

    @extend_schema(
        summary="Mark all notifications as read",
        responses={200: OpenApiResponse(description="All notifications marked as read")},
        tags=["Notifications"],
    )
    @action(detail=False, methods=["post"], url_path="read-all")
    def mark_all_read(self, request):
        """Mark all unread notifications as read."""
        from django.utils import timezone

        count = Notification.objects.filter(
            recipient=request.user,
            is_read=False,
        ).update(is_read=True, read_at=timezone.now())

        return Response({"marked_read": count})

    @extend_schema(
        summary="Get unread notification count",
        responses={200: OpenApiResponse(description="Unread count")},
        tags=["Notifications"],
    )
    @action(detail=False, methods=["get"], url_path="unread-count")
    def unread_count(self, request):
        """Get count of unread notifications."""
        count = Notification.objects.filter(
            recipient=request.user,
            is_read=False,
        ).count()
        return Response({"unread_count": count})


class DeviceTokenView(APIView):
    """Register/unregister FCM device tokens."""

    permission_classes = [IsAuthenticated]

    @extend_schema(
        summary="Register device token",
        request=DeviceTokenSerializer,
        responses={201: DeviceTokenSerializer},
        tags=["Notifications"],
    )
    def post(self, request):
        """Register a device token."""
        serializer = DeviceTokenSerializer(
            data=request.data, context={"request": request}
        )
        serializer.is_valid(raise_exception=True)
        serializer.save()
        return Response(serializer.data, status=status.HTTP_201_CREATED)

    @extend_schema(
        summary="Unregister device token",
        responses={200: OpenApiResponse(description="Token deactivated")},
        tags=["Notifications"],
    )
    def delete(self, request):
        """Deactivate a device token."""
        token = request.data.get("token")
        if not token:
            return Response(
                {"detail": "Token is required."},
                status=status.HTTP_400_BAD_REQUEST,
            )
        DeviceToken.objects.filter(
            user=request.user, token=token
        ).update(is_active=False)
        return Response({"detail": "Token deactivated."})


class NotificationPreferencesView(APIView):
    """Get/update notification preferences."""

    permission_classes = [IsAuthenticated]

    @extend_schema(
        summary="Get notification preferences",
        responses={200: NotificationPreferencesSerializer},
        tags=["Notifications"],
    )
    def get(self, request):
        """Get user's notification preferences."""
        profile = request.user.hotel_profile
        return Response({
            "receive_notifications": profile.receive_notifications,
        })

    @extend_schema(
        summary="Update notification preferences",
        request=NotificationPreferencesSerializer,
        responses={200: NotificationPreferencesSerializer},
        tags=["Notifications"],
    )
    def put(self, request):
        """Update user's notification preferences."""
        serializer = NotificationPreferencesSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        profile = request.user.hotel_profile
        profile.receive_notifications = serializer.validated_data["receive_notifications"]
        profile.save(update_fields=["receive_notifications"])

        return Response({
            "receive_notifications": profile.receive_notifications,
        })


# ===== Phase 5.3: Guest Messaging Views =====


@extend_schema_view(
    list=extend_schema(summary="List message templates", tags=["Guest Messaging"]),
    create=extend_schema(summary="Create message template", tags=["Guest Messaging"]),
    retrieve=extend_schema(summary="Get message template", tags=["Guest Messaging"]),
    update=extend_schema(summary="Update message template", tags=["Guest Messaging"]),
    partial_update=extend_schema(summary="Partial update template", tags=["Guest Messaging"]),
    destroy=extend_schema(summary="Delete message template", tags=["Guest Messaging"]),
)
class MessageTemplateViewSet(viewsets.ModelViewSet):
    """CRUD for message templates."""

    permission_classes = [IsAuthenticated]
    serializer_class = MessageTemplateSerializer
    queryset = MessageTemplate.objects.all()

    def get_serializer_class(self):
        if self.action == "list":
            return MessageTemplateListSerializer
        return MessageTemplateSerializer

    def get_queryset(self):
        qs = super().get_queryset()
        # Filter by template_type
        template_type = self.request.query_params.get("template_type")
        if template_type:
            qs = qs.filter(template_type=template_type)
        # Filter by channel
        channel = self.request.query_params.get("channel")
        if channel:
            qs = qs.filter(channel=channel)
        # Filter by active status
        is_active = self.request.query_params.get("is_active")
        if is_active is not None:
            qs = qs.filter(is_active=is_active.lower() == "true")
        return qs

    @extend_schema(
        summary="Preview rendered template",
        request=PreviewMessageSerializer,
        responses={200: OpenApiResponse(description="Rendered message preview")},
        tags=["Guest Messaging"],
    )
    @action(detail=False, methods=["post"], url_path="preview")
    def preview(self, request):
        """Preview a rendered template with guest/booking context."""
        serializer = PreviewMessageSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        try:
            template = MessageTemplate.objects.get(pk=serializer.validated_data["template"])
        except MessageTemplate.DoesNotExist:
            return Response(
                {"detail": "Không tìm thấy mẫu tin nhắn."},
                status=status.HTTP_404_NOT_FOUND,
            )

        try:
            guest = Guest.objects.get(pk=serializer.validated_data["guest"])
        except Guest.DoesNotExist:
            return Response(
                {"detail": "Không tìm thấy khách."},
                status=status.HTTP_404_NOT_FOUND,
            )

        booking = None
        booking_id = serializer.validated_data.get("booking")
        if booking_id:
            try:
                booking = Booking.objects.select_related("room", "room__room_type").get(
                    pk=booking_id
                )
            except Booking.DoesNotExist:
                pass

        from .messaging_service import GuestMessagingService

        rendered_subject, rendered_body = GuestMessagingService.render_template(
            template, guest, booking
        )

        recipient = GuestMessagingService.get_recipient_address(guest, template.channel)

        return Response({
            "subject": rendered_subject,
            "body": rendered_body,
            "recipient_address": recipient,
            "channel": template.channel,
        })


@extend_schema_view(
    list=extend_schema(summary="List guest messages", tags=["Guest Messaging"]),
    retrieve=extend_schema(summary="Get guest message", tags=["Guest Messaging"]),
)
class GuestMessageViewSet(viewsets.ReadOnlyModelViewSet):
    """ViewSet for viewing guest messages and sending new ones."""

    permission_classes = [IsAuthenticated]
    serializer_class = GuestMessageSerializer

    def get_queryset(self):
        qs = GuestMessage.objects.select_related(
            "guest", "booking", "booking__room", "template", "sent_by"
        ).all()

        # Filter by guest
        guest_id = self.request.query_params.get("guest")
        if guest_id:
            qs = qs.filter(guest_id=guest_id)

        # Filter by booking
        booking_id = self.request.query_params.get("booking")
        if booking_id:
            qs = qs.filter(booking_id=booking_id)

        # Filter by channel
        channel = self.request.query_params.get("channel")
        if channel:
            qs = qs.filter(channel=channel)

        # Filter by status
        msg_status = self.request.query_params.get("status")
        if msg_status:
            qs = qs.filter(status=msg_status)

        return qs

    def get_serializer_class(self):
        if self.action == "list":
            return GuestMessageListSerializer
        return GuestMessageSerializer

    @extend_schema(
        summary="Send a message to a guest",
        request=SendMessageSerializer,
        responses={201: GuestMessageSerializer},
        tags=["Guest Messaging"],
    )
    @action(detail=False, methods=["post"], url_path="send")
    def send_message(self, request):
        """Send a new message to a guest."""
        serializer = SendMessageSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        data = serializer.validated_data

        # Validate guest
        try:
            guest = Guest.objects.get(pk=data["guest"])
        except Guest.DoesNotExist:
            return Response(
                {"detail": "Không tìm thấy khách."},
                status=status.HTTP_404_NOT_FOUND,
            )

        # Validate booking (optional)
        booking = None
        if data.get("booking"):
            try:
                booking = Booking.objects.get(pk=data["booking"])
            except Booking.DoesNotExist:
                return Response(
                    {"detail": "Không tìm thấy đặt phòng."},
                    status=status.HTTP_404_NOT_FOUND,
                )

        # Validate template (optional)
        template = None
        if data.get("template"):
            try:
                template = MessageTemplate.objects.get(pk=data["template"])
            except MessageTemplate.DoesNotExist:
                pass

        from .messaging_service import GuestMessagingService

        # Get recipient address
        recipient = GuestMessagingService.get_recipient_address(
            guest, data["channel"]
        )

        # Create message record
        message = GuestMessage.objects.create(
            guest=guest,
            booking=booking,
            template=template,
            channel=data["channel"],
            subject=data["subject"],
            body=data["body"],
            recipient_address=recipient,
            sent_by=request.user,
        )

        # Send the message
        GuestMessagingService.send_message(message)

        # Refresh from DB
        message.refresh_from_db()

        return Response(
            GuestMessageSerializer(message).data,
            status=status.HTTP_201_CREATED,
        )

    @extend_schema(
        summary="Resend a failed message",
        responses={200: GuestMessageSerializer},
        tags=["Guest Messaging"],
    )
    @action(detail=True, methods=["post"], url_path="resend")
    def resend(self, request, pk=None):
        """Resend a failed message."""
        message = self.get_object()
        if message.status not in ("failed", "draft"):
            return Response(
                {"detail": "Chỉ có thể gửi lại tin nhắn thất bại hoặc nháp."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        from .messaging_service import GuestMessagingService

        GuestMessagingService.send_message(message)
        message.refresh_from_db()

        return Response(GuestMessageSerializer(message).data)
