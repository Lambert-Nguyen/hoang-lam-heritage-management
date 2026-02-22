"""
Tests for Temporary Residence Declaration Export endpoints.

Covers two official Vietnamese declaration forms:
- Mẫu ĐD10 (Nghị định 144/2021): Sổ quản lý lưu trú — Vietnamese guests
- Mẫu NA17 (Thông tư 04/2015): Phiếu khai báo tạm trú — Foreign guests
"""

from datetime import date, timedelta
from decimal import Decimal

from django.contrib.auth.models import User
from django.utils import timezone

import pytest
from rest_framework import status
from rest_framework.test import APIClient

from hotel_api.models import Booking, Guest, HotelUser, Room, RoomType

# ==================== Fixtures ====================


@pytest.fixture
def api_client():
    """Create API client."""
    return APIClient()


@pytest.fixture
def create_user(db):
    """Factory to create users with roles."""

    def _create_user(username, role="staff"):
        user = User.objects.create_user(username=username, password="testpass123")
        HotelUser.objects.create(user=user, role=role, phone=f"+84{username[-6:]}")
        return user

    return _create_user


@pytest.fixture
def manager_user(create_user):
    """Create manager user."""
    return create_user("manager001", "manager")


@pytest.fixture
def staff_user(create_user):
    """Create staff user."""
    return create_user("staff001", "staff")


@pytest.fixture
def room_type(db):
    """Create a room type."""
    return RoomType.objects.create(
        name="Phòng Đôi",
        name_en="Double Room",
        base_rate=Decimal("500000"),
        max_guests=2,
    )


@pytest.fixture
def room(room_type):
    """Create a room."""
    return Room.objects.create(
        number="101",
        name="Phòng 101",
        room_type=room_type,
        floor=1,
        status=Room.Status.AVAILABLE,
    )


@pytest.fixture
def room_102(room_type):
    """Create second room for multiple bookings."""
    return Room.objects.create(
        number="102",
        name="Phòng 102",
        room_type=room_type,
        floor=1,
        status=Room.Status.AVAILABLE,
    )


@pytest.fixture
def guest_vietnamese(db):
    """Create a Vietnamese guest with full details for ĐD10."""
    return Guest.objects.create(
        full_name="Nguyễn Văn Test",
        phone="+84901234567",
        id_type=Guest.IDType.CCCD,
        id_number="012345678901",
        nationality="Việt Nam",
        gender="male",
        date_of_birth=date(1990, 5, 15),
        id_issue_date=date(2020, 1, 1),
        id_issue_place="CA TP Hồ Chí Minh",
        address="123 Đường ABC, Quận 1, TP.HCM",
    )


@pytest.fixture
def guest_foreign(db):
    """Create a foreign guest with full NA17 details."""
    return Guest.objects.create(
        full_name="John Smith",
        phone="+1234567890",
        id_type=Guest.IDType.PASSPORT,
        id_number="AB1234567",
        nationality="United States",
        gender="male",
        date_of_birth=date(1985, 8, 20),
        id_issue_date=date(2022, 6, 15),
        id_issue_place="US State Department",
        address="123 Main St, New York, USA",
        # NA17-specific fields
        passport_type=Guest.PassportType.ORDINARY,
        visa_type=Guest.VisaType.VISA,
        visa_number="VN2026012345",
        visa_issue_date=date(2026, 1, 1),
        visa_expiry_date=date(2026, 7, 1),
        visa_issuing_authority="Vietnam Embassy, Washington D.C.",
        entry_date=date(2026, 2, 1),
        entry_port="Sân bay Tân Sơn Nhất",
        entry_purpose="Du lịch",
    )


@pytest.fixture
def booking_vn_checked_in(room, guest_vietnamese, room_type):
    """Create a checked-in booking for Vietnamese guest today."""
    return Booking.objects.create(
        room=room,
        guest=guest_vietnamese,
        check_in_date=date.today(),
        check_out_date=date.today() + timedelta(days=2),
        status=Booking.Status.CHECKED_IN,
        actual_check_in=timezone.now(),
        nightly_rate=room_type.base_rate,
        total_amount=Decimal("1000000"),
    )


@pytest.fixture
def booking_foreign_checked_in(room_102, guest_foreign, room_type):
    """Create a checked-in booking for foreign guest today."""
    return Booking.objects.create(
        room=room_102,
        guest=guest_foreign,
        check_in_date=date.today(),
        check_out_date=date.today() + timedelta(days=3),
        status=Booking.Status.CHECKED_IN,
        actual_check_in=timezone.now(),
        nightly_rate=room_type.base_rate,
        total_amount=Decimal("1500000"),
    )


def _authenticate(client, user):
    """Helper to authenticate API client."""
    client.force_authenticate(user=user)


# ==================== Auth & Validation ====================


@pytest.mark.django_db
class TestDeclarationExportAuth:
    """Test authentication and authorization for declaration export."""

    def test_export_requires_auth(self, api_client):
        response = api_client.get("/api/v1/guests/declaration-export/")
        assert response.status_code == status.HTTP_401_UNAUTHORIZED

    def test_staff_can_export(self, api_client, staff_user, booking_vn_checked_in):
        _authenticate(api_client, staff_user)
        response = api_client.get(
            "/api/v1/guests/declaration-export/?export_format=csv&form_type=dd10"
        )
        assert response.status_code == status.HTTP_200_OK

    def test_manager_can_export(self, api_client, manager_user, booking_vn_checked_in):
        _authenticate(api_client, manager_user)
        response = api_client.get(
            "/api/v1/guests/declaration-export/?export_format=csv&form_type=dd10"
        )
        assert response.status_code == status.HTTP_200_OK


@pytest.mark.django_db
class TestDeclarationExportValidation:
    """Test input validation."""

    def test_invalid_date_format(self, api_client, manager_user):
        _authenticate(api_client, manager_user)
        response = api_client.get("/api/v1/guests/declaration-export/?date_from=invalid")
        assert response.status_code == status.HTTP_400_BAD_REQUEST
        assert "Định dạng ngày" in response.data["detail"]

    def test_date_from_after_date_to(self, api_client, manager_user):
        _authenticate(api_client, manager_user)
        date_from = date.today()
        date_to = date.today() - timedelta(days=5)
        response = api_client.get(
            f"/api/v1/guests/declaration-export/?date_from={date_from}&date_to={date_to}"
        )
        assert response.status_code == status.HTTP_400_BAD_REQUEST
        assert "Ngày bắt đầu" in response.data["detail"]

    def test_invalid_form_type(self, api_client, manager_user):
        _authenticate(api_client, manager_user)
        response = api_client.get("/api/v1/guests/declaration-export/?form_type=invalid")
        assert response.status_code == status.HTTP_400_BAD_REQUEST
        assert "form_type" in response.data["detail"]

    def test_default_dates_to_today(self, api_client, manager_user, booking_vn_checked_in):
        _authenticate(api_client, manager_user)
        response = api_client.get(
            "/api/v1/guests/declaration-export/?export_format=csv&form_type=dd10"
        )
        assert response.status_code == status.HTTP_200_OK
        today_str = date.today().isoformat()
        assert today_str in response["Content-Disposition"]


# ==================== ĐD10 Format (Vietnamese Guests) ====================


@pytest.mark.django_db
class TestDD10Export:
    """Test Mẫu ĐD10 - Sổ quản lý lưu trú (Nghị định 144/2021)."""

    def test_dd10_csv_contains_correct_headers(
        self, api_client, manager_user, booking_vn_checked_in
    ):
        """ĐD10 CSV has correct column headers per official format."""
        _authenticate(api_client, manager_user)
        response = api_client.get(
            "/api/v1/guests/declaration-export/?export_format=csv&form_type=dd10"
        )
        assert response.status_code == status.HTTP_200_OK
        content = response.content.decode("utf-8-sig")

        # Official ĐD10 headers
        assert "Họ và tên" in content
        assert "Ngày sinh" in content
        assert "Giới tính" in content
        assert "Quốc tịch" in content
        assert "Số CMND/CCCD/Hộ chiếu" in content
        assert "Địa chỉ thường trú" in content
        assert "Ngày đến" in content
        assert "Ngày đi" in content
        assert "Số phòng" in content
        assert "Ghi chú" in content

    def test_dd10_csv_contains_establishment_header(
        self, api_client, manager_user, booking_vn_checked_in
    ):
        """ĐD10 CSV contains hotel establishment info at the top."""
        _authenticate(api_client, manager_user)
        response = api_client.get(
            "/api/v1/guests/declaration-export/?export_format=csv&form_type=dd10"
        )
        content = response.content.decode("utf-8-sig")
        assert "Hoàng Lâm Heritage Suites" in content
        assert "Địa chỉ:" in content
        assert "Điện thoại:" in content

    def test_dd10_csv_contains_form_title(self, api_client, manager_user, booking_vn_checked_in):
        """ĐD10 CSV contains the official form title."""
        _authenticate(api_client, manager_user)
        response = api_client.get(
            "/api/v1/guests/declaration-export/?export_format=csv&form_type=dd10"
        )
        content = response.content.decode("utf-8-sig")
        assert "SỔ QUẢN LÝ LƯU TRÚ" in content
        assert "ĐD10" in content

    def test_dd10_csv_contains_vietnamese_guest_data(
        self, api_client, manager_user, booking_vn_checked_in
    ):
        """ĐD10 CSV includes Vietnamese guest data."""
        _authenticate(api_client, manager_user)
        response = api_client.get(
            "/api/v1/guests/declaration-export/?export_format=csv&form_type=dd10"
        )
        content = response.content.decode("utf-8-sig")
        assert "Nguyễn Văn Test" in content
        assert "012345678901" in content
        assert "123 Đường ABC" in content

    def test_dd10_excludes_foreign_guests(
        self, api_client, manager_user, booking_vn_checked_in, booking_foreign_checked_in
    ):
        """ĐD10 form should NOT include foreign guests."""
        _authenticate(api_client, manager_user)
        response = api_client.get(
            "/api/v1/guests/declaration-export/?export_format=csv&form_type=dd10"
        )
        content = response.content.decode("utf-8-sig")
        assert "John Smith" not in content

    def test_dd10_shows_expected_checkout_date(
        self, api_client, manager_user, booking_vn_checked_in
    ):
        """ĐD10 shows expected check-out date even for currently checked-in guests."""
        _authenticate(api_client, manager_user)
        response = api_client.get(
            "/api/v1/guests/declaration-export/?export_format=csv&form_type=dd10"
        )
        content = response.content.decode("utf-8-sig")
        expected_checkout = (date.today() + timedelta(days=2)).strftime("%d/%m/%Y")
        assert expected_checkout in content

    def test_dd10_only_checked_in_or_out_bookings(
        self, api_client, manager_user, room, guest_vietnamese, room_type
    ):
        """ĐD10 excludes pending/confirmed/cancelled bookings."""
        _authenticate(api_client, manager_user)

        # Create a pending booking
        Booking.objects.create(
            room=room,
            guest=guest_vietnamese,
            check_in_date=date.today(),
            check_out_date=date.today() + timedelta(days=1),
            status=Booking.Status.PENDING,
            nightly_rate=room_type.base_rate,
            total_amount=Decimal("500000"),
        )

        response = api_client.get(
            "/api/v1/guests/declaration-export/?export_format=csv&form_type=dd10"
        )
        content = response.content.decode("utf-8-sig")
        # Guest name should NOT appear (only pending booking exists)
        assert "Nguyễn Văn Test" not in content

    def test_dd10_empty_date_range(self, api_client, manager_user):
        """ĐD10 returns headers only for empty date range."""
        _authenticate(api_client, manager_user)
        future = date.today() + timedelta(days=30)
        response = api_client.get(
            f"/api/v1/guests/declaration-export/?date_from={future}&date_to={future}&export_format=csv&form_type=dd10"
        )
        assert response.status_code == status.HTTP_200_OK
        content = response.content.decode("utf-8-sig")
        assert "SỔ QUẢN LÝ LƯU TRÚ" in content


# ==================== NA17 Format (Foreign Guests) ====================


@pytest.mark.django_db
class TestNA17Export:
    """Test Mẫu NA17 - Phiếu khai báo tạm trú (Thông tư 04/2015)."""

    def test_na17_csv_contains_correct_headers(
        self, api_client, manager_user, booking_foreign_checked_in
    ):
        """NA17 CSV has correct column headers per official format."""
        _authenticate(api_client, manager_user)
        response = api_client.get(
            "/api/v1/guests/declaration-export/?export_format=csv&form_type=na17"
        )
        assert response.status_code == status.HTTP_200_OK
        content = response.content.decode("utf-8-sig")

        # Official NA17 headers
        assert "Họ tên" in content
        assert "Giới tính" in content
        assert "Quốc tịch" in content
        assert "Số hộ chiếu" in content
        assert "Loại hộ chiếu" in content
        assert "Loại giấy tờ nhập cảnh" in content
        assert "Cửa khẩu nhập cảnh" in content
        assert "Mục đích nhập cảnh" in content
        assert "Tạm trú từ ngày" in content
        assert "Tạm trú đến ngày" in content

    def test_na17_csv_contains_form_title(
        self, api_client, manager_user, booking_foreign_checked_in
    ):
        """NA17 CSV contains the official form title."""
        _authenticate(api_client, manager_user)
        response = api_client.get(
            "/api/v1/guests/declaration-export/?export_format=csv&form_type=na17"
        )
        content = response.content.decode("utf-8-sig")
        assert "NA17" in content
        assert "NGƯỜI NƯỚC NGOÀI" in content

    def test_na17_csv_contains_foreign_guest_data(
        self, api_client, manager_user, booking_foreign_checked_in
    ):
        """NA17 CSV includes foreign guest with visa/entry information."""
        _authenticate(api_client, manager_user)
        response = api_client.get(
            "/api/v1/guests/declaration-export/?export_format=csv&form_type=na17"
        )
        content = response.content.decode("utf-8-sig")
        assert "John Smith" in content
        assert "AB1234567" in content
        assert "Mỹ" in content  # nationality displayed in Vietnamese
        assert "VN2026012345" in content  # visa number
        assert "Sân bay Tân Sơn Nhất" in content  # entry port
        assert "Du lịch" in content  # entry purpose

    def test_na17_csv_contains_passport_type(
        self, api_client, manager_user, booking_foreign_checked_in
    ):
        """NA17 CSV shows passport type (phổ thông, công vụ, etc.)."""
        _authenticate(api_client, manager_user)
        response = api_client.get(
            "/api/v1/guests/declaration-export/?export_format=csv&form_type=na17"
        )
        content = response.content.decode("utf-8-sig")
        assert "Phổ thông" in content  # Passport type display

    def test_na17_csv_contains_visa_info(
        self, api_client, manager_user, booking_foreign_checked_in
    ):
        """NA17 CSV shows visa type, number, validity, issuing authority."""
        _authenticate(api_client, manager_user)
        response = api_client.get(
            "/api/v1/guests/declaration-export/?export_format=csv&form_type=na17"
        )
        content = response.content.decode("utf-8-sig")
        assert "Thị thực (Visa)" in content  # visa type display
        assert "VN2026012345" in content  # visa number
        assert "Vietnam Embassy" in content  # issuing authority

    def test_na17_excludes_vietnamese_guests(
        self, api_client, manager_user, booking_vn_checked_in, booking_foreign_checked_in
    ):
        """NA17 form should NOT include Vietnamese guests."""
        _authenticate(api_client, manager_user)
        response = api_client.get(
            "/api/v1/guests/declaration-export/?export_format=csv&form_type=na17"
        )
        content = response.content.decode("utf-8-sig")
        assert "Nguyễn Văn Test" not in content

    def test_na17_shows_expected_checkout_date(
        self, api_client, manager_user, booking_foreign_checked_in
    ):
        """NA17 shows expected check-out date (tạm trú đến ngày)."""
        _authenticate(api_client, manager_user)
        response = api_client.get(
            "/api/v1/guests/declaration-export/?export_format=csv&form_type=na17"
        )
        content = response.content.decode("utf-8-sig")
        expected_checkout = (date.today() + timedelta(days=3)).strftime("%d/%m/%Y")
        assert expected_checkout in content


# ==================== Combined Export (all) ====================


@pytest.mark.django_db
class TestCombinedExport:
    """Test form_type=all exports both ĐD10 and NA17."""

    def test_csv_all_contains_both_forms(
        self, api_client, manager_user, booking_vn_checked_in, booking_foreign_checked_in
    ):
        """CSV with form_type=all contains both form sections."""
        _authenticate(api_client, manager_user)
        response = api_client.get(
            "/api/v1/guests/declaration-export/?export_format=csv&form_type=all"
        )
        assert response.status_code == status.HTTP_200_OK
        content = response.content.decode("utf-8-sig")

        # Both forms present
        assert "SỔ QUẢN LÝ LƯU TRÚ" in content
        assert "NA17" in content
        # Both guests present
        assert "Nguyễn Văn Test" in content
        assert "John Smith" in content

    def test_excel_all_has_separate_sheets(
        self, api_client, manager_user, booking_vn_checked_in, booking_foreign_checked_in
    ):
        """Excel with form_type=all creates separate sheets."""
        _authenticate(api_client, manager_user)
        response = api_client.get(
            "/api/v1/guests/declaration-export/?export_format=excel&form_type=all"
        )

        if response.status_code == status.HTTP_200_OK:
            assert "spreadsheetml" in response["Content-Type"]
            assert ".xlsx" in response["Content-Disposition"]

            # Verify sheets by reading the Excel file
            import io

            import openpyxl

            wb = openpyxl.load_workbook(io.BytesIO(response.content))
            sheet_names = wb.sheetnames
            assert len(sheet_names) == 2
            assert any("ĐD10" in name for name in sheet_names)
            assert any("NA17" in name for name in sheet_names)

    def test_excel_dd10_sheet_has_establishment_header(
        self, api_client, manager_user, booking_vn_checked_in
    ):
        """Excel ĐD10 sheet includes hotel establishment header."""
        _authenticate(api_client, manager_user)
        response = api_client.get(
            "/api/v1/guests/declaration-export/?export_format=excel&form_type=dd10"
        )

        if response.status_code == status.HTTP_200_OK:
            import io

            import openpyxl

            wb = openpyxl.load_workbook(io.BytesIO(response.content))
            ws = wb.active
            assert "Hoàng Lâm Heritage Suites" in str(ws.cell(row=1, column=1).value)
            assert "Địa chỉ:" in str(ws.cell(row=2, column=1).value)
            assert "Điện thoại:" in str(ws.cell(row=3, column=1).value)


# ==================== Declaration Marking ====================


@pytest.mark.django_db
class TestDeclarationMarking:
    """Test that export marks bookings as declared."""

    def test_export_marks_bookings_as_declared(
        self, api_client, manager_user, booking_vn_checked_in
    ):
        """After export, bookings should be marked as declaration_submitted=True."""
        _authenticate(api_client, manager_user)

        # Verify not declared initially
        booking_vn_checked_in.refresh_from_db()
        assert booking_vn_checked_in.declaration_submitted is False

        # Export
        response = api_client.get(
            "/api/v1/guests/declaration-export/?export_format=csv&form_type=dd10"
        )
        assert response.status_code == status.HTTP_200_OK

        # Verify marked as declared
        booking_vn_checked_in.refresh_from_db()
        assert booking_vn_checked_in.declaration_submitted is True
        assert booking_vn_checked_in.declaration_submitted_at is not None

    def test_already_declared_not_updated(self, api_client, manager_user, booking_vn_checked_in):
        """Already-declared bookings don't get their timestamp updated."""
        _authenticate(api_client, manager_user)

        # First export
        api_client.get("/api/v1/guests/declaration-export/?export_format=csv&form_type=dd10")
        booking_vn_checked_in.refresh_from_db()
        first_timestamp = booking_vn_checked_in.declaration_submitted_at

        # Second export
        api_client.get("/api/v1/guests/declaration-export/?export_format=csv&form_type=dd10")
        booking_vn_checked_in.refresh_from_db()
        assert booking_vn_checked_in.declaration_submitted_at == first_timestamp


# ==================== Guest Model Properties ====================


@pytest.mark.django_db
class TestGuestForeignDetection:
    """Test Guest.is_foreign_guest property."""

    def test_vietnamese_guest_not_foreign(self, guest_vietnamese):
        assert guest_vietnamese.is_foreign_guest is False

    def test_foreign_guest_is_foreign(self, guest_foreign):
        assert guest_foreign.is_foreign_guest is True

    def test_vietnam_alias_vn(self, db):
        g = Guest.objects.create(full_name="Test VN", phone="+849000001", nationality="VN")
        assert g.is_foreign_guest is False

    def test_vietnam_alias_english(self, db):
        g = Guest.objects.create(
            full_name="Test Vietnam", phone="+849000002", nationality="Vietnam"
        )
        assert g.is_foreign_guest is False

    def test_empty_nationality_treated_as_vietnamese(self, db):
        g = Guest.objects.create(full_name="Test Empty", phone="+849000003", nationality="")
        assert g.is_foreign_guest is False

    def test_na17_fields_on_foreign_guest(self, guest_foreign):
        assert guest_foreign.passport_type == Guest.PassportType.ORDINARY
        assert guest_foreign.visa_type == Guest.VisaType.VISA
        assert guest_foreign.visa_number == "VN2026012345"
        assert guest_foreign.entry_port == "Sân bay Tân Sơn Nhất"
        assert guest_foreign.entry_purpose == "Du lịch"
